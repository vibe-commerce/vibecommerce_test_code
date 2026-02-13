"""
ABCDX-анализ товаров на основе данных из generate_ads_data.py.

Использует те же seed(42) и логику генерации для точного воспроизведения данных,
добавляет расчёт маржи и классификацию ABCDX.

Результат: новый лист «ABCDX» в файле ads_data_v2.0.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import random

random.seed(42)  # СОВПАДАЕТ с generate_ads_data.py

# ============================================================
# 1. PRODUCT CATALOG (из generate_ads_data.py)
# ============================================================
products = [
    # (id, name, category, cost, price_wb, price_ozon, weight_kg, base_vol_wb, base_vol_ozon)
    (1,  "Подушка бамбук 50×70",            "Подушки",  350,  990,  1050, 0.8,  45, 25),
    (2,  "Подушка бамбук 70×70",            "Подушки",  400,  1190, 1250, 1.0,  35, 20),
    (3,  "Подушка лебяжий пух 50×70",       "Подушки",  300,  890,  940,  0.7,  55, 30),
    (4,  "Подушка лебяжий пух 70×70",       "Подушки",  350,  1090, 1140, 0.9,  40, 22),
    (5,  "Подушка холлофайбер 50×70",       "Подушки",  200,  590,  620,  0.6,  70, 40),
    (6,  "Подушка холлофайбер 70×70",       "Подушки",  250,  690,  730,  0.8,  50, 28),
    (7,  "Подушка ортопед. Memory Foam",    "Подушки",  800,  2490, 2590, 1.2,  20, 15),
    (8,  "Подушка ортопед. с гелем",        "Подушки",  900,  2990, 3090, 1.3,  12, 10),
    (9,  "Подушка для беременных U-обр.",   "Подушки",  600,  1990, 2090, 2.5,  15, 12),
    (10, "Подушка-валик для шеи",           "Подушки",  250,  790,  840,  0.4,  30, 18),
    (11, "Одеяло бамбук 1.5-сп",           "Одеяла",   500,  1490, 1550, 1.8,  35, 20),
    (12, "Одеяло бамбук 2-сп",             "Одеяла",   650,  1890, 1990, 2.2,  25, 15),
    (13, "Одеяло бамбук Евро",             "Одеяла",   750,  2190, 2290, 2.5,  20, 12),
    (14, "Одеяло лебяжий пух 1.5-сп",      "Одеяла",   400,  1290, 1350, 1.5,  40, 22),
    (15, "Одеяло лебяжий пух 2-сп",        "Одеяла",   550,  1590, 1690, 1.9,  28, 16),
    (16, "Одеяло лебяжий пух Евро",        "Одеяла",   650,  1890, 1990, 2.3,  18, 10),
    (17, "Одеяло шерстяное 1.5-сп",        "Одеяла",   700,  2490, 2590, 2.0,  12,  8),
    (18, "Одеяло облегч. летнее 1.5-сп",   "Одеяла",   350,  990,  1050, 1.0,   5,  3),
    (19, "КПБ сатин 1.5-сп",              "Постельное бельё", 600, 1790, 1890, 1.5, 30, 18),
    (20, "КПБ сатин 2-сп",                "Постельное бельё", 700, 2190, 2290, 1.8, 22, 14),
    (21, "КПБ сатин Евро",                "Постельное бельё", 800, 2490, 2590, 2.0, 18, 11),
    (22, "КПБ поплин 1.5-сп",             "Постельное бельё", 400, 1190, 1250, 1.3, 40, 25),
    (23, "КПБ поплин 2-сп",               "Постельное бельё", 500, 1490, 1550, 1.6, 30, 18),
    (24, "КПБ поплин Евро",               "Постельное бельё", 600, 1690, 1790, 1.8, 22, 14),
    (25, "КПБ бязь 1.5-сп",               "Постельное бельё", 300,  890,  940, 1.2, 55, 35),
    (26, "КПБ бязь 2-сп",                 "Постельное бельё", 380, 1090, 1150, 1.4, 40, 25),
    (27, "Простыня на рез. сатин 160×200", "Постельное бельё", 300,  890,  940, 0.5, 35, 20),
    (28, "Простыня на рез. поплин 160×200","Постельное бельё", 200,  590,  640, 0.4, 50, 30),
    (29, "Наматрасник стёганый 160×200",      "Наматрасники", 400, 1290, 1350, 1.5, 20, 12),
    (30, "Наматрасник стёганый 180×200",      "Наматрасники", 450, 1490, 1550, 1.7, 15, 10),
    (31, "Наматрасник водонепрон. 160×200",   "Наматрасники", 500, 1590, 1690, 1.6, 18, 14),
    (32, "Наматрасник водонепрон. 180×200",   "Наматрасники", 550, 1790, 1890, 1.8, 12,  9),
    (33, "Наматрасник с бортами 160×200",     "Наматрасники", 350, 1090, 1150, 1.4, 15,  8),
    (34, "Наматрасник с бортами 180×200",     "Наматрасники", 400, 1290, 1350, 1.6, 10,  6),
    (35, "Плед микрофибра 150×200",   "Пледы", 300,  990, 1050, 1.0, 40, 22),
    (36, "Плед микрофибра 200×220",   "Пледы", 400, 1390, 1450, 1.4, 25, 15),
    (37, "Плед шерпа 150×200",        "Пледы", 450, 1590, 1690, 1.3, 30, 18),
    (38, "Плед велсофт 150×200",      "Пледы", 250,  790,  840, 0.9, 35, 20),
    (39, "Плед с рукавами",           "Пледы", 350, 1190, 1250, 1.2, 20, 15),
    (40, "Маска для сна шёлковая",     "Аксессуары", 150,  590,  640, 0.05, 60, 35),
    (41, "Маска для сна 3D",           "Аксессуары", 100,  390,  420, 0.04, 80, 50),
    (42, "Маска для сна с охлаждением","Аксессуары", 200,  690,  740, 0.08, 30, 18),
    (43, "Беруши силиконовые (набор)", "Аксессуары",  80,  290,  320, 0.02,100, 60),
    (44, "Спрей для подушки лаванда",  "Аксессуары", 120,  490,  540, 0.15, 45, 28),
    (45, "Ночник-проектор звёзд",      "Аксессуары", 350, 1290, 1350, 0.30, 25, 20),
    (46, "Подушка детская 40×60",  "Детский сон", 200,  690,  740, 0.4, 25, 15),
    (47, "Одеяло детское 110×140", "Детский сон", 300,  990, 1050, 1.0, 20, 12),
    (48, "КПБ детский поплин",     "Детский сон", 350, 1090, 1150, 0.8, 18, 12),
    (49, "Бортики в кроватку",     "Детский сон", 400, 1390, 1450, 1.5, 12,  8),
    (50, "Ночник детский LED",     "Детский сон", 250,  890,  940, 0.2, 20, 15),
]

product_map = {p[0]: p for p in products}

# ============================================================
# 2. WEEKS
# ============================================================
weeks = [
    datetime(2025, 11, 17), datetime(2025, 11, 24),
    datetime(2025, 12,  1), datetime(2025, 12,  8),
    datetime(2025, 12, 15), datetime(2025, 12, 22),
    datetime(2025, 12, 29), datetime(2026,  1,  5),
    datetime(2026,  1, 12), datetime(2026,  1, 19),
    datetime(2026,  1, 26), datetime(2026,  2,  2),
]

# Последний месяц = последние 4 недели (W9-W12: 12 янв — 2 фев 2026)
LAST_MONTH_WEEKS = set(range(8, 12))  # индексы 8,9,10,11
# Новинки = первое появление за последние 3 месяца (≈12 недель)
# Все данные = 12 недель, так что новинка = SKU впервые в W4+ (≈ после 8 дек)
NOVELTY_THRESHOLD_WEEK = 4  # появился с W5 или позже → новинка

# ============================================================
# 3. ПАРАМЕТРЫ РАСЧЁТА МАРЖИ
# ============================================================
season_overall = [0.90, 1.25, 1.00, 1.05, 1.15, 1.45, 1.15, 0.55, 0.65, 0.78, 0.85, 0.90]
cat_season = {
    "Подушки":          [1.00, 1.10, 1.00, 1.00, 1.05, 1.15, 1.00, 0.85, 0.90, 1.00, 1.00, 1.00],
    "Одеяла":           [1.10, 1.10, 1.10, 1.15, 1.20, 1.30, 1.10, 0.70, 0.80, 0.90, 0.95, 1.00],
    "Постельное бельё": [1.00, 1.05, 1.00, 1.00, 1.00, 1.15, 1.00, 0.90, 0.95, 1.00, 1.00, 1.00],
    "Наматрасники":     [1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 0.90, 0.95, 1.00, 1.00, 1.00],
    "Пледы":            [1.10, 1.15, 1.10, 1.15, 1.20, 1.40, 1.10, 0.60, 0.70, 0.80, 0.85, 0.90],
    "Аксессуары":       [0.90, 1.00, 1.00, 1.00, 1.10, 1.50, 1.15, 0.70, 0.80, 0.90, 0.95, 1.00],
    "Детский сон":      [1.00, 1.00, 1.00, 1.00, 1.05, 1.30, 1.05, 0.80, 0.85, 0.90, 0.95, 1.00],
}

commission_rates = {
    "WB":  {"Подушки": 0.15, "Одеяла": 0.15, "Постельное бельё": 0.15,
            "Наматрасники": 0.15, "Пледы": 0.15, "Аксессуары": 0.18, "Детский сон": 0.15},
    "Ozon": {"Подушки": 0.12, "Одеяла": 0.12, "Постельное бельё": 0.12,
             "Наматрасники": 0.12, "Пледы": 0.12, "Аксессуары": 0.15, "Детский сон": 0.12},
}

drr_base = {
    "Подушки": 0.08, "Одеяла": 0.07, "Постельное бельё": 0.09,
    "Наматрасники": 0.06, "Пледы": 0.07, "Аксессуары": 0.10, "Детский сон": 0.07,
}
drr_week_mult = [1.0, 1.3, 1.0, 1.0, 1.1, 1.2, 0.9, 0.5, 0.7, 0.8, 0.9, 1.0]

oos_rules = {(5, "WB", 8): 0, (5, "WB", 9): 0}
discount_rules = {}
for p in products:
    if p[3] and p[0] in range(11, 19):
        for wi in range(7, 12):
            discount_rules[(p[0], "WB", wi)] = 0.90

trend_rules = {7: 1.05}
promo_boost = {(37, 4): 1.8, (37, 5): 2.0}


def logistics_per_unit(weight_kg, platform):
    if platform == "WB":
        return round(55 + weight_kg * 12)
    else:
        return round(45 + weight_kg * 14)


def gen_qty(base, overall_m, cat_m, noise=0.20):
    val = base * overall_m * cat_m
    val *= random.uniform(1 - noise, 1 + noise)
    return max(0, round(val))


# ============================================================
# 4. ВОСПРОИЗВЕДЕНИЕ ДАННЫХ (идентично generate_ads_data.py)
# ============================================================
sales_rows = []

for wi, week_date in enumerate(weeks):
    for p in products:
        pid, name, cat, cost, price_wb, price_ozon, weight, base_wb, base_ozon = p
        for platform in ["WB", "Ozon"]:
            base_vol = base_wb if platform == "WB" else base_ozon
            price = price_wb if platform == "WB" else price_ozon

            overall_m = season_overall[wi]
            cat_m = cat_season[cat][wi]

            trend_m = 1.0
            if pid in trend_rules:
                trend_m = trend_rules[pid] ** wi

            if (pid, platform, wi) in oos_rules:
                qty = 0
            else:
                qty = gen_qty(base_vol, overall_m, cat_m, noise=0.20)
                qty = round(qty * trend_m)

            if pid == 1 and wi % 2 == 0:
                qty = round(qty * 1.12)
            elif pid == 1 and wi % 2 == 1:
                qty = round(qty * 0.88)
            elif pid == 3 and wi % 2 == 0:
                qty = round(qty * 0.88)
            elif pid == 3 and wi % 2 == 1:
                qty = round(qty * 1.12)

            if (pid, platform, wi) in discount_rules:
                price = round(price * discount_rules[(pid, platform, wi)])

            promo_m = 1.0
            if (pid, wi) in promo_boost:
                promo_m = promo_boost[(pid, wi)]
                qty = round(qty * promo_m)

            if pid == 18:
                qty = max(0, round(qty * 0.3))

            comm_rate = commission_rates[platform][cat]
            logi = logistics_per_unit(weight, platform) * qty

            revenue = price * qty
            drr = drr_base[cat] * drr_week_mult[wi]
            if (pid, wi) in promo_boost:
                drr *= 1.5
            ad_spend = round(revenue * drr)

            if qty > 0:
                logi = round(logi * random.uniform(0.92, 1.08))

            # Маржа = Выручка - Себестоимость - Комиссия - Логистика - Реклама
            cogs_total = cost * qty
            commission_total = round(revenue * comm_rate)
            margin = revenue - cogs_total - commission_total - logi - ad_spend

            sales_rows.append({
                "week": week_date,
                "wi": wi,
                "sku_id": pid,
                "name": name,
                "category": cat,
                "platform": platform,
                "qty": qty,
                "price": price,
                "cost": cost,
                "weight": weight,
                "revenue": revenue,
                "cogs": cogs_total,
                "commission_rate": comm_rate,
                "commission": commission_total,
                "logistics": logi,
                "ad_spend": ad_spend,
                "margin": margin,
            })

# Прогоним ads_rows random для синхронизации seed
# (в generate_ads_data.py после sales_rows идёт генерация ads_rows с random)
for sr in sales_rows:
    qty = sr["qty"]
    cat = sr["category"]
    platform = sr["platform"]
    wi = sr["wi"]
    if qty == 0:
        random.randint(80, 400)
        random.uniform(0.7, 1.3)
        random.uniform(0.3, 0.6)
        random.uniform(0.6, 1.0)
        random.uniform(0.2, 0.5)
    else:
        random.uniform(0.82, 1.18)
        random.uniform(0.82, 1.18)
        random.uniform(0.85, 1.15)
        random.uniform(0.88, 1.12)
        random.uniform(0.88, 1.12)
        random.uniform(0.85, 1.15)


# ============================================================
# 5. АГРЕГАЦИЯ ЗА ПОСЛЕДНИЙ МЕСЯЦ (W9-W12)
# ============================================================
sku_data = {}  # sku_id → {revenue, cogs, commission, logistics, ad_spend, margin, qty}
sku_first_week = {}  # sku_id → первый wi, когда qty > 0

# Определяем первую неделю появления каждого SKU
for row in sales_rows:
    sid = row["sku_id"]
    if row["qty"] > 0:
        if sid not in sku_first_week:
            sku_first_week[sid] = row["wi"]
        else:
            sku_first_week[sid] = min(sku_first_week[sid], row["wi"])

# Агрегация за последний месяц
for row in sales_rows:
    if row["wi"] not in LAST_MONTH_WEEKS:
        continue

    sid = row["sku_id"]
    if sid not in sku_data:
        sku_data[sid] = {
            "name": row["name"],
            "category": row["category"],
            "cost": row["cost"],
            "revenue": 0,
            "cogs": 0,
            "commission": 0,
            "logistics": 0,
            "ad_spend": 0,
            "margin": 0,
            "qty": 0,
        }
    d = sku_data[sid]
    d["revenue"] += row["revenue"]
    d["cogs"] += row["cogs"]
    d["commission"] += row["commission"]
    d["logistics"] += row["logistics"]
    d["ad_spend"] += row["ad_spend"]
    d["margin"] += row["margin"]
    d["qty"] += row["qty"]


# ============================================================
# 6. КЛАССИФИКАЦИЯ ABCDX
# ============================================================
results = []

for sid, d in sorted(sku_data.items()):
    marginality = (d["margin"] / d["revenue"] * 100) if d["revenue"] > 0 else 0

    # Определяем первое появление
    first_week = sku_first_week.get(sid, 0)
    is_novelty = first_week >= NOVELTY_THRESHOLD_WEEK  # появился с W5+

    results.append({
        "sku_id": sid,
        "name": d["name"],
        "category": d["category"],
        "cost": d["cost"],
        "qty": d["qty"],
        "revenue": d["revenue"],
        "cogs": d["cogs"],
        "commission": d["commission"],
        "logistics": d["logistics"],
        "ad_spend": d["ad_spend"],
        "margin": d["margin"],
        "marginality": marginality,
        "first_week": first_week,
        "is_novelty": is_novelty,
        "class": None,
    })

# Шаг 1: X — новинки (приоритет над всем)
for r in results:
    if r["is_novelty"]:
        r["class"] = "X"

# Шаг 2: D — убыточные (маржа <= 0), если не X
for r in results:
    if r["class"] is None and r["margin"] <= 0:
        r["class"] = "D"

# Шаг 3: ABC — из оставшихся
unclassified = [r for r in results if r["class"] is None]
unclassified.sort(key=lambda r: r["margin"], reverse=True)

total_margin = sum(r["margin"] for r in unclassified)
cumulative = 0

for r in unclassified:
    cumulative += r["margin"]
    cum_pct = cumulative / total_margin if total_margin > 0 else 0

    if cum_pct <= 0.60 and r["marginality"] >= 20:
        r["class"] = "A"
    elif cum_pct <= 0.60 and r["marginality"] >= 15:
        # Попал в top-60% по марже, но маржинальность 15-20% → B
        r["class"] = "B"
    elif cum_pct <= 0.60 and r["marginality"] < 15:
        # Попал в top-60% по марже, но маржинальность < 15% → C
        r["class"] = "C"
    elif cum_pct <= 0.80 and r["marginality"] >= 15:
        r["class"] = "B"
    elif cum_pct <= 0.80 and r["marginality"] < 15:
        r["class"] = "C"
    else:
        r["class"] = "C"


# ============================================================
# 7. ЗАПИСЬ В EXCEL
# ============================================================
INPUT_FILE = "/Users/vadimbakanov/Documents/_CODE/vibecommerce_test_code/PRJ_MARKETPLACE/ads_data_v1.0.xlsx"
OUTPUT_FILE = "/Users/vadimbakanov/Documents/_CODE/vibecommerce_test_code/PRJ_MARKETPLACE/ads_data_v2.0.xlsx"

wb = openpyxl.load_workbook(INPUT_FILE)

# --- Обновим лист «Каталог» — добавим колонки себестоимости, комиссий, логистики ---
ws_cat = wb["Каталог"]

new_headers = [
    ("Себестоимость, ₽", 16),
    ("Комиссия WB, %", 14),
    ("Комиссия Ozon, %", 15),
    ("Логистика WB, ₽/шт", 17),
    ("Логистика Ozon, ₽/шт", 18),
]

header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, size=11, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
num_fmt_rub = '#,##0" ₽"'
num_fmt_pct = "0.0%"
num_fmt_pct2 = "0.0%"
num_fmt_int = "#,##0"

for ci_offset, (hdr, width) in enumerate(new_headers):
    col = 6 + ci_offset  # F, G, H, I, J
    cell = ws_cat.cell(row=1, column=col, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws_cat.column_dimensions[get_column_letter(col)].width = width

for ri, p in enumerate(products, 2):
    pid, name, cat, cost, price_wb, price_ozon, weight, _, _ = p
    comm_wb = commission_rates["WB"][cat]
    comm_oz = commission_rates["Ozon"][cat]
    logi_wb = logistics_per_unit(weight, "WB")
    logi_oz = logistics_per_unit(weight, "Ozon")

    vals = [cost, comm_wb, comm_oz, logi_wb, logi_oz]
    fmts = [num_fmt_rub, num_fmt_pct, num_fmt_pct, num_fmt_rub, num_fmt_rub]
    for ci_offset, (v, fmt) in enumerate(zip(vals, fmts)):
        cell = ws_cat.cell(row=ri, column=6 + ci_offset, value=v)
        cell.number_format = fmt
        cell.border = thin_border

ws_cat.auto_filter.ref = f"A1:J{len(products) + 1}"

# --- Новый лист «ABCDX» ---
if "ABCDX" in wb.sheetnames:
    del wb["ABCDX"]

ws_abcdx = wb.create_sheet("ABCDX")

abcdx_headers = [
    ("SKU", 6),
    ("Название", 35),
    ("Категория", 18),
    ("Класс ABCDX", 13),
    ("Заказы, шт", 12),
    ("Выручка, ₽", 14),
    ("Себестоимость, ₽", 16),
    ("Комиссия МП, ₽", 15),
    ("Логистика, ₽", 13),
    ("Реклама, ₽", 12),
    ("Маржа, ₽", 13),
    ("Маржинальность, %", 17),
    ("Доля в общей марже, %", 18),
    ("Накопл. доля маржи, %", 19),
]

# Цвета для классов
class_colors = {
    "A": PatternFill("solid", fgColor="92D050"),  # зелёный
    "B": PatternFill("solid", fgColor="FFC000"),  # жёлтый
    "C": PatternFill("solid", fgColor="FF8C00"),  # оранжевый
    "D": PatternFill("solid", fgColor="FF4444"),  # красный
    "X": PatternFill("solid", fgColor="00B0F0"),  # голубой
}
class_fonts = {
    "A": Font(bold=True, size=11, color="000000"),
    "B": Font(bold=True, size=11, color="000000"),
    "C": Font(bold=True, size=11, color="000000"),
    "D": Font(bold=True, size=11, color="FFFFFF"),
    "X": Font(bold=True, size=11, color="FFFFFF"),
}

# Заголовки
for ci, (hdr, width) in enumerate(abcdx_headers, 1):
    cell = ws_abcdx.cell(row=1, column=ci, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws_abcdx.column_dimensions[get_column_letter(ci)].width = width

# Сортировка: A → B → C → D → X, внутри — по марже desc
class_order = {"A": 0, "B": 1, "C": 2, "D": 3, "X": 4}
results.sort(key=lambda r: (class_order.get(r["class"], 9), -r["margin"]))

# Рассчитаем доли маржи
positive_margin_total = sum(r["margin"] for r in results if r["margin"] > 0)

# Для накопительной доли маржи — считаем только по ABC (не D, не X)
abc_results = [r for r in results if r["class"] in ("A", "B", "C")]
abc_margin_total = sum(r["margin"] for r in abc_results)
abc_cumulative = 0

for ri, r in enumerate(results, 2):
    cls = r["class"]

    # Доля и накопительная доля
    share = r["margin"] / positive_margin_total * 100 if positive_margin_total > 0 and r["margin"] > 0 else 0
    if cls in ("A", "B", "C"):
        abc_cumulative += r["margin"]
        cum_share = abc_cumulative / abc_margin_total * 100 if abc_margin_total > 0 else 0
    else:
        cum_share = None

    vals = [
        r["sku_id"], r["name"], r["category"], cls,
        r["qty"], r["revenue"], r["cogs"], r["commission"],
        r["logistics"], r["ad_spend"], r["margin"],
        r["marginality"] / 100, share / 100,
        cum_share / 100 if cum_share is not None else "—",
    ]
    fmts = [
        None, None, None, None,
        num_fmt_int, num_fmt_rub, num_fmt_rub, num_fmt_rub,
        num_fmt_rub, num_fmt_rub, num_fmt_rub,
        num_fmt_pct, num_fmt_pct,
        num_fmt_pct if cum_share is not None else None,
    ]

    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws_abcdx.cell(row=ri, column=ci, value=v)
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt

    # Подсветка класса
    class_cell = ws_abcdx.cell(row=ri, column=4)
    if cls in class_colors:
        class_cell.fill = class_colors[cls]
        class_cell.font = class_fonts[cls]
        class_cell.alignment = Alignment(horizontal="center")

ws_abcdx.auto_filter.ref = f"A1:N{len(results) + 1}"
ws_abcdx.freeze_panes = "A2"


# --- Лист «Сводка ABCDX» ---
if "Сводка ABCDX" in wb.sheetnames:
    del wb["Сводка ABCDX"]

ws_summary = wb.create_sheet("Сводка ABCDX")

sum_headers = [
    ("Класс", 10),
    ("Кол-во SKU", 12),
    ("Заказы, шт", 12),
    ("Выручка, ₽", 14),
    ("Маржа, ₽", 14),
    ("Средн. маржинальность, %", 22),
    ("Доля выручки, %", 15),
    ("Доля маржи, %", 13),
]

for ci, (hdr, width) in enumerate(sum_headers, 1):
    cell = ws_summary.cell(row=1, column=ci, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws_summary.column_dimensions[get_column_letter(ci)].width = width

total_revenue_all = sum(r["revenue"] for r in results)
total_margin_all = sum(r["margin"] for r in results if r["margin"] > 0)

for ri, cls in enumerate(["A", "B", "C", "D", "X"], 2):
    cls_items = [r for r in results if r["class"] == cls]
    if not cls_items:
        vals = [cls, 0, 0, 0, 0, 0, 0, 0]
    else:
        n = len(cls_items)
        qty = sum(r["qty"] for r in cls_items)
        rev = sum(r["revenue"] for r in cls_items)
        mar = sum(r["margin"] for r in cls_items)
        avg_marginality = sum(r["marginality"] for r in cls_items) / n if n > 0 else 0
        rev_share = rev / total_revenue_all if total_revenue_all > 0 else 0
        mar_share = mar / total_margin_all if total_margin_all > 0 else 0
        vals = [cls, n, qty, rev, mar, avg_marginality / 100, rev_share, mar_share]

    fmts = [None, num_fmt_int, num_fmt_int, num_fmt_rub, num_fmt_rub, num_fmt_pct, num_fmt_pct, num_fmt_pct]
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cell = ws_summary.cell(row=ri, column=ci, value=v)
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt

    # Подсветка класса
    class_cell = ws_summary.cell(row=ri, column=1)
    if cls in class_colors:
        class_cell.fill = class_colors[cls]
        class_cell.font = class_fonts[cls]
        class_cell.alignment = Alignment(horizontal="center")

# Строка «Итого»
total_row = ri + 1
ws_summary.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True, size=11)
ws_summary.cell(row=total_row, column=2, value=len(results)).number_format = num_fmt_int
ws_summary.cell(row=total_row, column=3, value=sum(r["qty"] for r in results)).number_format = num_fmt_int
ws_summary.cell(row=total_row, column=4, value=total_revenue_all).number_format = num_fmt_rub
ws_summary.cell(row=total_row, column=5, value=sum(r["margin"] for r in results)).number_format = num_fmt_rub
for ci in range(1, 9):
    ws_summary.cell(row=total_row, column=ci).border = thin_border
    ws_summary.cell(row=total_row, column=ci).font = Font(bold=True, size=11)

ws_summary.freeze_panes = "A2"


# ============================================================
# 8. СОХРАНЕНИЕ
# ============================================================
wb.save(OUTPUT_FILE)

# ============================================================
# 9. ВЫВОД РЕЗУЛЬТАТОВ
# ============================================================
print(f"\n{'='*70}")
print(f"  ABCDX-АНАЛИЗ ТОВАРОВ")
print(f"  Период: последний месяц (12 янв — 2 фев 2026)")
print(f"{'='*70}\n")

print(f"{'Класс':<7} {'SKU':>4} {'Заказы':>8} {'Выручка':>12} {'Маржа':>12} {'Маржин.':>8}")
print("-" * 55)

for cls in ["A", "B", "C", "D", "X"]:
    cls_items = [r for r in results if r["class"] == cls]
    n = len(cls_items)
    qty = sum(r["qty"] for r in cls_items)
    rev = sum(r["revenue"] for r in cls_items)
    mar = sum(r["margin"] for r in cls_items)
    avg_m = sum(r["marginality"] for r in cls_items) / n if n > 0 else 0
    print(f"  {cls:<5} {n:>4} {qty:>8} {rev:>11,}₽ {mar:>11,}₽ {avg_m:>7.1f}%")

print("-" * 55)
total_qty = sum(r["qty"] for r in results)
total_rev = sum(r["revenue"] for r in results)
total_mar = sum(r["margin"] for r in results)
avg_total = sum(r["marginality"] for r in results) / len(results)
print(f"  ИТОГО {len(results):>3} {total_qty:>8} {total_rev:>11,}₽ {total_mar:>11,}₽ {avg_total:>7.1f}%")

print(f"\n--- Детализация по классам ---\n")

for cls, label in [("A", "ЛИДЕРЫ"), ("B", "СЕРЕДНЯКИ"), ("C", "АУТСАЙДЕРЫ"), ("D", "НА УДАЛЕНИЕ"), ("X", "НОВИНКИ")]:
    cls_items = [r for r in results if r["class"] == cls]
    if not cls_items:
        print(f"[{cls}] {label}: нет товаров\n")
        continue
    print(f"[{cls}] {label} ({len(cls_items)} SKU):")
    for r in cls_items:
        print(f"  SKU {r['sku_id']:>2}: {r['name']:<35} маржа={r['margin']:>8,}₽  маржин.={r['marginality']:>5.1f}%")
    print()

print(f"Файл сохранён: {OUTPUT_FILE}")
