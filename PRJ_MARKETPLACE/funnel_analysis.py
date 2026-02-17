"""
Анализ воронки продаж: Показы → Клики → Корзина → Заказы.
Источники: ads_data_v2.0.xlsx + sales_data_v1.0.xlsx
50 SKU × 2 площадки (WB + Ozon) × 12 недель (ноя 2025 – фев 2026).

Цель: найти узкие места (bottlenecks) воронки по SKU, категориям, площадкам.
Результат: Excel-отчёт (8 листов) + 8 PNG-графиков.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.lines import Line2D
from collections import defaultdict
import os

# ── Настройки ──
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['font.size'] = 10
plt.rcParams['axes.titlesize'] = 12
plt.rcParams['figure.dpi'] = 150

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ADS_FILE = os.path.join(SCRIPT_DIR, 'ads_data_v2.0.xlsx')
SALES_FILE = os.path.join(SCRIPT_DIR, 'sales_data_v1.0.xlsx')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'analysis_output')
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, 'funnel_report.xlsx')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Стили Excel ──
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, size=11, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
num_fmt_rub = '#,##0" ₽"'
num_fmt_pct = "0.0%"
num_fmt_int = "#,##0"
num_fmt_dec = "0.0"

# Цвета для подсветки
fill_red = PatternFill("solid", fgColor="FFC7CE")
fill_yellow = PatternFill("solid", fgColor="FFEB9C")
fill_green = PatternFill("solid", fgColor="C6EFCE")
fill_blue = PatternFill("solid", fgColor="BDD7EE")
fill_gray = PatternFill("solid", fgColor="F2F2F2")

# Цвета графиков
COLOR_WB = '#7b1fa2'
COLOR_OZON = '#1565c0'
COLORS_CAT = ['#1565c0', '#d32f2f', '#388e3c', '#f57c00', '#7b1fa2', '#00838f', '#c62828']


# ============================================================
# 1. ЗАГРУЗКА ДАННЫХ
# ============================================================
print("Загрузка данных...")

# --- ads_data_v2.0.xlsx: лист "Трафик и реклама" ---
# Читаем только RAW-колонки, формулы пересчитываем сами
wb_ads = openpyxl.load_workbook(ADS_FILE, read_only=True)
ws_ads = wb_ads['Трафик и реклама']

ads_data = []
for row in ws_ads.iter_rows(min_row=2, values_only=True):
    rec = {
        'week': row[0],
        'sku': row[1],
        'name': row[2],
        'category': row[3],
        'platform': row[4],
        'impr_total': row[5] or 0,
        'impr_ad': row[6] or 0,
        'clicks_total': row[8] or 0,
        'clicks_ad': row[9] or 0,
        'atc': row[12] or 0,
        'orders_total': row[14] or 0,
        'orders_ad': row[15] or 0,
        'price': row[18] or 0,
        'ad_spend': row[20] or 0,
    }
    # Вычисляемые метрики
    rec['revenue'] = rec['orders_total'] * rec['price']
    rec['ctr'] = rec['clicks_total'] / rec['impr_total'] if rec['impr_total'] > 0 else 0
    rec['cr_click_cart'] = rec['atc'] / rec['clicks_total'] if rec['clicks_total'] > 0 else 0
    rec['cr_cart_order'] = rec['orders_total'] / rec['atc'] if rec['atc'] > 0 else 0
    rec['cr_click_order'] = rec['orders_total'] / rec['clicks_total'] if rec['clicks_total'] > 0 else 0
    rec['ad_impr_share'] = rec['impr_ad'] / rec['impr_total'] if rec['impr_total'] > 0 else 0
    rec['ad_order_share'] = rec['orders_ad'] / rec['orders_total'] if rec['orders_total'] > 0 else 0
    rec['cpc'] = rec['ad_spend'] / rec['clicks_ad'] if rec['clicks_ad'] > 0 else 0
    rec['cpo'] = rec['ad_spend'] / rec['orders_ad'] if rec['orders_ad'] > 0 else 0
    rec['drr'] = rec['ad_spend'] / rec['revenue'] if rec['revenue'] > 0 else 0
    rec['roas'] = rec['revenue'] / rec['ad_spend'] if rec['ad_spend'] > 0 else 0
    ads_data.append(rec)

wb_ads.close()
print(f"  ads_data: {len(ads_data)} строк загружено")

# --- sales_data_v1.0.xlsx: лист "Продажи" ---
wb_sales = openpyxl.load_workbook(SALES_FILE, read_only=True)
ws_sales = wb_sales['Продажи']

sales_data = []
for row in ws_sales.iter_rows(min_row=2, values_only=True):
    rec = {
        'week': row[0],
        'sku': row[1],
        'name': row[2],
        'category': row[3],
        'platform': row[4],
        'qty': row[5] or 0,
        'price': row[6] or 0,
        'cost_unit': row[8] or 0,
        'comm_rate': row[10] or 0,
        'logistics': row[12] or 0,
        'storage': row[13] or 0,
        'ad_spend_sales': row[14] or 0,
    }
    rec['revenue'] = rec['qty'] * rec['price']
    rec['cogs'] = rec['qty'] * rec['cost_unit']
    rec['commission'] = rec['revenue'] * rec['comm_rate']
    rec['profit'] = rec['revenue'] - rec['cogs'] - rec['commission'] - rec['logistics'] - rec['storage'] - rec['ad_spend_sales']
    rec['margin_pct'] = rec['profit'] / rec['revenue'] if rec['revenue'] > 0 else 0
    sales_data.append(rec)

wb_sales.close()
print(f"  sales_data: {len(sales_data)} строк загружено")

# --- Связка P&L по ключу (week, sku, platform) ---
pnl_lookup = {}
for s in sales_data:
    key = (s['week'], s['sku'], s['platform'])
    pnl_lookup[key] = s

# --- Объединение ---
merged = []
matched = 0
for ad in ads_data:
    key = (ad['week'], ad['sku'], ad['platform'])
    pnl = pnl_lookup.get(key, {})
    combined = dict(ad)
    combined['profit'] = pnl.get('profit', 0)
    combined['margin_pct'] = pnl.get('margin_pct', 0)
    combined['cogs'] = pnl.get('cogs', 0)
    combined['commission'] = pnl.get('commission', 0)
    combined['logistics'] = pnl.get('logistics', 0)
    combined['storage'] = pnl.get('storage', 0)
    if pnl:
        matched += 1
    merged.append(combined)

print(f"  Объединено: {len(merged)} строк, из них с P&L: {matched}")


# ============================================================
# 2. АГРЕГАЦИЯ
# ============================================================
def aggregate_funnel(rows):
    """Агрегация метрик воронки для группы строк."""
    t_impr = sum(r['impr_total'] for r in rows)
    t_clicks = sum(r['clicks_total'] for r in rows)
    t_atc = sum(r['atc'] for r in rows)
    t_orders = sum(r['orders_total'] for r in rows)
    t_revenue = sum(r['revenue'] for r in rows)
    t_ad_spend = sum(r['ad_spend'] for r in rows)
    t_profit = sum(r['profit'] for r in rows)
    t_orders_ad = sum(r['orders_ad'] for r in rows)
    t_impr_ad = sum(r['impr_ad'] for r in rows)
    t_clicks_ad = sum(r['clicks_ad'] for r in rows)

    return {
        'impr': t_impr,
        'clicks': t_clicks,
        'atc': t_atc,
        'orders': t_orders,
        'revenue': t_revenue,
        'ad_spend': t_ad_spend,
        'profit': t_profit,
        'orders_ad': t_orders_ad,
        'impr_ad': t_impr_ad,
        'clicks_ad': t_clicks_ad,
        'ctr': t_clicks / t_impr if t_impr > 0 else 0,
        'cr_click_cart': t_atc / t_clicks if t_clicks > 0 else 0,
        'cr_cart_order': t_orders / t_atc if t_atc > 0 else 0,
        'cr_click_order': t_orders / t_clicks if t_clicks > 0 else 0,
        'drr': t_ad_spend / t_revenue if t_revenue > 0 else 0,
        'roas': t_revenue / t_ad_spend if t_ad_spend > 0 else 0,
        'cpo': t_ad_spend / t_orders_ad if t_orders_ad > 0 else 0,
        'cpc': t_ad_spend / t_clicks_ad if t_clicks_ad > 0 else 0,
        'margin_pct': t_profit / t_revenue if t_revenue > 0 else 0,
        'ad_order_share': t_orders_ad / t_orders if t_orders > 0 else 0,
    }


# --- Агрегация по SKU ---
print("\nАгрегация по SKU...")
sku_groups = defaultdict(list)
for r in merged:
    sku_groups[r['sku']].append(r)

sku_funnel = {}
for sku, rows in sorted(sku_groups.items()):
    agg = aggregate_funnel(rows)
    agg['sku'] = sku
    agg['name'] = rows[0]['name']
    agg['category'] = rows[0]['category']
    sku_funnel[sku] = agg

# --- Агрегация по категориям ---
cat_groups = defaultdict(list)
for r in merged:
    cat_groups[r['category']].append(r)

cat_funnel = {}
for cat, rows in sorted(cat_groups.items()):
    agg = aggregate_funnel(rows)
    agg['category'] = cat
    agg['sku_count'] = len(set(r['sku'] for r in rows))
    cat_funnel[cat] = agg

# --- Агрегация по (SKU, платформа) ---
sku_plat_groups = defaultdict(list)
for r in merged:
    sku_plat_groups[(r['sku'], r['platform'])].append(r)

sku_platform = defaultdict(dict)
for (sku, plat), rows in sku_plat_groups.items():
    sku_platform[sku][plat] = aggregate_funnel(rows)

# --- Агрегация по неделям ---
week_groups = defaultdict(list)
for r in merged:
    week_groups[r['week']].append(r)

weekly_funnel = []
for week in sorted(week_groups.keys()):
    agg = aggregate_funnel(week_groups[week])
    agg['week'] = week
    weekly_funnel.append(agg)

# --- Агрегация по (неделя, платформа) ---
week_plat_groups = defaultdict(list)
for r in merged:
    week_plat_groups[(r['week'], r['platform'])].append(r)


# ============================================================
# 3. ВЫЯВЛЕНИЕ УЗКИХ МЕСТ
# ============================================================
print("Выявление узких мест...")

# Бенчмарки: перцентили по категории
cat_benchmarks = defaultdict(lambda: defaultdict(list))
for sku, agg in sku_funnel.items():
    for metric in ['ctr', 'cr_click_cart', 'cr_cart_order']:
        cat_benchmarks[agg['category']][metric].append(agg[metric])

cat_p25 = {}
for cat, metrics in cat_benchmarks.items():
    cat_p25[cat] = {}
    for metric, values in metrics.items():
        sv = sorted(values)
        idx = max(0, len(sv) // 4 - 1)
        cat_p25[cat][metric] = sv[idx] if sv else 0

# Флаги узких мест + оценка потерянной выручки
stage_names = {
    'ctr': 'Показы\u2192Клики',
    'cr_click_cart': 'Клики\u2192Корзина',
    'cr_cart_order': 'Корзина\u2192Заказы',
}

# Медианы для сегментации
all_impr = sorted(agg['impr'] for agg in sku_funnel.values())
all_profit = sorted(agg['profit'] for agg in sku_funnel.values())
all_ad_spend = sorted(agg['ad_spend'] for agg in sku_funnel.values())
all_cr = sorted(agg['cr_click_order'] for agg in sku_funnel.values())
median_impr = all_impr[len(all_impr) // 2]
median_profit = all_profit[len(all_profit) // 2]
median_ad_spend = all_ad_spend[len(all_ad_spend) // 2]
median_cr = all_cr[len(all_cr) // 2]

for sku, agg in sku_funnel.items():
    cat = agg['category']
    cat_avg = cat_funnel[cat]

    # Узкие места
    bottlenecks = []
    for metric, stage_name in stage_names.items():
        if agg[metric] < cat_p25[cat].get(metric, 0):
            bottlenecks.append(stage_name)
    agg['bottlenecks'] = bottlenecks
    agg['bottleneck_str'] = '; '.join(bottlenecks) if bottlenecks else 'Нет'

    # Главное узкое место — этап с максимальным отставанием от среднего по категории
    max_gap = 0
    primary = 'Нет'
    for metric, stage_name in stage_names.items():
        cat_val = cat_avg[metric]
        if cat_val > 0:
            gap = (cat_val - agg[metric]) / cat_val
            if gap > max_gap:
                max_gap = gap
                primary = stage_name
    agg['primary_bottleneck'] = primary if max_gap > 0.05 else 'Нет'

    # Оценка потерянной выручки
    potential_clicks = agg['impr'] * cat_avg['ctr']
    lost_clicks = max(0, potential_clicks - agg['clicks'])
    potential_orders_from_clicks = lost_clicks * cat_avg['cr_click_order']

    potential_cart = agg['clicks'] * cat_avg['cr_click_cart']
    lost_cart = max(0, potential_cart - agg['atc'])
    potential_orders_from_cart = lost_cart * cat_avg['cr_cart_order']

    avg_order_value = agg['revenue'] / agg['orders'] if agg['orders'] > 0 else 0
    agg['lost_revenue'] = (potential_orders_from_clicks + potential_orders_from_cart) * avg_order_value

    # Сегмент
    if agg['impr'] > median_impr and agg['profit'] < median_profit:
        agg['segment'] = 'Трафик-пожиратель'
    elif agg['margin_pct'] > 0.15 and agg['impr'] < median_impr:
        agg['segment'] = 'Точка роста'
    elif agg['ad_spend'] > median_ad_spend and agg['profit'] < 0:
        agg['segment'] = 'Убыточный'
    elif agg['cr_click_order'] > median_cr and agg['impr'] < median_impr * 1.5:
        agg['segment'] = 'Масштабировать'
    else:
        agg['segment'] = 'Стабильный'


# ============================================================
# 4. РЕКОМЕНДАЦИИ
# ============================================================
print("Генерация рекомендаций...")

recommendations = []
for sku in sorted(sku_funnel.keys(), key=lambda s: -sku_funnel[s]['lost_revenue']):
    agg = sku_funnel[sku]
    cat_avg = cat_funnel[agg['category']]
    recs = []

    if agg['ctr'] < cat_avg['ctr'] * 0.8:
        recs.append(f"CTR {agg['ctr']*100:.2f}% < среднего {cat_avg['ctr']*100:.2f}%: улучшить главное фото и заголовок")
    if agg['cr_click_cart'] < cat_avg['cr_click_cart'] * 0.8:
        recs.append(f"CR клик\u2192корзина {agg['cr_click_cart']*100:.1f}% < {cat_avg['cr_click_cart']*100:.1f}%: доработать карточку (описание, фото, отзывы)")
    if agg['cr_cart_order'] < cat_avg['cr_cart_order'] * 0.8:
        recs.append(f"CR корзина\u2192заказ {agg['cr_cart_order']*100:.1f}% < {cat_avg['cr_cart_order']*100:.1f}%: проверить цену/доставку/условия")
    if agg['drr'] > 0.12:
        recs.append(f"ДРР {agg['drr']*100:.1f}% > 12%: оптимизировать рекламу")
    if agg['profit'] < 0:
        recs.append(f"Убыточен (прибыль {agg['profit']:,.0f} руб.): пересмотреть цену или снизить расход на рекламу")
    if agg['margin_pct'] > 0.2 and agg['impr'] < median_impr:
        recs.append(f"Высокая маржа {agg['margin_pct']*100:.1f}%, мало трафика: увеличить рекламу")

    if recs:
        recommendations.append({
            'sku': sku,
            'name': agg['name'],
            'category': agg['category'],
            'lost_revenue': agg['lost_revenue'],
            'primary_bottleneck': agg['primary_bottleneck'],
            'recs_text': '\n'.join(recs),
            'recs_list': recs,
        })


# ============================================================
# 5. ЗАПИСЬ EXCEL-ОТЧЁТА
# ============================================================
print("\nЗапись Excel-отчёта...")


def write_header_row(ws, headers_widths):
    """Заголовки с форматированием."""
    for ci, (hdr, width) in enumerate(headers_widths, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = width


def write_row(ws, row_idx, values, formats=None):
    """Запись строки с бордерами и форматами."""
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        cell.border = thin_border
        if formats and ci <= len(formats) and formats[ci - 1]:
            cell.number_format = formats[ci - 1]


wb_out = openpyxl.Workbook()

# ── ЛИСТ 1: Воронка по SKU ──
ws1 = wb_out.active
ws1.title = "Воронка по SKU"

headers1 = [
    ("SKU", 6), ("Название", 35), ("Категория", 18),
    ("Показы", 12), ("Клики", 10), ("Корзина", 10), ("Заказы", 10),
    ("CTR, %", 10), ("CR клик\u2192корз, %", 16), ("CR корз\u2192заказ, %", 16), ("CR клик\u2192заказ, %", 16),
    ("Выручка, \u20bd", 14), ("Прибыль, \u20bd", 13), ("Маржа, %", 10),
    ("ДРР, %", 9), ("ROAS", 8),
    ("Узкое место", 30), ("Сегмент", 18),
]
write_header_row(ws1, headers1)

fmts1 = [
    None, None, None,
    num_fmt_int, num_fmt_int, num_fmt_int, num_fmt_int,
    num_fmt_pct, num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_rub, num_fmt_rub, num_fmt_pct,
    num_fmt_pct, num_fmt_dec,
    None, None,
]

for ri, sku in enumerate(sorted(sku_funnel.keys()), 2):
    a = sku_funnel[sku]
    vals = [
        a['sku'], a['name'], a['category'],
        a['impr'], a['clicks'], a['atc'], a['orders'],
        a['ctr'], a['cr_click_cart'], a['cr_cart_order'], a['cr_click_order'],
        a['revenue'], a['profit'], a['margin_pct'],
        a['drr'], a['roas'],
        a['bottleneck_str'], a['segment'],
    ]
    write_row(ws1, ri, vals, fmts1)

    # Подсветка узких мест
    if a['bottlenecks']:
        ws1.cell(row=ri, column=17).fill = fill_red
    # Подсветка сегмента
    seg_cell = ws1.cell(row=ri, column=18)
    if a['segment'] == 'Убыточный':
        seg_cell.fill = fill_red
    elif a['segment'] == 'Трафик-пожиратель':
        seg_cell.fill = fill_yellow
    elif a['segment'] == 'Точка роста':
        seg_cell.fill = fill_green
    elif a['segment'] == 'Масштабировать':
        seg_cell.fill = fill_blue

ws1.auto_filter.ref = f"A1:R{len(sku_funnel) + 1}"
ws1.freeze_panes = "D2"

# ── ЛИСТ 2: WB vs Ozon ──
ws2 = wb_out.create_sheet("WB vs Ozon")

headers2 = [
    ("SKU", 6), ("Название", 30), ("Категория", 16),
    ("CTR WB", 9), ("CTR Ozon", 10), ("\u0394 CTR", 9),
    ("CR\u2192корз WB", 12), ("CR\u2192корз Ozon", 14), ("\u0394 CR\u2192корз", 12),
    ("CR\u2192заказ WB", 14), ("CR\u2192заказ Ozon", 15), ("\u0394 CR\u2192заказ", 13),
    ("Заказы WB", 11), ("Заказы Ozon", 12),
    ("Выручка WB", 13), ("Выручка Ozon", 14),
    ("ДРР WB", 9), ("ДРР Ozon", 10),
    ("Прибыль WB", 13), ("Прибыль Ozon", 14),
]
write_header_row(ws2, headers2)

fmts2 = [
    None, None, None,
    num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_int, num_fmt_int,
    num_fmt_rub, num_fmt_rub,
    num_fmt_pct, num_fmt_pct,
    num_fmt_rub, num_fmt_rub,
]

for ri, sku in enumerate(sorted(sku_funnel.keys()), 2):
    wb_data = sku_platform[sku].get('WB', {})
    oz_data = sku_platform[sku].get('Ozon', {})
    info = sku_funnel[sku]

    vals = [
        sku, info['name'], info['category'],
        wb_data.get('ctr', 0), oz_data.get('ctr', 0),
        wb_data.get('ctr', 0) - oz_data.get('ctr', 0),
        wb_data.get('cr_click_cart', 0), oz_data.get('cr_click_cart', 0),
        wb_data.get('cr_click_cart', 0) - oz_data.get('cr_click_cart', 0),
        wb_data.get('cr_click_order', 0), oz_data.get('cr_click_order', 0),
        wb_data.get('cr_click_order', 0) - oz_data.get('cr_click_order', 0),
        wb_data.get('orders', 0), oz_data.get('orders', 0),
        wb_data.get('revenue', 0), oz_data.get('revenue', 0),
        wb_data.get('drr', 0), oz_data.get('drr', 0),
        wb_data.get('profit', 0), oz_data.get('profit', 0),
    ]
    write_row(ws2, ri, vals, fmts2)

    # Подсветка дельт: зелёный если WB лучше, красный если хуже
    for col_idx in [6, 9, 12]:  # дельта-колонки
        cell = ws2.cell(row=ri, column=col_idx)
        if isinstance(cell.value, (int, float)):
            if cell.value > 0.002:
                cell.fill = fill_green
            elif cell.value < -0.002:
                cell.fill = fill_red

# Итоговая строка по платформам
ri_total = len(sku_funnel) + 2
wb_total = aggregate_funnel([r for r in merged if r['platform'] == 'WB'])
oz_total = aggregate_funnel([r for r in merged if r['platform'] == 'Ozon'])
ws2.cell(row=ri_total, column=1, value="ИТОГО").font = Font(bold=True, size=11)
total_vals = [
    None, "Все SKU", None,
    wb_total['ctr'], oz_total['ctr'], wb_total['ctr'] - oz_total['ctr'],
    wb_total['cr_click_cart'], oz_total['cr_click_cart'], wb_total['cr_click_cart'] - oz_total['cr_click_cart'],
    wb_total['cr_click_order'], oz_total['cr_click_order'], wb_total['cr_click_order'] - oz_total['cr_click_order'],
    wb_total['orders'], oz_total['orders'],
    wb_total['revenue'], oz_total['revenue'],
    wb_total['drr'], oz_total['drr'],
    wb_total['profit'], oz_total['profit'],
]
write_row(ws2, ri_total, total_vals, fmts2)
for ci in range(1, 21):
    ws2.cell(row=ri_total, column=ci).font = Font(bold=True, size=11)

ws2.auto_filter.ref = f"A1:T{len(sku_funnel) + 1}"
ws2.freeze_panes = "D2"

# ── ЛИСТ 3: Категории ──
ws3 = wb_out.create_sheet("Категории")

headers3 = [
    ("Категория", 18), ("SKU", 6),
    ("Показы", 12), ("Клики", 10), ("Корзина", 10), ("Заказы", 10),
    ("CTR, %", 10), ("CR клик\u2192корз, %", 16), ("CR корз\u2192заказ, %", 16), ("CR клик\u2192заказ, %", 16),
    ("Выручка, \u20bd", 14), ("Прибыль, \u20bd", 13), ("Маржа, %", 10),
    ("ДРР, %", 9), ("ROAS", 8),
]
write_header_row(ws3, headers3)

fmts3 = [
    None, num_fmt_int,
    num_fmt_int, num_fmt_int, num_fmt_int, num_fmt_int,
    num_fmt_pct, num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_rub, num_fmt_rub, num_fmt_pct,
    num_fmt_pct, num_fmt_dec,
]

for ri, cat in enumerate(sorted(cat_funnel.keys()), 2):
    a = cat_funnel[cat]
    vals = [
        cat, a['sku_count'],
        a['impr'], a['clicks'], a['atc'], a['orders'],
        a['ctr'], a['cr_click_cart'], a['cr_cart_order'], a['cr_click_order'],
        a['revenue'], a['profit'], a['margin_pct'],
        a['drr'], a['roas'],
    ]
    write_row(ws3, ri, vals, fmts3)

ws3.auto_filter.ref = f"A1:O{len(cat_funnel) + 1}"
ws3.freeze_panes = "A2"

# ── ЛИСТ 4: Рекламная эффективность ──
ws4 = wb_out.create_sheet("Рекламная эффективность")

headers4 = [
    ("SKU", 6), ("Название", 30), ("Категория", 16),
    ("Выручка, \u20bd", 13), ("Расход рекл., \u20bd", 15), ("Прибыль, \u20bd", 13),
    ("ДРР, %", 9), ("ROAS", 8), ("CPC, \u20bd", 9), ("CPO, \u20bd", 10),
    ("Доля рекл. заказов, %", 20),
    ("Заказы органика", 15), ("Заказы реклама", 15),
]
write_header_row(ws4, headers4)

fmts4 = [
    None, None, None,
    num_fmt_rub, num_fmt_rub, num_fmt_rub,
    num_fmt_pct, num_fmt_dec, num_fmt_rub, num_fmt_rub,
    num_fmt_pct,
    num_fmt_int, num_fmt_int,
]

# Сортировка по ДРР desc (самые неэффективные сверху)
sorted_by_drr = sorted(sku_funnel.values(), key=lambda a: -a['drr'])

for ri, a in enumerate(sorted_by_drr, 2):
    vals = [
        a['sku'], a['name'], a['category'],
        a['revenue'], a['ad_spend'], a['profit'],
        a['drr'], a['roas'], a['cpc'], a['cpo'],
        a['ad_order_share'],
        a['orders'] - a['orders_ad'], a['orders_ad'],
    ]
    write_row(ws4, ri, vals, fmts4)

    # Подсветка ДРР
    drr_cell = ws4.cell(row=ri, column=7)
    if a['drr'] > 0.15:
        drr_cell.fill = fill_red
    elif a['drr'] > 0.10:
        drr_cell.fill = fill_yellow
    else:
        drr_cell.fill = fill_green

    # Подсветка убыточных
    if a['profit'] < 0:
        ws4.cell(row=ri, column=6).fill = fill_red

ws4.auto_filter.ref = f"A1:M{len(sku_funnel) + 1}"
ws4.freeze_panes = "D2"

# ── ЛИСТ 5: Динамика по неделям ──
ws5 = wb_out.create_sheet("Динамика по неделям")

headers5 = [
    ("Неделя", 12),
    ("Показы", 12), ("Клики", 10), ("Корзина", 10), ("Заказы", 10),
    ("CTR, %", 10), ("CR клик\u2192корз, %", 16), ("CR корз\u2192заказ, %", 16), ("CR клик\u2192заказ, %", 16),
    ("Выручка, \u20bd", 14), ("Расход рекл., \u20bd", 15),
    ("ДРР, %", 9), ("ROAS", 8),
]
write_header_row(ws5, headers5)

fmts5 = [
    "DD.MM.YYYY",
    num_fmt_int, num_fmt_int, num_fmt_int, num_fmt_int,
    num_fmt_pct, num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_rub, num_fmt_rub,
    num_fmt_pct, num_fmt_dec,
]

for ri, wf in enumerate(weekly_funnel, 2):
    vals = [
        wf['week'],
        wf['impr'], wf['clicks'], wf['atc'], wf['orders'],
        wf['ctr'], wf['cr_click_cart'], wf['cr_cart_order'], wf['cr_click_order'],
        wf['revenue'], wf['ad_spend'],
        wf['drr'], wf['roas'],
    ]
    write_row(ws5, ri, vals, fmts5)

ws5.freeze_panes = "A2"

# ── ЛИСТ 6: Трафик vs Прибыль ──
ws6 = wb_out.create_sheet("Трафик vs Прибыль")

headers6 = [
    ("SKU", 6), ("Название", 30), ("Категория", 16),
    ("Показы", 12), ("Заказы", 10),
    ("Выручка, \u20bd", 13), ("Прибыль, \u20bd", 13), ("Маржа, %", 10),
    ("CR клик\u2192заказ, %", 16), ("ДРР, %", 9),
    ("Сегмент", 20),
]
write_header_row(ws6, headers6)

fmts6 = [
    None, None, None,
    num_fmt_int, num_fmt_int,
    num_fmt_rub, num_fmt_rub, num_fmt_pct,
    num_fmt_pct, num_fmt_pct,
    None,
]

sorted_by_impr = sorted(sku_funnel.values(), key=lambda a: -a['impr'])
for ri, a in enumerate(sorted_by_impr, 2):
    vals = [
        a['sku'], a['name'], a['category'],
        a['impr'], a['orders'],
        a['revenue'], a['profit'], a['margin_pct'],
        a['cr_click_order'], a['drr'],
        a['segment'],
    ]
    write_row(ws6, ri, vals, fmts6)

    seg_cell = ws6.cell(row=ri, column=11)
    if a['segment'] == 'Убыточный':
        seg_cell.fill = fill_red
    elif a['segment'] == 'Трафик-пожиратель':
        seg_cell.fill = fill_yellow
    elif a['segment'] == 'Точка роста':
        seg_cell.fill = fill_green
    elif a['segment'] == 'Масштабировать':
        seg_cell.fill = fill_blue

ws6.auto_filter.ref = f"A1:K{len(sku_funnel) + 1}"
ws6.freeze_panes = "D2"

# ── ЛИСТ 7: Топ проблемных SKU ──
ws7 = wb_out.create_sheet("Топ проблемных SKU")

headers7 = [
    ("SKU", 6), ("Название", 30), ("Категория", 16),
    ("Потер. выручка, \u20bd", 17), ("Главное узкое место", 22),
    ("CTR, %", 10), ("CR\u2192корз, %", 12), ("CR\u2192заказ, %", 12),
    ("CTR категории, %", 16), ("CR\u2192корз кат., %", 16), ("CR\u2192заказ кат., %", 16),
    ("Выручка, \u20bd", 13), ("Прибыль, \u20bd", 13),
]
write_header_row(ws7, headers7)

fmts7 = [
    None, None, None,
    num_fmt_rub, None,
    num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_pct, num_fmt_pct, num_fmt_pct,
    num_fmt_rub, num_fmt_rub,
]

top_problems = sorted(sku_funnel.values(), key=lambda a: -a['lost_revenue'])[:20]
for ri, a in enumerate(top_problems, 2):
    ca = cat_funnel[a['category']]
    vals = [
        a['sku'], a['name'], a['category'],
        a['lost_revenue'], a['primary_bottleneck'],
        a['ctr'], a['cr_click_cart'], a['cr_cart_order'],
        ca['ctr'], ca['cr_click_cart'], ca['cr_cart_order'],
        a['revenue'], a['profit'],
    ]
    write_row(ws7, ri, vals, fmts7)

    # Подсветка отстающих конверсий (красным если < 80% от категории)
    for sku_col, cat_col in [(6, 9), (7, 10), (8, 11)]:
        sku_val = ws7.cell(row=ri, column=sku_col).value
        cat_val = ws7.cell(row=ri, column=cat_col).value
        if isinstance(sku_val, (int, float)) and isinstance(cat_val, (int, float)) and cat_val > 0:
            if sku_val < cat_val * 0.8:
                ws7.cell(row=ri, column=sku_col).fill = fill_red
            elif sku_val < cat_val * 0.95:
                ws7.cell(row=ri, column=sku_col).fill = fill_yellow

ws7.freeze_panes = "D2"

# ── ЛИСТ 8: Рекомендации ──
ws8 = wb_out.create_sheet("Рекомендации")

headers8 = [
    ("SKU", 6), ("Название", 30), ("Категория", 16),
    ("Потер. выручка, \u20bd", 17), ("Главное узкое место", 22),
    ("Рекомендации", 80),
]
write_header_row(ws8, headers8)

fmts8 = [None, None, None, num_fmt_rub, None, None]

for ri, rec in enumerate(recommendations[:30], 2):
    vals = [
        rec['sku'], rec['name'], rec['category'],
        rec['lost_revenue'], rec['primary_bottleneck'],
        rec['recs_text'],
    ]
    write_row(ws8, ri, vals, fmts8)
    ws8.cell(row=ri, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    ws8.row_dimensions[ri].height = max(30, len(rec['recs_list']) * 16)

ws8.freeze_panes = "D2"

# Сохранение
wb_out.save(OUTPUT_XLSX)
print(f"  Excel: {OUTPUT_XLSX}")


# ============================================================
# 6. ГРАФИКИ
# ============================================================
print("\nГенерация графиков...")

# ── ГРАФИК 06: Общая воронка WB vs Ozon ──
fig, ax = plt.subplots(figsize=(14, 8))

stages = ['Показы', 'Клики', 'Корзина', 'Заказы']
wb_vals = [wb_total['impr'], wb_total['clicks'], wb_total['atc'], wb_total['orders']]
oz_vals = [oz_total['impr'], oz_total['clicks'], oz_total['atc'], oz_total['orders']]

x = range(len(stages))
bar_w = 0.35
bars_wb = ax.bar([i - bar_w/2 for i in x], wb_vals, bar_w, label='WB', color=COLOR_WB, alpha=0.85)
bars_oz = ax.bar([i + bar_w/2 for i in x], oz_vals, bar_w, label='Ozon', color=COLOR_OZON, alpha=0.85)

# Подписи значений
for bars in [bars_wb, bars_oz]:
    for bar in bars:
        h = bar.get_height()
        ax.text(bar.get_x() + bar.get_width() / 2, h, f'{h:,.0f}',
                ha='center', va='bottom', fontsize=9, fontweight='bold')

# Конверсии между этапами
wb_crs = [
    f"CTR: {wb_total['ctr']*100:.2f}%",
    f"CR: {wb_total['cr_click_cart']*100:.1f}%",
    f"CR: {wb_total['cr_cart_order']*100:.1f}%",
]
oz_crs = [
    f"CTR: {oz_total['ctr']*100:.2f}%",
    f"CR: {oz_total['cr_click_cart']*100:.1f}%",
    f"CR: {oz_total['cr_cart_order']*100:.1f}%",
]

for i in range(3):
    mid_x = i + 0.5
    y_pos = max(wb_vals[i], oz_vals[i]) * 0.65
    ax.annotate(f"WB: {wb_crs[i]}\nOzon: {oz_crs[i]}",
                xy=(mid_x, y_pos), fontsize=8, ha='center',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow', alpha=0.8))

ax.set_xticks(x)
ax.set_xticklabels(stages, fontsize=12, fontweight='bold')
ax.set_ylabel('Количество')
ax.set_title('Воронка продаж: WB vs Ozon (общая за 12 недель)\nПоказы \u2192 Клики \u2192 Корзина \u2192 Заказы',
             fontsize=14, fontweight='bold')
ax.legend(fontsize=12)
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
ax.grid(axis='y', alpha=0.2)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '06_funnel_overview.png'), bbox_inches='tight')
plt.close()
print("  06_funnel_overview.png")

# ── ГРАФИК 07: Тепловая карта конверсий по категориям ──
fig, ax = plt.subplots(figsize=(12, 7))

categories = sorted(cat_funnel.keys())
metrics_names = ['CTR, %', 'CR клик\u2192корзина, %', 'CR корзина\u2192заказ, %']
metrics_keys = ['ctr', 'cr_click_cart', 'cr_cart_order']

matrix = []
for cat in categories:
    row = [cat_funnel[cat][k] * 100 for k in metrics_keys]
    matrix.append(row)

im = ax.imshow(matrix, cmap='RdYlGn', aspect='auto', vmin=0)

ax.set_xticks(range(len(metrics_names)))
ax.set_xticklabels(metrics_names, fontsize=11, fontweight='bold')
ax.set_yticks(range(len(categories)))
ax.set_yticklabels(categories, fontsize=11)

# Значения в ячейках
for i in range(len(categories)):
    for j in range(len(metrics_keys)):
        val = matrix[i][j]
        color = 'white' if val < 3 or val > 25 else 'black'
        ax.text(j, i, f'{val:.2f}%', ha='center', va='center',
                fontsize=11, fontweight='bold', color=color)

ax.set_title('Конверсии воронки по категориям\n(среднее за 12 недель, WB + Ozon)',
             fontsize=14, fontweight='bold')
plt.colorbar(im, ax=ax, label='Конверсия, %', shrink=0.8)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '07_funnel_by_category.png'), bbox_inches='tight')
plt.close()
print("  07_funnel_by_category.png")

# ── ГРАФИК 08: Scatter — трафик vs конверсия ──
fig, ax = plt.subplots(figsize=(14, 10))

cat_colors = {cat: COLORS_CAT[i % len(COLORS_CAT)] for i, cat in enumerate(sorted(cat_funnel.keys()))}

for sku, a in sku_funnel.items():
    size = max(30, min(400, a['revenue'] / 5000))
    ax.scatter(a['impr'], a['cr_click_order'] * 100, s=size,
               c=cat_colors[a['category']], alpha=0.7, edgecolors='white', linewidth=0.5)

# Подписи для крайних SKU
for sku, a in sku_funnel.items():
    if a['lost_revenue'] > sorted([s['lost_revenue'] for s in sku_funnel.values()])[-6]:
        ax.annotate(f"SKU{sku}", (a['impr'], a['cr_click_order'] * 100),
                    fontsize=7, alpha=0.8, xytext=(5, 5), textcoords='offset points')

# Медианные линии (квадранты)
ax.axvline(x=median_impr, color='gray', linestyle='--', alpha=0.4)
ax.axhline(y=median_cr * 100, color='gray', linestyle='--', alpha=0.4)

# Подписи квадрантов
ax.text(0.02, 0.98, 'Нишевые\n(мало трафика, высокая CR)',
        transform=ax.transAxes, fontsize=9, va='top', color='#388e3c', style='italic')
ax.text(0.98, 0.98, 'Лидеры\n(много трафика, высокая CR)',
        transform=ax.transAxes, fontsize=9, va='top', ha='right', color='#1565c0', style='italic')
ax.text(0.02, 0.02, 'Проблемные\n(мало трафика, низкая CR)',
        transform=ax.transAxes, fontsize=9, va='bottom', color='gray', style='italic')
ax.text(0.98, 0.02, 'Узкое место\n(много трафика, низкая CR)',
        transform=ax.transAxes, fontsize=9, va='bottom', ha='right', color='#d32f2f', style='italic')

# Легенда по категориям
legend_els = [Line2D([0], [0], marker='o', color='w', markerfacecolor=cat_colors[c],
              markersize=10, label=c) for c in sorted(cat_colors.keys())]
ax.legend(handles=legend_els, loc='center left', bbox_to_anchor=(1.01, 0.5), fontsize=9)

ax.set_xlabel('Показы (всего за 12 недель)', fontsize=11)
ax.set_ylabel('CR клик \u2192 заказ, %', fontsize=11)
ax.set_title('Трафик vs Конверсия по SKU\n(размер = выручка)', fontsize=14, fontweight='bold')
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
ax.grid(True, alpha=0.2)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '08_bottleneck_scatter.png'), bbox_inches='tight')
plt.close()
print("  08_bottleneck_scatter.png")

# ── ГРАФИК 09: WB vs Ozon по категориям ──
fig, axes = plt.subplots(1, 3, figsize=(18, 6))

cat_plat_groups = defaultdict(lambda: defaultdict(list))
for r in merged:
    cat_plat_groups[r['category']][r['platform']].append(r)

cats_sorted = sorted(cat_funnel.keys())
metrics_plot = [
    ('ctr', 'CTR, %', axes[0]),
    ('cr_click_cart', 'CR клик\u2192корзина, %', axes[1]),
    ('cr_cart_order', 'CR корзина\u2192заказ, %', axes[2]),
]

for metric_key, metric_label, ax in metrics_plot:
    wb_vals_cat = []
    oz_vals_cat = []
    for cat in cats_sorted:
        wb_agg = aggregate_funnel(cat_plat_groups[cat].get('WB', []))
        oz_agg = aggregate_funnel(cat_plat_groups[cat].get('Ozon', []))
        wb_vals_cat.append(wb_agg[metric_key] * 100)
        oz_vals_cat.append(oz_agg[metric_key] * 100)

    y = range(len(cats_sorted))
    bar_h = 0.35
    ax.barh([i - bar_h/2 for i in y], wb_vals_cat, bar_h, label='WB', color=COLOR_WB, alpha=0.85)
    ax.barh([i + bar_h/2 for i in y], oz_vals_cat, bar_h, label='Ozon', color=COLOR_OZON, alpha=0.85)

    ax.set_yticks(list(y))
    ax.set_yticklabels(cats_sorted, fontsize=9)
    ax.set_xlabel(metric_label, fontsize=10)
    ax.set_title(metric_label, fontsize=12, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(axis='x', alpha=0.2)

fig.suptitle('Конверсии воронки: WB vs Ozon по категориям', fontsize=14, fontweight='bold', y=1.02)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '09_wb_vs_ozon_comparison.png'), bbox_inches='tight')
plt.close()
print("  09_wb_vs_ozon_comparison.png")

# ── ГРАФИК 10: Динамика конверсий по неделям ──
fig, axes = plt.subplots(3, 1, figsize=(14, 12), sharex=True)

weeks_dates = [wf['week'] for wf in weekly_funnel]
weeks_labels = [w.strftime('%d.%m') for w in weeks_dates]

# Данные по платформам
weeks_sorted = sorted(week_plat_groups.keys())
wb_weekly = []
oz_weekly = []
for week in sorted(set(w for w, _ in weeks_sorted)):
    wb_rows = week_plat_groups.get((week, 'WB'), [])
    oz_rows = week_plat_groups.get((week, 'Ozon'), [])
    if wb_rows:
        wb_weekly.append(aggregate_funnel(wb_rows))
    if oz_rows:
        oz_weekly.append(aggregate_funnel(oz_rows))

metrics_trend = [
    ('ctr', 'CTR (Показы \u2192 Клики), %', axes[0]),
    ('cr_click_cart', 'CR (Клики \u2192 Корзина), %', axes[1]),
    ('cr_cart_order', 'CR (Корзина \u2192 Заказы), %', axes[2]),
]

for metric_key, metric_label, ax in metrics_trend:
    overall_vals = [wf[metric_key] * 100 for wf in weekly_funnel]
    wb_vals_w = [w[metric_key] * 100 for w in wb_weekly]
    oz_vals_w = [w[metric_key] * 100 for w in oz_weekly]

    ax.plot(range(len(weeks_labels)), overall_vals, 'o-', color='#333333', linewidth=2.5,
            markersize=6, label='Общее', zorder=3)
    ax.plot(range(len(weeks_labels)), wb_vals_w, 's--', color=COLOR_WB, linewidth=1.5,
            markersize=5, alpha=0.7, label='WB')
    ax.plot(range(len(weeks_labels)), oz_vals_w, '^--', color=COLOR_OZON, linewidth=1.5,
            markersize=5, alpha=0.7, label='Ozon')

    # Вертикальные линии для BF и НГ
    ax.axvline(x=1, color='#d32f2f', linestyle=':', alpha=0.5, label='Black Friday' if metric_key == 'ctr' else None)
    ax.axvline(x=5, color='#388e3c', linestyle=':', alpha=0.5, label='НГ-пик' if metric_key == 'ctr' else None)

    ax.set_ylabel(metric_label, fontsize=10)
    ax.legend(fontsize=9, loc='best')
    ax.grid(True, alpha=0.2)

axes[2].set_xticks(range(len(weeks_labels)))
axes[2].set_xticklabels(weeks_labels, fontsize=9, rotation=45)
axes[2].set_xlabel('Неделя', fontsize=11)

fig.suptitle('Динамика конверсий воронки по неделям\n(ноя 2025 \u2014 фев 2026)',
             fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '10_funnel_trends_weekly.png'), bbox_inches='tight')
plt.close()
print("  10_funnel_trends_weekly.png")

# ── ГРАФИК 11: Scatter — ДРР vs ROAS ──
fig, ax = plt.subplots(figsize=(14, 10))

for sku, a in sku_funnel.items():
    size = max(30, min(400, a['ad_spend'] / 200))
    color = '#388e3c' if a['profit'] > 0 else '#d32f2f'
    ax.scatter(a['drr'] * 100, a['roas'], s=size, c=color, alpha=0.7,
               edgecolors='white', linewidth=0.5)

# Подписи для крайних
for sku, a in sku_funnel.items():
    if a['drr'] > 0.14 or a['roas'] > 18 or a['profit'] < -5000:
        ax.annotate(f"SKU{sku}", (a['drr'] * 100, a['roas']),
                    fontsize=7, alpha=0.8, xytext=(5, 5), textcoords='offset points')

ax.axvline(x=12, color='red', linestyle='--', alpha=0.4, label='ДРР = 12%')
ax.axhline(y=8, color='blue', linestyle='--', alpha=0.4, label='ROAS = 8')

ax.text(0.02, 0.98, 'Эффективная реклама\n(низкий ДРР, высокий ROAS)',
        transform=ax.transAxes, fontsize=9, va='top', color='#388e3c', style='italic')
ax.text(0.98, 0.02, 'Неэффективная реклама\n(высокий ДРР, низкий ROAS)',
        transform=ax.transAxes, fontsize=9, va='bottom', ha='right', color='#d32f2f', style='italic')

legend_els = [
    Line2D([0], [0], marker='o', color='w', markerfacecolor='#388e3c', markersize=10, label='Прибыльный'),
    Line2D([0], [0], marker='o', color='w', markerfacecolor='#d32f2f', markersize=10, label='Убыточный'),
]
ax.legend(handles=legend_els + ax.get_legend_handles_labels()[0], loc='upper right', fontsize=9)

ax.set_xlabel('ДРР, %', fontsize=11)
ax.set_ylabel('ROAS', fontsize=11)
ax.set_title('Эффективность рекламы по SKU\n(размер = расход на рекламу)', fontsize=14, fontweight='bold')
ax.grid(True, alpha=0.2)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '11_ad_efficiency_scatter.png'), bbox_inches='tight')
plt.close()
print("  11_ad_efficiency_scatter.png")

# ── ГРАФИК 12: Scatter — трафик vs прибыль ──
fig, ax = plt.subplots(figsize=(14, 10))

for sku, a in sku_funnel.items():
    size = max(30, min(400, a['revenue'] / 5000))
    ax.scatter(a['impr'], a['profit'], s=size,
               c=cat_colors[a['category']], alpha=0.7, edgecolors='white', linewidth=0.5)

# Подписи для крайних
for sku, a in sku_funnel.items():
    if a['profit'] < 0 or a['impr'] > all_impr[-3] or a['profit'] > all_profit[-3]:
        ax.annotate(f"SKU{sku}", (a['impr'], a['profit']),
                    fontsize=7, alpha=0.8, xytext=(5, 5), textcoords='offset points')

ax.axvline(x=median_impr, color='gray', linestyle='--', alpha=0.4)
ax.axhline(y=median_profit, color='gray', linestyle='--', alpha=0.4)
ax.axhline(y=0, color='red', linestyle='-', alpha=0.3)

ax.text(0.02, 0.98, 'Точки роста\n(мало трафика, высокая прибыль)',
        transform=ax.transAxes, fontsize=9, va='top', color='#388e3c', style='italic')
ax.text(0.98, 0.98, 'Лидеры\n(много трафика, высокая прибыль)',
        transform=ax.transAxes, fontsize=9, va='top', ha='right', color='#1565c0', style='italic')
ax.text(0.98, 0.02, 'Трафик-пожиратели\n(много трафика, мало прибыли)',
        transform=ax.transAxes, fontsize=9, va='bottom', ha='right', color='#d32f2f', style='italic')

legend_els = [Line2D([0], [0], marker='o', color='w', markerfacecolor=cat_colors[c],
              markersize=10, label=c) for c in sorted(cat_colors.keys())]
ax.legend(handles=legend_els, loc='center left', bbox_to_anchor=(1.01, 0.5), fontsize=9)

ax.set_xlabel('Показы (всего за 12 недель)', fontsize=11)
ax.set_ylabel('Прибыль, \u20bd', fontsize=11)
ax.set_title('Трафик vs Прибыль по SKU\n(размер = выручка)', fontsize=14, fontweight='bold')
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
ax.grid(True, alpha=0.2)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '12_traffic_vs_profit.png'), bbox_inches='tight')
plt.close()
print("  12_traffic_vs_profit.png")

# ── ГРАФИК 13: Топ проблемных SKU по потерянной выручке ──
fig, ax = plt.subplots(figsize=(14, 10))

top15 = sorted(sku_funnel.values(), key=lambda a: -a['lost_revenue'])[:15]

# Цвета по типу узкого места
bn_colors = {
    'Показы\u2192Клики': '#d32f2f',
    'Клики\u2192Корзина': '#f57c00',
    'Корзина\u2192Заказы': '#7b1fa2',
    'Нет': '#999999',
}

names = [f"SKU{a['sku']} {a['name'][:25]}" for a in top15]
values_lost = [a['lost_revenue'] for a in top15]
colors_bars = [bn_colors.get(a['primary_bottleneck'], '#999999') for a in top15]

bars = ax.barh(range(len(top15)), values_lost, color=colors_bars, height=0.7,
               edgecolor='white', linewidth=0.5)

# Подписи
for i, (bar, a) in enumerate(zip(bars, top15)):
    w = bar.get_width()
    ax.text(w + max(values_lost) * 0.01, i, f'{w:,.0f} \u20bd | {a["primary_bottleneck"]}',
            va='center', fontsize=8)

ax.set_yticks(range(len(top15)))
ax.set_yticklabels(names, fontsize=9)
ax.invert_yaxis()
ax.set_xlabel('Потерянная выручка (оценка), \u20bd', fontsize=11)
ax.set_title('Топ-15 SKU по потерянной выручке из-за проседания воронки\n(оценка: если бы конверсии были на уровне категории)',
             fontsize=13, fontweight='bold')

legend_els = [Line2D([0], [0], marker='s', color='w', markerfacecolor=c, markersize=12, label=k)
              for k, c in bn_colors.items() if k != 'Нет']
ax.legend(handles=legend_els, title='Тип узкого места', loc='lower right', fontsize=9)
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
ax.grid(axis='x', alpha=0.2)

plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, '13_top_bottlenecks.png'), bbox_inches='tight')
plt.close()
print("  13_top_bottlenecks.png")


# ============================================================
# 7. ИТОГОВЫЙ ВЫВОД В КОНСОЛЬ
# ============================================================
print(f"\n{'='*70}")
print(f"  АНАЛИЗ ВОРОНКИ ПРОДАЖ")
print(f"  Период: ноя 2025 \u2014 фев 2026 (12 недель)")
print(f"  SKU: 50 | Площадки: WB + Ozon")
print(f"{'='*70}")

# Общая воронка
total = aggregate_funnel(merged)
print(f"\n  ОБЩАЯ ВОРОНКА:")
print(f"  Показы:  {total['impr']:>12,}")
print(f"  Клики:   {total['clicks']:>12,}  (CTR: {total['ctr']*100:.2f}%)")
print(f"  Корзина: {total['atc']:>12,}  (CR клик\u2192корз: {total['cr_click_cart']*100:.1f}%)")
print(f"  Заказы:  {total['orders']:>12,}  (CR корз\u2192заказ: {total['cr_cart_order']*100:.1f}%)")

print(f"\n  WB vs OZON:")
print(f"  {'':>20} {'WB':>10} {'Ozon':>10}")
print(f"  {'CTR':>20} {wb_total['ctr']*100:>9.2f}% {oz_total['ctr']*100:>9.2f}%")
print(f"  {'CR клик\u2192корзина':>20} {wb_total['cr_click_cart']*100:>9.1f}% {oz_total['cr_click_cart']*100:>9.1f}%")
print(f"  {'CR корз\u2192заказ':>20} {wb_total['cr_cart_order']*100:>9.1f}% {oz_total['cr_cart_order']*100:>9.1f}%")
print(f"  {'ДРР':>20} {wb_total['drr']*100:>9.1f}% {oz_total['drr']*100:>9.1f}%")
print(f"  {'ROAS':>20} {wb_total['roas']:>9.1f}x {oz_total['roas']:>9.1f}x")

print(f"\n  КАТЕГОРИИ (CR клик\u2192заказ):")
for cat in sorted(cat_funnel.keys(), key=lambda c: -cat_funnel[c]['cr_click_order']):
    a = cat_funnel[cat]
    print(f"  {cat:<20} CR={a['cr_click_order']*100:.2f}%  CTR={a['ctr']*100:.2f}%  ДРР={a['drr']*100:.1f}%  маржа={a['margin_pct']*100:.1f}%")

# Топ-5 узких мест
print(f"\n  ТОП-5 ПРОБЛЕМНЫХ SKU (по потерянной выручке):")
for a in sorted(sku_funnel.values(), key=lambda a: -a['lost_revenue'])[:5]:
    print(f"  SKU{a['sku']:>2} {a['name']:<35} потеря ~{a['lost_revenue']:>10,.0f} \u20bd  |  {a['primary_bottleneck']}")

# Сегменты
segments = defaultdict(list)
for a in sku_funnel.values():
    segments[a['segment']].append(a)

print(f"\n  СЕГМЕНТЫ SKU:")
for seg in ['Точка роста', 'Масштабировать', 'Стабильный', 'Трафик-пожиратель', 'Убыточный']:
    items = segments.get(seg, [])
    print(f"  {seg:<22} {len(items):>3} SKU")

print(f"\n{'='*70}")
print(f"  ФАЙЛЫ:")
print(f"  Excel: {OUTPUT_XLSX}")
print(f"  Графики: {OUTPUT_DIR}/06-13_*.png (8 файлов)")
print(f"{'='*70}")
