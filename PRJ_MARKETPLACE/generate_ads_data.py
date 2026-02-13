"""
Генератор тестового датасета по рекламе и трафику для товаров для сна.
50 SKU × 2 площадки (WB + Ozon) × 12 недель (ноя 2025 – фев 2026).

ВАЖНО: данные мэтчатся с sales_data_v1.0.xlsx.
Общие ключи: (Неделя, SKU, Площадка).
Совпадающие поля: Заказы (шт), Цена (₽), Выручка (₽), Расход на рекламу (₽).

Для воспроизводимости используется random.seed(42) и идентичная логика
генерации sales_rows из generate_sales_data.py (секции 1–6).
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import random

random.seed(42)  # СОВПАДАЕТ с generate_sales_data.py — критично для мэтчинга!

# ============================================================
# 1. PRODUCT CATALOG  (идентично generate_sales_data.py)
# ============================================================
products = [
    # (id, name, category, cost, price_wb, price_ozon, weight_kg, base_vol_wb, base_vol_ozon)
    # --- Подушки (10 SKU) ---
    (1,  "Подушка бамбук 50×70",            "Подушки",  350,  990,  1050, 0.8,  45, 25),
    (2,  "Подушка бамбук 70×70",            "Подушки",  400,  1190, 1250, 1.0,  35, 20),
    (3,  "Подушка лебяжий пух 50×70",       "Подушки",  300,  890,  940,  0.7,  55, 30),
    (4,  "Подушка лебяжий пух 70×70",       "Подушки",  350,  1090, 1140, 0.9,  40, 22),
    (5,  "Подушка холлофайбер 50×70",       "Подушки",  200,  590,  620,  0.6,  70, 40),
    (6,  "Подушка холлофайбер 70×70",       "Подушки",  250,  690,  730,  0.8,  50, 28),
    (7,  "Подушка ортопед. Memory Foam",    "Подушки",  800,  2490, 2590, 1.2,  20, 15),
    (8,  "Подушка ортопед. с гелем",        "Подушки",  900,  2990, 3090, 1.3,  12, 10),
    (9,  "Подушка для беременных U-обр.",   "Подушки",  600,  1990, 2090, 2.5,  15, 12),
    (10, "Подушка-валик для шеи",           "Подушки",  250,  790,  840,  0.4,  30, 18),
    # --- Одеяла (8 SKU) ---
    (11, "Одеяло бамбук 1.5-сп",           "Одеяла",   500,  1490, 1550, 1.8,  35, 20),
    (12, "Одеяло бамбук 2-сп",             "Одеяла",   650,  1890, 1990, 2.2,  25, 15),
    (13, "Одеяло бамбук Евро",             "Одеяла",   750,  2190, 2290, 2.5,  20, 12),
    (14, "Одеяло лебяжий пух 1.5-сп",      "Одеяла",   400,  1290, 1350, 1.5,  40, 22),
    (15, "Одеяло лебяжий пух 2-сп",        "Одеяла",   550,  1590, 1690, 1.9,  28, 16),
    (16, "Одеяло лебяжий пух Евро",        "Одеяла",   650,  1890, 1990, 2.3,  18, 10),
    (17, "Одеяло шерстяное 1.5-сп",        "Одеяла",   700,  2490, 2590, 2.0,  12,  8),
    (18, "Одеяло облегч. летнее 1.5-сп",   "Одеяла",   350,  990,  1050, 1.0,   5,  3),
    # --- Постельное бельё (10 SKU) ---
    (19, "КПБ сатин 1.5-сп",              "Постельное бельё", 600, 1790, 1890, 1.5, 30, 18),
    (20, "КПБ сатин 2-сп",                "Постельное бельё", 700, 2190, 2290, 1.8, 22, 14),
    (21, "КПБ сатин Евро",                "Постельное бельё", 800, 2490, 2590, 2.0, 18, 11),
    (22, "КПБ поплин 1.5-сп",             "Постельное бельё", 400, 1190, 1250, 1.3, 40, 25),
    (23, "КПБ поплин 2-сп",               "Постельное бельё", 500, 1490, 1550, 1.6, 30, 18),
    (24, "КПБ поплин Евро",               "Постельное бельё", 600, 1690, 1790, 1.8, 22, 14),
    (25, "КПБ бязь 1.5-сп",               "Постельное бельё", 300,  890,  940, 1.2, 55, 35),
    (26, "КПБ бязь 2-сп",                 "Постельное бельё", 380, 1090, 1150, 1.4, 40, 25),
    (27, "Простыня на рез. сатин 160×200", "Постельное бельё", 300,  890,  940, 0.5, 35, 20),
    (28, "Простыня на рез. поплин 160×200","Постельное бельё", 200,  590,  640, 0.4, 50, 30),
    # --- Наматрасники (6 SKU) ---
    (29, "Наматрасник стёганый 160×200",      "Наматрасники", 400, 1290, 1350, 1.5, 20, 12),
    (30, "Наматрасник стёганый 180×200",      "Наматрасники", 450, 1490, 1550, 1.7, 15, 10),
    (31, "Наматрасник водонепрон. 160×200",   "Наматрасники", 500, 1590, 1690, 1.6, 18, 14),
    (32, "Наматрасник водонепрон. 180×200",   "Наматрасники", 550, 1790, 1890, 1.8, 12,  9),
    (33, "Наматрасник с бортами 160×200",     "Наматрасники", 350, 1090, 1150, 1.4, 15,  8),
    (34, "Наматрасник с бортами 180×200",     "Наматрасники", 400, 1290, 1350, 1.6, 10,  6),
    # --- Пледы (5 SKU) ---
    (35, "Плед микрофибра 150×200",   "Пледы", 300,  990, 1050, 1.0, 40, 22),
    (36, "Плед микрофибра 200×220",   "Пледы", 400, 1390, 1450, 1.4, 25, 15),
    (37, "Плед шерпа 150×200",        "Пледы", 450, 1590, 1690, 1.3, 30, 18),
    (38, "Плед велсофт 150×200",      "Пледы", 250,  790,  840, 0.9, 35, 20),
    (39, "Плед с рукавами",           "Пледы", 350, 1190, 1250, 1.2, 20, 15),
    # --- Аксессуары для сна (6 SKU) ---
    (40, "Маска для сна шёлковая",     "Аксессуары", 150,  590,  640, 0.05, 60, 35),
    (41, "Маска для сна 3D",           "Аксессуары", 100,  390,  420, 0.04, 80, 50),
    (42, "Маска для сна с охлаждением","Аксессуары", 200,  690,  740, 0.08, 30, 18),
    (43, "Беруши силиконовые (набор)", "Аксессуары",  80,  290,  320, 0.02,100, 60),
    (44, "Спрей для подушки лаванда",  "Аксессуары", 120,  490,  540, 0.15, 45, 28),
    (45, "Ночник-проектор звёзд",      "Аксессуары", 350, 1290, 1350, 0.30, 25, 20),
    # --- Детский сон (5 SKU) ---
    (46, "Подушка детская 40×60",  "Детский сон", 200,  690,  740, 0.4, 25, 15),
    (47, "Одеяло детское 110×140", "Детский сон", 300,  990, 1050, 1.0, 20, 12),
    (48, "КПБ детский поплин",     "Детский сон", 350, 1090, 1150, 0.8, 18, 12),
    (49, "Бортики в кроватку",     "Детский сон", 400, 1390, 1450, 1.5, 12,  8),
    (50, "Ночник детский LED",     "Детский сон", 250,  890,  940, 0.2, 20, 15),
]

# ============================================================
# 2. WEEKS  (идентично generate_sales_data.py)
# ============================================================
weeks = [
    datetime(2025, 11, 17),  # W1
    datetime(2025, 11, 24),  # W2 — Black Friday
    datetime(2025, 12,  1),  # W3
    datetime(2025, 12,  8),  # W4
    datetime(2025, 12, 15),  # W5
    datetime(2025, 12, 22),  # W6 — НГ-пик
    datetime(2025, 12, 29),  # W7 — НГ-каникулы
    datetime(2026,  1,  5),  # W8 — пост-НГ спад
    datetime(2026,  1, 12),  # W9
    datetime(2026,  1, 19),  # W10
    datetime(2026,  1, 26),  # W11
    datetime(2026,  2,  2),  # W12
]

# ============================================================
# 3. SEASONAL & CATEGORY MULTIPLIERS  (идентично generate_sales_data.py)
# ============================================================
season_overall = [0.90, 1.25, 1.00, 1.05, 1.15, 1.45, 1.15, 0.55, 0.65, 0.78, 0.85, 0.90]

cat_season = {
    "Подушки":          [1.00, 1.10, 1.00, 1.00, 1.05, 1.15, 1.00, 0.85, 0.90, 1.00, 1.00, 1.00],
    "Одеяла":           [1.10, 1.10, 1.10, 1.15, 1.20, 1.30, 1.10, 0.70, 0.80, 0.90, 0.95, 1.00],
    "Постельное бельё": [1.00, 1.05, 1.00, 1.00, 1.00, 1.15, 1.00, 0.90, 0.95, 1.00, 1.00, 1.00],
    "Наматрасники":     [1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 0.90, 0.95, 1.00, 1.00, 1.00],
    "Пледы":            [1.10, 1.15, 1.10, 1.15, 1.20, 1.40, 1.10, 0.60, 0.70, 0.80, 0.85, 0.90],
    "Аксессуары":       [0.90, 1.00, 1.00, 1.00, 1.10, 1.50, 1.15, 0.70, 0.80, 0.90, 0.95, 1.00],
    "Детский сон":      [1.00, 1.00, 1.00, 1.00, 1.05, 1.30, 1.05, 0.80, 0.85, 0.90, 0.95, 1.00],
}

# ============================================================
# 4. COMMISSION, DRR, LOGISTICS  (идентично generate_sales_data.py)
# ============================================================
commission_rates = {
    "WB":  {"Подушки": 0.15, "Одеяла": 0.15, "Постельное бельё": 0.15,
            "Наматрасники": 0.15, "Пледы": 0.15, "Аксессуары": 0.18, "Детский сон": 0.15},
    "Ozon": {"Подушки": 0.12, "Одеяла": 0.12, "Постельное бельё": 0.12,
             "Наматрасники": 0.12, "Пледы": 0.12, "Аксессуары": 0.15, "Детский сон": 0.12},
}

drr_base = {
    "Подушки": 0.08, "Одеяла": 0.07, "Постельное бельё": 0.09,
    "Наматрасники": 0.06, "Пледы": 0.07, "Аксессуары": 0.10, "Детский сон": 0.07,
}
drr_week_mult = [1.0, 1.3, 1.0, 1.0, 1.1, 1.2, 0.9, 0.5, 0.7, 0.8, 0.9, 1.0]


def logistics_per_unit(weight_kg, platform):
    if platform == "WB":
        return round(55 + weight_kg * 12)
    else:
        return round(45 + weight_kg * 14)


def storage_weekly(weight_kg, qty, platform):
    per_unit = 2.5 + weight_kg * 1.8 if platform == "WB" else 2.0 + weight_kg * 1.5
    return round(per_unit * qty)


# ============================================================
# 5. SPECIAL PATTERNS  (идентично generate_sales_data.py)
# ============================================================
oos_rules = {(5, "WB", 8): 0, (5, "WB", 9): 0}

discount_rules = {}
for p in products:
    if p[3] and p[0] in range(11, 19):
        for wi in range(7, 12):
            discount_rules[(p[0], "WB", wi)] = 0.90

trend_rules = {7: 1.05}
promo_boost = {(37, 4): 1.8, (37, 5): 2.0}


# ============================================================
# 6. REPRODUCE SALES_ROWS  (идентичная логика для мэтчинга)
# ============================================================
def gen_qty(base, overall_m, cat_m, noise=0.20):
    val = base * overall_m * cat_m
    val *= random.uniform(1 - noise, 1 + noise)
    return max(0, round(val))


sales_rows = []

for wi, week_date in enumerate(weeks):
    for p in products:
        pid, name, cat, cost, price_wb, price_ozon, weight, base_wb, base_ozon = p

        for platform in ["WB", "Ozon"]:
            base_vol = base_wb if platform == "WB" else base_ozon
            price = price_wb if platform == "WB" else price_ozon

            overall_m = season_overall[wi]
            cat_m = cat_season[cat][wi]

            trend_m = 1.0
            if pid in trend_rules:
                trend_m = trend_rules[pid] ** wi

            if (pid, platform, wi) in oos_rules:
                qty = 0
            else:
                qty = gen_qty(base_vol, overall_m, cat_m, noise=0.20)
                qty = round(qty * trend_m)

            # каннибализация SKU 1 ↔ 3
            if pid == 1 and wi % 2 == 0:
                qty = round(qty * 1.12)
            elif pid == 1 and wi % 2 == 1:
                qty = round(qty * 0.88)
            elif pid == 3 and wi % 2 == 0:
                qty = round(qty * 0.88)
            elif pid == 3 and wi % 2 == 1:
                qty = round(qty * 1.12)

            if (pid, platform, wi) in discount_rules:
                price = round(price * discount_rules[(pid, platform, wi)])

            promo_m = 1.0
            if (pid, wi) in promo_boost:
                promo_m = promo_boost[(pid, wi)]
                qty = round(qty * promo_m)

            if pid == 18:  # летнее одеяло
                qty = max(0, round(qty * 0.3))

            comm_rate = commission_rates[platform][cat]
            logi = logistics_per_unit(weight, platform) * qty
            stor = storage_weekly(weight, qty, platform)

            revenue = price * qty
            drr = drr_base[cat] * drr_week_mult[wi]
            if (pid, wi) in promo_boost:
                drr *= 1.5
            ad_spend = round(revenue * drr)

            if qty > 0:
                logi = round(logi * random.uniform(0.92, 1.08))
                stor = round(stor * random.uniform(0.90, 1.10))

            sales_rows.append({
                "week": week_date,
                "wi": wi,
                "sku_id": pid,
                "name": name,
                "category": cat,
                "platform": platform,
                "qty": qty,
                "price": price,
                "revenue": revenue,
                "ad_spend": ad_spend,
            })


# ============================================================
# 7. ПАРАМЕТРЫ ТРАФИКА И РЕКЛАМЫ
# ============================================================

# Конверсия клик → заказ (CR) по категориям
base_cr = {
    "Подушки": 0.040, "Одеяла": 0.032, "Постельное бельё": 0.038,
    "Наматрасники": 0.048, "Пледы": 0.042, "Аксессуары": 0.028,
    "Детский сон": 0.035,
}

# CTR (клик / показ) по категориям
base_ctr = {
    "Подушки": 0.055, "Одеяла": 0.048, "Постельное бельё": 0.058,
    "Наматрасники": 0.052, "Пледы": 0.062, "Аксессуары": 0.044,
    "Детский сон": 0.050,
}

# Добавление в корзину (от кликов) по категориям
base_atc = {
    "Подушки": 0.18, "Одеяла": 0.15, "Постельное бельё": 0.20,
    "Наматрасники": 0.22, "Пледы": 0.19, "Аксессуары": 0.13,
    "Детский сон": 0.17,
}

# Доля рекламного трафика по площадкам
ad_params = {
    "WB":   {"impr_share": 0.28, "click_share": 0.38, "order_share": 0.35},
    "Ozon": {"impr_share": 0.33, "click_share": 0.43, "order_share": 0.40},
}

# Сезонный множитель интенсивности рекламы (больше вокруг BF/НГ)
ad_intensity = [1.0, 1.3, 1.0, 1.0, 1.1, 1.2, 0.9, 0.5, 0.7, 0.8, 0.9, 1.0]


# ============================================================
# 8. ГЕНЕРАЦИЯ ДАННЫХ ПО РЕКЛАМЕ И ТРАФИКУ
# ============================================================
ads_rows = []

for sr in sales_rows:
    qty = sr["qty"]
    price = sr["price"]
    revenue = sr["revenue"]
    ad_spend = sr["ad_spend"]
    cat = sr["category"]
    platform = sr["platform"]
    wi = sr["wi"]

    if qty == 0:
        # OOS / нулевые продажи — есть немного показов, но нет конверсий
        base_impr = random.randint(80, 400)
        impr_total = base_impr
        impr_ad = min(
            round(impr_total * ad_params[platform]["impr_share"]
                  * random.uniform(0.7, 1.3)),
            impr_total,
        )
        clicks_total = max(
            round(impr_total * base_ctr[cat] * random.uniform(0.3, 0.6)), 0
        )
        clicks_ad = min(
            max(round(clicks_total * ad_params[platform]["click_share"]
                      * random.uniform(0.6, 1.0)), 0),
            clicks_total,
        )
        atc = max(
            round(clicks_total * base_atc[cat] * random.uniform(0.2, 0.5)), 0
        )

        ads_rows.append({
            "week": sr["week"], "sku_id": sr["sku_id"], "name": sr["name"],
            "category": cat, "platform": platform,
            "impr_total": impr_total, "impr_ad": impr_ad,
            "clicks_total": clicks_total, "clicks_ad": clicks_ad,
            "atc": atc,
            "orders_total": 0, "orders_ad": 0,
            "price": price, "revenue": 0, "ad_spend": ad_spend,
        })
        continue

    # --- Нормальный кейс: qty > 0 ---

    # Воронка: заказы → клики → показы (обратное вычисление)
    cr = base_cr[cat] * random.uniform(0.82, 1.18)
    clicks_total = max(round(qty / cr), qty + 1)

    ctr = base_ctr[cat] * random.uniform(0.82, 1.18)
    impr_total = max(round(clicks_total / ctr), clicks_total + 1)

    atc_rate = base_atc[cat] * random.uniform(0.85, 1.15)
    atc = max(round(clicks_total * atc_rate), qty)

    # Разделение на рекламный и органический трафик
    intensity = ad_intensity[wi]

    ad_click_share = min(
        ad_params[platform]["click_share"] * intensity
        * random.uniform(0.88, 1.12),
        0.65,
    )
    clicks_ad = max(round(clicks_total * ad_click_share), 1)

    ad_impr_share = min(
        ad_params[platform]["impr_share"] * intensity
        * random.uniform(0.88, 1.12),
        0.55,
    )
    impr_ad = max(round(impr_total * ad_impr_share), clicks_ad)

    ad_order_share = min(
        ad_params[platform]["order_share"] * random.uniform(0.85, 1.15),
        0.60,
    )
    orders_ad = min(max(round(qty * ad_order_share), 1), qty)

    ads_rows.append({
        "week": sr["week"], "sku_id": sr["sku_id"], "name": sr["name"],
        "category": cat, "platform": platform,
        "impr_total": impr_total, "impr_ad": impr_ad,
        "clicks_total": clicks_total, "clicks_ad": clicks_ad,
        "atc": atc,
        "orders_total": qty, "orders_ad": orders_ad,
        "price": price, "revenue": revenue, "ad_spend": ad_spend,
    })


# ============================================================
# 9. ЗАПИСЬ EXCEL — ads_data_v1.0.xlsx
# ============================================================
wb = openpyxl.Workbook()

# --- Стили ---
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
num_fmt_rub = '#,##0" ₽"'
num_fmt_pct = "0.0%"
num_fmt_int = "#,##0"
num_fmt_dec = "#,##0.0"
date_fmt = "DD.MM.YYYY"


def write_header(ws, headers, widths):
    """Записать заголовки с форматированием."""
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w


# ==================== ЛИСТ 1: Каталог ====================
ws_cat = wb.active
ws_cat.title = "Каталог"

write_header(ws_cat,
    ["SKU", "Название", "Категория", "Цена WB, ₽", "Цена Ozon, ₽"],
    [6, 35, 18, 12, 13],
)

for ri, p in enumerate(products, 2):
    vals = [p[0], p[1], p[2], p[4], p[5]]
    for ci, v in enumerate(vals, 1):
        cell = ws_cat.cell(row=ri, column=ci, value=v)
        cell.border = thin_border
        if ci in (4, 5):
            cell.number_format = num_fmt_rub

ws_cat.auto_filter.ref = f"A1:E{len(products) + 1}"
ws_cat.freeze_panes = "A2"


# ==================== ЛИСТ 2: Трафик и реклама ====================
ws = wb.create_sheet("Трафик и реклама")

ad_headers = [
    "Неделя",                    # A
    "SKU",                       # B
    "Название",                  # C
    "Категория",                 # D
    "Площадка",                  # E
    "Показы (всего)",            # F
    "Показы (реклама)",          # G
    "Доля рекл. показов, %",    # H — formula
    "Клики (всего)",             # I
    "Клики (реклама)",           # J
    "CTR общий, %",              # K — formula
    "CTR реклама, %",            # L — formula
    "Добавления в корзину",      # M
    "CR клик→корзина, %",        # N — formula
    "Заказы (всего)",            # O — = qty из sales_data
    "Заказы (с рекламы)",        # P
    "CR клик→заказ, %",          # Q — formula
    "Доля рекл. заказов, %",     # R — formula
    "Цена, ₽",                   # S
    "Выручка, ₽",                # T — formula
    "Расход на рекламу, ₽",      # U
    "CPC, ₽",                    # V — formula
    "CPO, ₽",                    # W — formula
    "ДРР, %",                    # X — formula
    "ROAS",                      # Y — formula
]
ad_widths = [
    12, 6, 35, 18, 10,     # A–E
    14, 14, 17,             # F–H
    12, 14,                 # I–J
    12, 14,                 # K–L
    18, 17,                 # M–N
    13, 16,                 # O–P
    15, 18,                 # Q–R
    10, 13, 18,             # S–U
    10, 10, 10, 8,          # V–Y
]

write_header(ws, ad_headers, ad_widths)

for ri, row in enumerate(ads_rows, 2):
    r = ri

    # A–E: идентификация строки
    ws.cell(row=r, column=1, value=row["week"]).number_format = date_fmt
    ws.cell(row=r, column=2, value=row["sku_id"])
    ws.cell(row=r, column=3, value=row["name"])
    ws.cell(row=r, column=4, value=row["category"])
    ws.cell(row=r, column=5, value=row["platform"])

    # F–G: показы
    ws.cell(row=r, column=6, value=row["impr_total"]).number_format = num_fmt_int
    ws.cell(row=r, column=7, value=row["impr_ad"]).number_format = num_fmt_int

    # H: доля рекламных показов (формула)
    ws.cell(row=r, column=8).value = f"=IF(F{r}=0,0,G{r}/F{r})"
    ws.cell(row=r, column=8).number_format = num_fmt_pct

    # I–J: клики
    ws.cell(row=r, column=9, value=row["clicks_total"]).number_format = num_fmt_int
    ws.cell(row=r, column=10, value=row["clicks_ad"]).number_format = num_fmt_int

    # K: CTR общий (формула)
    ws.cell(row=r, column=11).value = f"=IF(F{r}=0,0,I{r}/F{r})"
    ws.cell(row=r, column=11).number_format = num_fmt_pct

    # L: CTR реклама (формула)
    ws.cell(row=r, column=12).value = f"=IF(G{r}=0,0,J{r}/G{r})"
    ws.cell(row=r, column=12).number_format = num_fmt_pct

    # M: добавления в корзину
    ws.cell(row=r, column=13, value=row["atc"]).number_format = num_fmt_int

    # N: CR клик→корзина (формула)
    ws.cell(row=r, column=14).value = f"=IF(I{r}=0,0,M{r}/I{r})"
    ws.cell(row=r, column=14).number_format = num_fmt_pct

    # O: заказы всего (= qty из sales_data!)
    ws.cell(row=r, column=15, value=row["orders_total"]).number_format = num_fmt_int

    # P: заказы с рекламы
    ws.cell(row=r, column=16, value=row["orders_ad"]).number_format = num_fmt_int

    # Q: CR клик→заказ (формула)
    ws.cell(row=r, column=17).value = f"=IF(I{r}=0,0,O{r}/I{r})"
    ws.cell(row=r, column=17).number_format = num_fmt_pct

    # R: доля рекламных заказов (формула)
    ws.cell(row=r, column=18).value = f"=IF(O{r}=0,0,P{r}/O{r})"
    ws.cell(row=r, column=18).number_format = num_fmt_pct

    # S: цена
    ws.cell(row=r, column=19, value=row["price"]).number_format = num_fmt_rub

    # T: выручка (формула = заказы × цена)
    ws.cell(row=r, column=20).value = f"=O{r}*S{r}"
    ws.cell(row=r, column=20).number_format = num_fmt_rub

    # U: расход на рекламу
    ws.cell(row=r, column=21, value=row["ad_spend"]).number_format = num_fmt_rub

    # V: CPC (формула)
    ws.cell(row=r, column=22).value = f"=IF(J{r}=0,0,U{r}/J{r})"
    ws.cell(row=r, column=22).number_format = num_fmt_dec

    # W: CPO (формула)
    ws.cell(row=r, column=23).value = f"=IF(P{r}=0,0,U{r}/P{r})"
    ws.cell(row=r, column=23).number_format = num_fmt_rub

    # X: ДРР (формула)
    ws.cell(row=r, column=24).value = f"=IF(T{r}=0,0,U{r}/T{r})"
    ws.cell(row=r, column=24).number_format = num_fmt_pct

    # Y: ROAS (формула)
    ws.cell(row=r, column=25).value = f"=IF(U{r}=0,0,T{r}/U{r})"
    ws.cell(row=r, column=25).number_format = "0.0"

    # бордеры для всех ячеек строки
    for ci in range(1, 26):
        ws.cell(row=r, column=ci).border = thin_border

ws.auto_filter.ref = f"A1:Y{len(ads_rows) + 1}"
ws.freeze_panes = "A2"


# ==================== ЛИСТ 3: Сводка по категориям ====================
ws_sum = wb.create_sheet("Сводка")

# Агрегация: категория × неделя (WB + Ozon суммарно)
summary = defaultdict(lambda: {
    "impr": 0, "impr_ad": 0, "clicks": 0, "clicks_ad": 0,
    "atc": 0, "orders": 0, "orders_ad": 0, "revenue": 0, "ad_spend": 0,
})

for row in ads_rows:
    key = (row["category"], row["week"])
    s = summary[key]
    s["impr"] += row["impr_total"]
    s["impr_ad"] += row["impr_ad"]
    s["clicks"] += row["clicks_total"]
    s["clicks_ad"] += row["clicks_ad"]
    s["atc"] += row["atc"]
    s["orders"] += row["orders_total"]
    s["orders_ad"] += row["orders_ad"]
    s["revenue"] += row["revenue"]
    s["ad_spend"] += row["ad_spend"]

sum_headers = [
    "Категория",           # A
    "Неделя",              # B
    "Показы",              # C
    "Показы (рекл.)",      # D
    "Клики",               # E
    "Клики (рекл.)",       # F
    "Корзина",             # G
    "Заказы",              # H
    "Заказы (рекл.)",      # I
    "Выручка, ₽",          # J
    "Расход рекл., ₽",     # K
    "CTR, %",              # L — formula
    "CR клик→заказ, %",    # M — formula
    "ДРР, %",              # N — formula
    "ROAS",                # O — formula
]
sum_widths = [18, 12, 12, 14, 10, 14, 10, 10, 14, 14, 16, 10, 16, 10, 8]

write_header(ws_sum, sum_headers, sum_widths)

sorted_keys = sorted(summary.keys(), key=lambda x: (x[0], x[1]))

for ri, key in enumerate(sorted_keys, 2):
    r = ri
    cat, week = key
    s = summary[key]

    ws_sum.cell(row=r, column=1, value=cat).border = thin_border
    ws_sum.cell(row=r, column=2, value=week).number_format = date_fmt
    ws_sum.cell(row=r, column=2).border = thin_border

    # C–K: значения
    vals = [
        s["impr"], s["impr_ad"], s["clicks"], s["clicks_ad"],
        s["atc"], s["orders"], s["orders_ad"], s["revenue"], s["ad_spend"],
    ]
    for ci, v in enumerate(vals, 3):
        cell = ws_sum.cell(row=r, column=ci, value=v)
        cell.border = thin_border
        if ci in (10, 11):
            cell.number_format = num_fmt_rub
        else:
            cell.number_format = num_fmt_int

    # L: CTR = клики / показы
    c = ws_sum.cell(row=r, column=12)
    c.value = f"=IF(C{r}=0,0,E{r}/C{r})"
    c.number_format = num_fmt_pct
    c.border = thin_border

    # M: CR = заказы / клики
    c = ws_sum.cell(row=r, column=13)
    c.value = f"=IF(E{r}=0,0,H{r}/E{r})"
    c.number_format = num_fmt_pct
    c.border = thin_border

    # N: ДРР = расход / выручка
    c = ws_sum.cell(row=r, column=14)
    c.value = f"=IF(J{r}=0,0,K{r}/J{r})"
    c.number_format = num_fmt_pct
    c.border = thin_border

    # O: ROAS = выручка / расход
    c = ws_sum.cell(row=r, column=15)
    c.value = f"=IF(K{r}=0,0,J{r}/K{r})"
    c.number_format = "0.0"
    c.border = thin_border

ws_sum.auto_filter.ref = f"A1:O{len(sorted_keys) + 1}"
ws_sum.freeze_panes = "A2"


# ============================================================
# 10. СОХРАНЕНИЕ И ВЕРИФИКАЦИЯ
# ============================================================
output_path = "/Users/vadimbakanov/Documents/_CODE/vibecommerce_test_code/PRJ_MARKETPLACE/ads_data_v1.0.xlsx"
wb.save(output_path)

print(f"✓ Файл сохранён: {output_path}")
print(f"  Каталог: {len(products)} SKU")
print(f"  Трафик и реклама: {len(ads_rows)} строк")
print(f"  Сводка: {len(sorted_keys)} строк (категория × неделя)")

# --- Верификация мэтчинга с sales_data ---
mismatches = 0
for i, (sr, ar) in enumerate(zip(sales_rows, ads_rows)):
    if sr["qty"] != ar["orders_total"]:
        print(f"  ⚠ MISMATCH строка {i}: sales qty={sr['qty']} != ads orders={ar['orders_total']}")
        mismatches += 1
    if sr["ad_spend"] != ar["ad_spend"]:
        print(f"  ⚠ MISMATCH строка {i}: sales ad_spend={sr['ad_spend']} != ads ad_spend={ar['ad_spend']}")
        mismatches += 1
    if sr["revenue"] != ar["revenue"]:
        print(f"  ⚠ MISMATCH строка {i}: sales revenue={sr['revenue']} != ads revenue={ar['revenue']}")
        mismatches += 1

if mismatches == 0:
    print("  ✓ Все данные мэтчатся с sales_data_v1.0.xlsx (qty, revenue, ad_spend)")
else:
    print(f"  ✗ Найдено {mismatches} несовпадений!")
