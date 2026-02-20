#!/usr/bin/env python3
"""Analyze a product (SKU) by ID on Ozon, Wildberries, or Yandex Market.

Fetches item data from MPStats API, optionally builds niche context,
and outputs a formatted report to console or Excel.

Usage:
    uv run --with httpx,python-dotenv PRJ_ANALYTICS/mpstats/analyze_sku.py 123456789
    uv run --with httpx,python-dotenv PRJ_ANALYTICS/mpstats/analyze_sku.py 123456789 --platform wb
    uv run --with httpx,python-dotenv,openpyxl PRJ_ANALYTICS/mpstats/analyze_sku.py 123456789 --xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from client import MPStatsClient, create_client
from models import ItemSummary, NicheContext, PLATFORM_NAMES, _fmt


def build_niche_context(
    client: MPStatsClient,
    category_path: str,
    d1: str,
    d2: str,
    platform: str,
) -> NicheContext | None:
    """Build niche context from category API data."""
    by_date = client.get_category_by_date(category_path, d1, d2, platform)
    sellers = client.get_category_sellers(category_path, d1, d2, platform)
    brands = client.get_category_brands(category_path, d1, d2, platform)

    if not by_date:
        return None

    # Calculate totals from by_date time series
    total_revenue = sum(d.get("revenue", 0) for d in by_date)
    total_sales = sum(d.get("sales", 0) for d in by_date)
    avg_check = total_revenue / total_sales if total_sales > 0 else 0

    # Trend: compare last 7 days vs previous 7 days
    if len(by_date) >= 14:
        recent = by_date[-7:]
        prev = by_date[-14:-7]
        avg_recent = sum(d.get("revenue", 0) for d in recent) / 7
        avg_prev = sum(d.get("revenue", 0) for d in prev) / 7
        trend_pct = ((avg_recent - avg_prev) / avg_prev * 100) if avg_prev > 0 else 0
    else:
        trend_pct = 0.0

    if trend_pct > 5:
        trend_label = "up"
    elif trend_pct < -5:
        trend_label = "down"
    else:
        trend_label = "stable"

    return NicheContext(
        platform=platform,
        category_name=category_path.split("/")[-1] if "/" in category_path else category_path,
        total_revenue=total_revenue,
        total_sales=total_sales,
        avg_check=avg_check,
        sellers_count=len(sellers) if sellers else 0,
        brands_count=len(brands) if brands else 0,
        trend_pct=trend_pct,
        trend_label=trend_label,
    )


def export_to_xlsx(item: ItemSummary, niche: NicheContext | None, filepath: str) -> None:
    """Export SKU analysis to Excel file."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKU Analysis"

    # Header
    ws["A1"] = "Анализ товара"
    ws["A1"].font = Font(bold=True, size=14)

    # Item data
    rows = [
        ("Маркетплейс", PLATFORM_NAMES.get(item.platform, item.platform)),
        ("ID", item.id),
        ("Название", item.name),
        ("Бренд", item.brand),
        ("Продавец", item.seller),
        ("Категория", item.category),
        ("Цена", item.final_price),
        ("Цена без скидки", item.price),
        ("Скидка %", item.discount),
        ("Рейтинг", round(item.rating, 1)),
        ("Отзывов", item.reviews_count),
        ("Остаток", item.balance),
    ]
    if item.delivery_scheme:
        rows.append(("Схема доставки", item.delivery_scheme))

    for i, (label, value) in enumerate(rows, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value

    # Niche context
    if niche:
        row = len(rows) + 5
        ws[f"A{row}"] = "Контекст ниши"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        niche_rows = [
            ("Категория", niche.category_name),
            ("Выручка категории", niche.total_revenue),
            ("Продажи", niche.total_sales),
            ("Ср. чек", round(niche.avg_check, 2)),
            ("Продавцов", niche.sellers_count),
            ("Брендов", niche.brands_count),
            ("Тренд %", round(niche.trend_pct, 1)),
            ("Тренд", niche.trend_label),
        ]
        for j, (label, value) in enumerate(niche_rows, start=row + 1):
            ws[f"A{j}"] = label
            ws[f"A{j}"].font = Font(bold=True)
            ws[f"B{j}"] = value

    # Auto-width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(filepath)


def main() -> None:
    parser = argparse.ArgumentParser(description="Анализ товара по ID через MPStats API")
    parser.add_argument("item_id", type=int, help="ID товара")
    parser.add_argument("--platform", choices=["oz", "wb", "ym"], default="oz", help="Маркетплейс (default: oz)")
    parser.add_argument("--niche", action="store_true", help="Добавить контекст ниши (доп. запросы к API)")
    parser.add_argument("--xlsx", metavar="FILE", help="Экспорт в Excel (путь к файлу)")
    args = parser.parse_args()

    try:
        client = create_client()
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    mp = PLATFORM_NAMES.get(args.platform, args.platform)
    print(f"Загрузка данных товара {args.item_id} с {mp}...")

    with client:
        raw = client.get_item(args.item_id, platform=args.platform)
        if raw is None:
            print(f"ERROR: Товар {args.item_id} не найден на {mp}.")
            sys.exit(1)

        item = ItemSummary.from_api(raw, platform=args.platform)
        print(item.format())

        # Niche context
        niche = None
        if args.niche and item.category:
            d2 = datetime.now().strftime("%Y-%m-%d")
            d1 = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            print(f"\nЗагрузка контекста ниши: {item.category}...")
            niche = build_niche_context(client, item.category, d1, d2, args.platform)
            if niche:
                print(niche.format())
            else:
                print("Не удалось получить контекст ниши.")

    # Excel export
    if args.xlsx:
        export_to_xlsx(item, niche, args.xlsx)
        print(f"\nExcel-отчёт сохранён: {args.xlsx}")


if __name__ == "__main__":
    main()
