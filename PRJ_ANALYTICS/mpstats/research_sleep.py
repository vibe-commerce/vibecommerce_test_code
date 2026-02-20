#!/usr/bin/env python3
"""Market research: Sleep products across Ozon, Wildberries, Yandex Market.

Phase 1: Discover sleep-related categories from MPStats category trees.
Phase 2: Analyze each category (revenue, trends, top products, sellers, brands).
Phase 3: Export consolidated report to Excel.

Usage:
    # Phase 1 — discover categories
    uv run --with httpx,python-dotenv PRJ_ANALYTICS/mpstats/research_sleep.py --discover

    # Phase 2 — full research (discover + analyze + Excel)
    uv run --with httpx,python-dotenv,openpyxl PRJ_ANALYTICS/mpstats/research_sleep.py

    # Analyze specific platforms only
    uv run --with httpx,python-dotenv,openpyxl PRJ_ANALYTICS/mpstats/research_sleep.py --platforms wb oz
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from client import MPStatsClient, create_client

# ── Sleep-related keywords for category search ──────────────────────

SLEEP_KEYWORDS = [
    "сон", "сна", "сну",
    "подушк",           # подушка, подушки
    "одеял",            # одеяло, одеяла
    "матрас", "матрац", # матрас, наматрасник
    "наматрасник",
    "постельн",         # постельное, постельный
    "простын",          # простыня, простыни
    "наволоч",          # наволочка, наволочки
    "пододеяльник",
    "покрывал",         # покрывало
    "плед",
    "пижам",            # пижама
    "ночн",             # ночная рубашка, ночник
    "маска для сна",
    "беруш",            # беруши
    "спальн",           # спальный, спальня
]

PLATFORM_NAMES = {"oz": "Ozon", "wb": "Wildberries", "ym": "Яндекс Маркет"}

# ── Manually curated sleep product categories per platform ──────────
# Auto-discovery produces too much noise; these are the real categories.

SLEEP_CATEGORIES = {
    "wb": [
        "Дом/Спальня/Постельные принадлежности/Подушки",
        "Дом/Спальня/Постельные принадлежности/Одеяла",
        "Дом/Спальня/Постельные принадлежности/Постельное белье",
        "Дом/Спальня/Постельные принадлежности/Наматрасники",
        "Дом/Спальня/Постельные принадлежности/Наволочки",
        "Дом/Спальня/Постельные принадлежности/Пододеяльники",
        "Дом/Спальня/Постельные принадлежности/Простыни",
        "Дом/Спальня/Постельные принадлежности/Пледы",
        "Дом/Спальня/Постельные принадлежности/Покрывала",
        "Дом/Спальня/Мебель/Матрасы",
        "Дом/Спальня/Постельные принадлежности/Матрасы-топперы",
    ],
    "oz": [
        "Дом и сад/Текстиль/Подушки",
        "Дом и сад/Текстиль/Одеяла",
        "Дом и сад/Текстиль/Постельное белье",
        "Дом и сад/Текстиль/Наматрасники и чехлы для матрасов",
        "Дом и сад/Текстиль/Пледы и покрывала",
        "Мебель/Мебель для спальни и комплектующие/Матрасы",
    ],
    "ym": [
        # YM API returns 500 for categories — not available yet
    ],
}


def analyze_category(
    client: MPStatsClient,
    category_path: str,
    d1: str,
    d2: str,
    platform: str,
) -> dict:
    """Analyze a single category — revenue, sales, trends, top items."""
    result = {
        "category": category_path,
        "platform": platform,
        "platform_name": PLATFORM_NAMES.get(platform, platform),
        "period": f"{d1} — {d2}",
    }

    # 1. Category by_date (trends)
    print(f"    by_date...", end=" ", flush=True)
    by_date = client.get_category_by_date(category_path, d1, d2, platform)
    if by_date:
        total_revenue = sum(d.get("revenue", 0) for d in by_date)
        total_sales = sum(d.get("sales", 0) for d in by_date)
        avg_check = total_revenue / total_sales if total_sales > 0 else 0

        # Week-over-week trend
        trend_pct = 0.0
        if len(by_date) >= 14:
            recent = by_date[-7:]
            prev = by_date[-14:-7]
            avg_r = sum(d.get("revenue", 0) for d in recent) / 7
            avg_p = sum(d.get("revenue", 0) for d in prev) / 7
            trend_pct = ((avg_r - avg_p) / avg_p * 100) if avg_p > 0 else 0

        result.update({
            "total_revenue": total_revenue,
            "total_sales": total_sales,
            "avg_check": round(avg_check, 2),
            "revenue_trend_pct": round(trend_pct, 1),
            "daily_data_points": len(by_date),
        })
        print(f"OK ({len(by_date)} дней)")
    else:
        print("нет данных")
        result.update({"total_revenue": 0, "total_sales": 0, "avg_check": 0, "revenue_trend_pct": 0})

    # 2. Top products
    print(f"    products...", end=" ", flush=True)
    products_raw = client.get_category_products(category_path, d1, d2, platform, end_row=50)
    top_products = []
    products_count = 0
    if products_raw and isinstance(products_raw, dict):
        rows = products_raw.get("data", [])
        products_count = products_raw.get("total", len(rows))
        for p in rows[:30]:
            top_products.append({
                "id": p.get("id", 0),
                "name": p.get("name", ""),
                "revenue": p.get("revenue", 0),
                "sales": p.get("sales", 0),
                "final_price": p.get("final_price", 0),
                "brand": p.get("brand", ""),
                "seller": p.get("seller", ""),
                "rating": p.get("rating", 0),
                "comments": p.get("comments", 0) or p.get("feedbacks", 0),
            })
        print(f"OK ({products_count} всего, топ-{len(top_products)} загружено)")
    else:
        print("нет данных")
    result["products_count"] = products_count
    result["top_products"] = top_products

    # 3. Top sellers
    print(f"    sellers...", end=" ", flush=True)
    sellers = client.get_category_sellers(category_path, d1, d2, platform)
    top_sellers = []
    if sellers:
        sorted_s = sorted(sellers, key=lambda s: s.get("revenue", 0), reverse=True)
        total_s_rev = sum(s.get("revenue", 0) for s in sorted_s)
        top5_rev = sum(s.get("revenue", 0) for s in sorted_s[:5])
        result["sellers_count"] = len(sellers)
        result["top5_sellers_pct"] = round((top5_rev / total_s_rev * 100) if total_s_rev > 0 else 0, 1)
        for s in sorted_s[:20]:
            top_sellers.append({
                "name": s.get("name", ""),
                "revenue": s.get("revenue", 0),
                "sales": s.get("sales", 0),
                "items": s.get("items", 0),
            })
        print(f"OK ({len(sellers)} продавцов)")
    else:
        print("нет данных")
        result["sellers_count"] = 0
        result["top5_sellers_pct"] = 0
    result["top_sellers"] = top_sellers

    # 4. Top brands
    print(f"    brands...", end=" ", flush=True)
    brands = client.get_category_brands(category_path, d1, d2, platform)
    top_brands = []
    if brands:
        sorted_b = sorted(brands, key=lambda b: b.get("revenue", 0), reverse=True)
        total_b_rev = sum(b.get("revenue", 0) for b in sorted_b)
        top5_b_rev = sum(b.get("revenue", 0) for b in sorted_b[:5])
        result["brands_count"] = len(brands)
        result["top5_brands_pct"] = round((top5_b_rev / total_b_rev * 100) if total_b_rev > 0 else 0, 1)
        for b in sorted_b[:20]:
            top_brands.append({
                "name": b.get("name", ""),
                "revenue": b.get("revenue", 0),
                "sales": b.get("sales", 0),
                "items": b.get("items", 0),
            })
        print(f"OK ({len(brands)} брендов)")
    else:
        print("нет данных")
        result["brands_count"] = 0
        result["top5_brands_pct"] = 0
    result["top_brands"] = top_brands

    return result


def fmt(n) -> str:
    """Format number with space separator."""
    if isinstance(n, float):
        if n == int(n):
            n = int(n)
        else:
            return f"{n:,.2f}".replace(",", " ")
    return f"{n:,}".replace(",", " ")


def print_report(all_results: list[dict]) -> None:
    """Print consolidated text report."""
    print()
    print("=" * 70)
    print("  ИССЛЕДОВАНИЕ РЫНКА: ТОВАРЫ ДЛЯ СНА")
    print("=" * 70)

    for r in all_results:
        mp = r.get("platform_name", r.get("platform", "?"))
        cat = r.get("category", "?")
        print(f"\n{'─' * 70}")
        print(f"  {cat} ({mp})")
        print(f"{'─' * 70}")
        print(f"  Период: {r.get('period', '?')}")
        print(f"  Выручка:         {fmt(r.get('total_revenue', 0))} руб.")
        print(f"  Продажи:         {fmt(r.get('total_sales', 0))} шт.")
        print(f"  Ср. чек:         {fmt(r.get('avg_check', 0))} руб.")
        print(f"  Тренд (WoW):     {r.get('revenue_trend_pct', 0):+.1f}%")
        print(f"  Товаров:         {fmt(r.get('products_count', 0))}")
        print(f"  Продавцов:       {fmt(r.get('sellers_count', 0))}")
        print(f"  Брендов:         {fmt(r.get('brands_count', 0))}")
        print(f"  Топ-5 продавцов: {r.get('top5_sellers_pct', 0):.1f}% выручки")
        print(f"  Топ-5 брендов:   {r.get('top5_brands_pct', 0):.1f}% выручки")

        if r.get("top_products"):
            print(f"\n  Топ-10 товаров:")
            for i, p in enumerate(r["top_products"][:10], 1):
                print(f"    {i:2d}. {p['name'][:50]}")
                print(f"        {fmt(p['revenue'])} руб. | {fmt(p['sales'])} шт. | "
                      f"{fmt(p['final_price'])} руб. | {p['brand']}")

        if r.get("top_brands"):
            print(f"\n  Топ-5 брендов:")
            for i, b in enumerate(r["top_brands"][:5], 1):
                print(f"    {i:2d}. {b['name'][:35]:35s}  {fmt(b['revenue']):>12s} руб.")

        if r.get("top_sellers"):
            print(f"\n  Топ-5 продавцов:")
            for i, s in enumerate(r["top_sellers"][:5], 1):
                print(f"    {i:2d}. {s['name'][:35]:35s}  {fmt(s['revenue']):>12s} руб.")


def export_xlsx(all_results: list[dict], filepath: str) -> None:
    """Export research to Excel with multiple sheets."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    money_fmt = '#,##0'
    pct_fmt = '0.0%'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ── Sheet 1: Summary across all platforms ──
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Исследование рынка: Товары для сна"
    ws["A1"].font = title_font
    ws["A2"] = f"Дата: {datetime.now().strftime('%Y-%m-%d')}"

    headers = [
        "Маркетплейс", "Категория", "Выручка (руб)", "Продажи (шт)",
        "Ср. чек (руб)", "Тренд WoW %", "Товаров", "Продавцов",
        "Брендов", "Топ-5 продавцов %", "Топ-5 брендов %",
    ]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = header_font
        cell.border = thin_border

    for i, r in enumerate(all_results, 5):
        ws.cell(row=i, column=1, value=r.get("platform_name", ""))
        ws.cell(row=i, column=2, value=r.get("category", ""))
        ws.cell(row=i, column=3, value=r.get("total_revenue", 0)).number_format = money_fmt
        ws.cell(row=i, column=4, value=r.get("total_sales", 0)).number_format = money_fmt
        ws.cell(row=i, column=5, value=r.get("avg_check", 0)).number_format = '#,##0.00'
        ws.cell(row=i, column=6, value=r.get("revenue_trend_pct", 0))
        ws.cell(row=i, column=7, value=r.get("products_count", 0))
        ws.cell(row=i, column=8, value=r.get("sellers_count", 0))
        ws.cell(row=i, column=9, value=r.get("brands_count", 0))
        ws.cell(row=i, column=10, value=r.get("top5_sellers_pct", 0))
        ws.cell(row=i, column=11, value=r.get("top5_brands_pct", 0))
        for j in range(1, 12):
            ws.cell(row=i, column=j).border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # ── Sheet 2: All top products ──
    ws2 = wb.create_sheet("Топ товары")
    prod_headers = ["Маркетплейс", "Категория", "#", "Название", "Выручка", "Продажи",
                    "Цена", "Бренд", "Продавец", "Рейтинг", "Отзывов"]
    for j, h in enumerate(prod_headers, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.border = thin_border

    row = 2
    for r in all_results:
        mp = r.get("platform_name", "")
        cat = r.get("category", "")
        for i, p in enumerate(r.get("top_products", [])[:20], 1):
            ws2.cell(row=row, column=1, value=mp)
            ws2.cell(row=row, column=2, value=cat)
            ws2.cell(row=row, column=3, value=i)
            ws2.cell(row=row, column=4, value=p.get("name", ""))
            ws2.cell(row=row, column=5, value=p.get("revenue", 0)).number_format = money_fmt
            ws2.cell(row=row, column=6, value=p.get("sales", 0)).number_format = money_fmt
            ws2.cell(row=row, column=7, value=p.get("final_price", 0)).number_format = money_fmt
            ws2.cell(row=row, column=8, value=p.get("brand", ""))
            ws2.cell(row=row, column=9, value=p.get("seller", ""))
            ws2.cell(row=row, column=10, value=p.get("rating", 0))
            ws2.cell(row=row, column=11, value=p.get("comments", 0))
            row += 1

    # ── Sheet 3: All top brands ──
    ws3 = wb.create_sheet("Топ бренды")
    brand_headers = ["Маркетплейс", "Категория", "#", "Бренд", "Выручка", "Продажи", "Товаров"]
    for j, h in enumerate(brand_headers, 1):
        ws3.cell(row=1, column=j, value=h).font = header_font

    row = 2
    for r in all_results:
        mp = r.get("platform_name", "")
        cat = r.get("category", "")
        for i, b in enumerate(r.get("top_brands", [])[:20], 1):
            ws3.cell(row=row, column=1, value=mp)
            ws3.cell(row=row, column=2, value=cat)
            ws3.cell(row=row, column=3, value=i)
            ws3.cell(row=row, column=4, value=b.get("name", ""))
            ws3.cell(row=row, column=5, value=b.get("revenue", 0)).number_format = money_fmt
            ws3.cell(row=row, column=6, value=b.get("sales", 0)).number_format = money_fmt
            ws3.cell(row=row, column=7, value=b.get("items", 0))
            row += 1

    # ── Sheet 4: All top sellers ──
    ws4 = wb.create_sheet("Топ продавцы")
    seller_headers = ["Маркетплейс", "Категория", "#", "Продавец", "Выручка", "Продажи", "Товаров"]
    for j, h in enumerate(seller_headers, 1):
        ws4.cell(row=1, column=j, value=h).font = header_font

    row = 2
    for r in all_results:
        mp = r.get("platform_name", "")
        cat = r.get("category", "")
        for i, s in enumerate(r.get("top_sellers", [])[:20], 1):
            ws4.cell(row=row, column=1, value=mp)
            ws4.cell(row=row, column=2, value=cat)
            ws4.cell(row=row, column=3, value=i)
            ws4.cell(row=row, column=4, value=s.get("name", ""))
            ws4.cell(row=row, column=5, value=s.get("revenue", 0)).number_format = money_fmt
            ws4.cell(row=row, column=6, value=s.get("sales", 0)).number_format = money_fmt
            ws4.cell(row=row, column=7, value=s.get("items", 0))
            row += 1

    wb.save(filepath)


def main() -> None:
    parser = argparse.ArgumentParser(description="Market research: sleep products")
    parser.add_argument("--discover", action="store_true", help="Only discover categories, don't analyze")
    parser.add_argument("--platforms", nargs="+", choices=["oz", "wb", "ym"], default=["wb", "oz", "ym"],
                        help="Platforms to research (default: all)")
    parser.add_argument("--days", type=int, default=30, help="Analysis period in days (default: 30)")
    parser.add_argument("--xlsx", metavar="FILE", help="Excel output path",
                        default="PRJ_ANALYTICS/mpstats/sleep_market_research.xlsx")
    parser.add_argument("--no-xlsx", action="store_true", help="Skip Excel export")
    args = parser.parse_args()

    try:
        client = create_client()
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")

    print("=" * 70)
    print("  ИССЛЕДОВАНИЕ РЫНКА: ТОВАРЫ ДЛЯ СНА")
    print(f"  Маркетплейсы: {', '.join(PLATFORM_NAMES[p] for p in args.platforms)}")
    print(f"  Период: {d1} — {d2} ({args.days} дней)")
    print("=" * 70)

    # ── Phase 1: List categories ──
    print("\n[Phase 1] Категории для анализа:")
    all_sleep_categories = []
    for platform in args.platforms:
        mp = PLATFORM_NAMES[platform]
        cats = SLEEP_CATEGORIES.get(platform, [])
        if not cats:
            print(f"  {mp}: нет категорий (API недоступен)")
            continue
        print(f"  {mp}: {len(cats)} категорий")
        for c in cats:
            all_sleep_categories.append({"path": c, "platform": platform})
            print(f"    - {c}")

    if not all_sleep_categories:
        print("\nНет категорий для анализа. Завершение.")
        sys.exit(0)

    total_cats = len(all_sleep_categories)
    print(f"\nИтого: {total_cats} категорий")
    # Each category = ~4 API calls → estimate usage
    est_calls = total_cats * 4
    print(f"Ожидаемое количество API-вызовов: ~{est_calls}")

    if args.discover:
        print("\n(--discover mode: анализ не запускается)")
        return

    # ── Phase 2: Analyze each category ──
    print(f"\n[Phase 2] Анализ категорий (по {args.days} дней)...")
    all_results = []

    with client:
        for i, cat_info in enumerate(all_sleep_categories, 1):
            platform = cat_info["platform"]
            path = cat_info["path"]
            mp = PLATFORM_NAMES[platform]
            print(f"\n  [{i}/{total_cats}] {path} ({mp})")

            result = analyze_category(client, path, d1, d2, platform)
            all_results.append(result)

            # Small delay to avoid rate limits
            time.sleep(0.5)

    # ── Phase 3: Report ──
    # Filter out categories with no data
    active_results = [r for r in all_results if r.get("total_revenue", 0) > 0]
    empty_results = [r for r in all_results if r.get("total_revenue", 0) == 0]

    if empty_results:
        print(f"\n  ({len(empty_results)} категорий без данных — пропущены)")

    # Sort by revenue descending
    active_results.sort(key=lambda r: r.get("total_revenue", 0), reverse=True)

    print_report(active_results)

    # ── Phase 4: Excel export ──
    if not args.no_xlsx and active_results:
        print(f"\n[Phase 4] Экспорт в Excel: {args.xlsx}")
        export_xlsx(active_results, args.xlsx)
        print(f"  Файл сохранён: {args.xlsx}")

    # Save raw JSON for further analysis
    json_path = args.xlsx.replace(".xlsx", ".json") if args.xlsx else "sleep_research.json"
    with open(json_path, "w", encoding="utf-8") as f:
        # Remove top_products etc for cleaner JSON summary
        summary = []
        for r in active_results:
            s = {k: v for k, v in r.items() if k not in ("top_products", "top_sellers", "top_brands")}
            s["top_products_count"] = len(r.get("top_products", []))
            s["top_sellers_count"] = len(r.get("top_sellers", []))
            s["top_brands_count"] = len(r.get("top_brands", []))
            summary.append(s)
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  JSON-сводка: {json_path}")


if __name__ == "__main__":
    main()
