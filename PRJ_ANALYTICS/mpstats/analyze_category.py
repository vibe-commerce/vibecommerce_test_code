#!/usr/bin/env python3
"""Analyze a category on Ozon, Wildberries, or Yandex Market.

Fetches category data (products, sellers, brands, trends) from MPStats API
and outputs a formatted report to console or Excel.

Usage:
    uv run --with httpx,python-dotenv PRJ_ANALYTICS/mpstats/analyze_category.py "Электроника/Смартфоны"
    uv run --with httpx,python-dotenv PRJ_ANALYTICS/mpstats/analyze_category.py "Электроника" --platform wb
    uv run --with httpx,python-dotenv,openpyxl PRJ_ANALYTICS/mpstats/analyze_category.py "Электроника" --xlsx report.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from client import MPStatsClient, create_client
from models import CategoryMetrics, CategoryProduct, PLATFORM_NAMES, _fmt


def analyze(
    client: MPStatsClient,
    category_path: str,
    d1: str,
    d2: str,
    platform: str,
) -> CategoryMetrics:
    """Build category metrics from multiple API calls."""
    # Fetch data
    by_date = client.get_category_by_date(category_path, d1, d2, platform)
    products_raw = client.get_category_products(category_path, d1, d2, platform, end_row=200)
    sellers = client.get_category_sellers(category_path, d1, d2, platform)
    brands = client.get_category_brands(category_path, d1, d2, platform)

    # Revenue and sales from by_date
    total_revenue = 0.0
    total_sales = 0
    if by_date:
        total_revenue = sum(d.get("revenue", 0) for d in by_date)
        total_sales = sum(d.get("sales", 0) for d in by_date)

    avg_check = total_revenue / total_sales if total_sales > 0 else 0

    # Trend
    revenue_trend_pct = 0.0
    if by_date and len(by_date) >= 14:
        recent = by_date[-7:]
        prev = by_date[-14:-7]
        avg_recent = sum(d.get("revenue", 0) for d in recent) / 7
        avg_prev = sum(d.get("revenue", 0) for d in prev) / 7
        revenue_trend_pct = ((avg_recent - avg_prev) / avg_prev * 100) if avg_prev > 0 else 0

    if revenue_trend_pct > 5:
        revenue_trend = "up"
    elif revenue_trend_pct < -5:
        revenue_trend = "down"
    else:
        revenue_trend = "stable"

    # Products
    products_count = 0
    top_products = []
    if products_raw and isinstance(products_raw, dict):
        rows = products_raw.get("data", [])
        products_count = products_raw.get("total", len(rows))
        for p in rows[:10]:
            top_products.append({
                "name": p.get("name", ""),
                "revenue": p.get("revenue", 0),
                "sales": p.get("sales", 0),
                "final_price": p.get("final_price", 0),
                "brand": p.get("brand", ""),
                "seller": p.get("seller", ""),
            })

    # Sellers concentration
    sellers_count = len(sellers) if sellers else 0
    top_sellers = []
    top5_sellers_revenue_pct = 0.0
    if sellers:
        sorted_sellers = sorted(sellers, key=lambda s: s.get("revenue", 0), reverse=True)
        total_sellers_revenue = sum(s.get("revenue", 0) for s in sorted_sellers)
        top5_revenue = sum(s.get("revenue", 0) for s in sorted_sellers[:5])
        top5_sellers_revenue_pct = (top5_revenue / total_sellers_revenue * 100) if total_sellers_revenue > 0 else 0
        for s in sorted_sellers[:10]:
            top_sellers.append({
                "name": s.get("name", ""),
                "revenue": s.get("revenue", 0),
                "sales": s.get("sales", 0),
                "items": s.get("items", 0),
            })

    # Brands concentration
    brands_count = len(brands) if brands else 0
    top_brands = []
    top5_brands_revenue_pct = 0.0
    if brands:
        sorted_brands = sorted(brands, key=lambda b: b.get("revenue", 0), reverse=True)
        total_brands_revenue = sum(b.get("revenue", 0) for b in sorted_brands)
        top5_brand_revenue = sum(b.get("revenue", 0) for b in sorted_brands[:5])
        top5_brands_revenue_pct = (top5_brand_revenue / total_brands_revenue * 100) if total_brands_revenue > 0 else 0
        for b in sorted_brands[:10]:
            top_brands.append({
                "name": b.get("name", ""),
                "revenue": b.get("revenue", 0),
                "sales": b.get("sales", 0),
                "items": b.get("items", 0),
            })

    return CategoryMetrics(
        total_revenue=total_revenue,
        total_sales=total_sales,
        avg_check=avg_check,
        products_count=products_count,
        sellers_count=sellers_count,
        brands_count=brands_count,
        revenue_trend_pct=revenue_trend_pct,
        revenue_trend=revenue_trend,
        top5_sellers_revenue_pct=top5_sellers_revenue_pct,
        top5_brands_revenue_pct=top5_brands_revenue_pct,
        top_sellers=top_sellers,
        top_brands=top_brands,
        top_products=top_products,
    )


def export_to_xlsx(
    metrics: CategoryMetrics,
    category_path: str,
    platform: str,
    d1: str,
    d2: str,
    filepath: str,
) -> None:
    """Export category analysis to Excel file."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Сводка"
    mp = PLATFORM_NAMES.get(platform, platform)

    ws["A1"] = f"Анализ категории: {category_path}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Маркетплейс: {mp} | Период: {d1} — {d2}"

    rows = [
        ("Выручка", metrics.total_revenue),
        ("Продажи", metrics.total_sales),
        ("Ср. чек", round(metrics.avg_check, 2)),
        ("Товаров", metrics.products_count),
        ("Продавцов", metrics.sellers_count),
        ("Брендов", metrics.brands_count),
        ("Тренд %", round(metrics.revenue_trend_pct, 1)),
        ("Тренд", metrics.revenue_trend),
        ("Топ-5 продавцов % выручки", round(metrics.top5_sellers_revenue_pct, 1)),
        ("Топ-5 брендов % выручки", round(metrics.top5_brands_revenue_pct, 1)),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value

    # Top products sheet
    if metrics.top_products:
        ws2 = wb.create_sheet("Топ товары")
        headers = ["#", "Название", "Выручка", "Продажи", "Цена", "Бренд", "Продавец"]
        for j, h in enumerate(headers, 1):
            ws2.cell(row=1, column=j, value=h).font = Font(bold=True)
        for i, p in enumerate(metrics.top_products, 1):
            ws2.cell(row=i + 1, column=1, value=i)
            ws2.cell(row=i + 1, column=2, value=p.get("name", ""))
            ws2.cell(row=i + 1, column=3, value=p.get("revenue", 0))
            ws2.cell(row=i + 1, column=4, value=p.get("sales", 0))
            ws2.cell(row=i + 1, column=5, value=p.get("final_price", 0))
            ws2.cell(row=i + 1, column=6, value=p.get("brand", ""))
            ws2.cell(row=i + 1, column=7, value=p.get("seller", ""))

    # Top sellers sheet
    if metrics.top_sellers:
        ws3 = wb.create_sheet("Топ продавцы")
        headers = ["#", "Продавец", "Выручка", "Продажи", "Товаров"]
        for j, h in enumerate(headers, 1):
            ws3.cell(row=1, column=j, value=h).font = Font(bold=True)
        for i, s in enumerate(metrics.top_sellers, 1):
            ws3.cell(row=i + 1, column=1, value=i)
            ws3.cell(row=i + 1, column=2, value=s.get("name", ""))
            ws3.cell(row=i + 1, column=3, value=s.get("revenue", 0))
            ws3.cell(row=i + 1, column=4, value=s.get("sales", 0))
            ws3.cell(row=i + 1, column=5, value=s.get("items", 0))

    # Top brands sheet
    if metrics.top_brands:
        ws4 = wb.create_sheet("Топ бренды")
        headers = ["#", "Бренд", "Выручка", "Продажи", "Товаров"]
        for j, h in enumerate(headers, 1):
            ws4.cell(row=1, column=j, value=h).font = Font(bold=True)
        for i, b in enumerate(metrics.top_brands, 1):
            ws4.cell(row=i + 1, column=1, value=i)
            ws4.cell(row=i + 1, column=2, value=b.get("name", ""))
            ws4.cell(row=i + 1, column=3, value=b.get("revenue", 0))
            ws4.cell(row=i + 1, column=4, value=b.get("sales", 0))
            ws4.cell(row=i + 1, column=5, value=b.get("items", 0))

    wb.save(filepath)


def main() -> None:
    parser = argparse.ArgumentParser(description="Анализ категории через MPStats API")
    parser.add_argument("category", help="Путь категории (напр. 'Электроника/Смартфоны')")
    parser.add_argument("--platform", choices=["oz", "wb", "ym"], default="oz", help="Маркетплейс (default: oz)")
    parser.add_argument("--days", type=int, default=30, help="Период анализа в днях (default: 30)")
    parser.add_argument("--xlsx", metavar="FILE", help="Экспорт в Excel (путь к файлу)")
    args = parser.parse_args()

    try:
        client = create_client()
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    d2 = datetime.now().strftime("%Y-%m-%d")
    d1 = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")
    mp = PLATFORM_NAMES.get(args.platform, args.platform)

    print(f"Анализ категории «{args.category}» на {mp}...")
    print(f"Период: {d1} — {d2} ({args.days} дней)")
    print()

    with client:
        metrics = analyze(client, args.category, d1, d2, args.platform)

    print(metrics.format(args.category, args.platform))

    if args.xlsx:
        export_to_xlsx(metrics, args.category, args.platform, d1, d2, args.xlsx)
        print(f"\nExcel-отчёт сохранён: {args.xlsx}")


if __name__ == "__main__":
    main()
