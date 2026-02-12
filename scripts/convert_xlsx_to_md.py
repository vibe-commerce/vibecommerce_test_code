#!/usr/bin/env python3
"""Конвертация Excel → Markdown таблицу.

Использование:
    uv run --with openpyxl scripts/convert_xlsx_to_md.py файл.xlsx
    uv run --with openpyxl scripts/convert_xlsx_to_md.py файл.xlsx -o output.md
    uv run --with openpyxl scripts/convert_xlsx_to_md.py файл.xlsx --sheet "Лист1"
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def xlsx_to_markdown(filepath: str, sheet_name: str | None = None) -> str:
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return "_Пустой лист_"

    # Header
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    md = "| " + " | ".join(headers) + " |\n"
    md += "| " + " | ".join(["---"] * len(headers)) + " |\n"

    # Data rows
    for row in rows[1:]:
        cells = [str(cell) if cell is not None else "" for cell in row]
        # Pad if row is shorter than header
        while len(cells) < len(headers):
            cells.append("")
        md += "| " + " | ".join(cells) + " |\n"

    wb.close()
    return md


def main():
    parser = argparse.ArgumentParser(description="Excel → Markdown")
    parser.add_argument("file", help="Путь к .xlsx файлу")
    parser.add_argument("-o", "--output", help="Выходной .md файл (по умолчанию — stdout)")
    parser.add_argument("--sheet", help="Имя листа (по умолчанию — активный)")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"Файл не найден: {args.file}", file=sys.stderr)
        sys.exit(1)

    result = xlsx_to_markdown(args.file, args.sheet)

    if args.output:
        Path(args.output).write_text(result, encoding="utf-8")
        print(f"Сохранено: {args.output}")
    else:
        print(result)


if __name__ == "__main__":
    main()
