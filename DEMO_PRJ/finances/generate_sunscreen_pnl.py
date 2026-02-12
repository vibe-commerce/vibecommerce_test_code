#!/usr/bin/env python3
"""
Генератор P&L модели для солнцезащитных кремов.
Запуск: uv run --with openpyxl DEMO_PRJ/finances/generate_sunscreen_pnl.py
"""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================================
# ПАРАМЕТРЫ МОДЕЛИ (можно менять)
# ============================================================================

# Старт продаж
START_YEAR = 2026
START_MONTH = 4  # Апрель — успеть к сезону
MONTHS_TOTAL = 36

# SKU матрица
SKUS = [
    {"name": "SPF50+ Крем для лица 50мл", "price": 1490, "cost": 180, "share": 0.35},
    {"name": "SPF50+ Спрей для тела 150мл", "price": 1290, "cost": 160, "share": 0.40},
    {"name": "SPF50+ Молочко для тела 200мл", "price": 1590, "cost": 200, "share": 0.25},
]

# Средневзвешенные значения (рассчитываются автоматически)
AVG_PRICE = sum(s["price"] * s["share"] for s in SKUS)
AVG_COST = sum(s["cost"] * s["share"] for s in SKUS)

# Каналы продаж
CHANNELS = {
    "ozon": {"share_y1": 0.45, "share_y2": 0.40, "share_y3": 0.35, "commission": 0.18},
    "wb": {"share_y1": 0.45, "share_y2": 0.45, "share_y3": 0.45, "commission": 0.15},
    "site": {"share_y1": 0.10, "share_y2": 0.15, "share_y3": 0.20, "commission": 0.025},
}

# Сезонные коэффициенты (среднее = 1.0)
SEASONALITY = {
    1: 0.30,   # Январь
    2: 0.40,   # Февраль
    3: 0.60,   # Март
    4: 0.90,   # Апрель
    5: 1.40,   # Май
    6: 1.80,   # Июнь
    7: 2.00,   # Июль
    8: 1.60,   # Август
    9: 0.80,   # Сентябрь
    10: 0.40,  # Октябрь
    11: 0.30,  # Ноябрь
    12: 0.30,  # Декабрь
}

# Базовые продажи (штук в месяц при коэфф. сезонности = 1.0)
BASE_SALES_START = 150  # Начальный объём
BASE_SALES_GROWTH_Y1 = 1.12  # Рост м/м в первый год
BASE_SALES_GROWTH_Y2 = 1.08  # Рост м/м во второй год
BASE_SALES_GROWTH_Y3 = 1.04  # Рост м/м в третий год

# Маркетинг (% от выручки)
MARKETING_PCT_Y1 = 0.35
MARKETING_PCT_Y2 = 0.25
MARKETING_PCT_Y3 = 0.18

# Логистика (₽ за единицу)
LOGISTICS_PER_UNIT = 120

# Постоянные расходы (₽/мес)
FIXED_COSTS_Y1 = 80000   # ЗП, подписки, хранение
FIXED_COSTS_Y2 = 100000
FIXED_COSTS_Y3 = 120000

# Единоразовые расходы (месяц 1)
ONETIME_BRAND = 300000      # Бренд/упаковка
ONETIME_CERT = 150000       # Сертификация
ONETIME_CONTENT = 150000    # Фото/видео

# Инвестиции
INVESTMENT = 3_000_000

# ============================================================================
# ГЕНЕРАЦИЯ МОДЕЛИ
# ============================================================================

def get_year_number(month_idx):
    """Возвращает номер года (1, 2 или 3) по индексу месяца."""
    if month_idx < 12:
        return 1
    elif month_idx < 24:
        return 2
    else:
        return 3

def get_channel_share(channel, year_num):
    """Возвращает долю канала для заданного года."""
    ch = CHANNELS[channel]
    if year_num == 1:
        return ch["share_y1"]
    elif year_num == 2:
        return ch["share_y2"]
    else:
        return ch["share_y3"]

def get_marketing_pct(year_num):
    """Возвращает % маркетинга для заданного года."""
    if year_num == 1:
        return MARKETING_PCT_Y1
    elif year_num == 2:
        return MARKETING_PCT_Y2
    else:
        return MARKETING_PCT_Y3

def get_fixed_costs(year_num):
    """Возвращает постоянные расходы для заданного года."""
    if year_num == 1:
        return FIXED_COSTS_Y1
    elif year_num == 2:
        return FIXED_COSTS_Y2
    else:
        return FIXED_COSTS_Y3

def get_growth_rate(year_num):
    """Возвращает коэффициент роста м/м для заданного года."""
    if year_num == 1:
        return BASE_SALES_GROWTH_Y1
    elif year_num == 2:
        return BASE_SALES_GROWTH_Y2
    else:
        return BASE_SALES_GROWTH_Y3

def generate_pnl():
    """Генерирует данные P&L."""
    data = []
    base_sales = BASE_SALES_START
    cumulative_ebitda = -INVESTMENT  # Начинаем с минуса (инвестиции)

    for month_idx in range(MONTHS_TOTAL):
        # Дата
        year = START_YEAR + (START_MONTH - 1 + month_idx) // 12
        month = (START_MONTH - 1 + month_idx) % 12 + 1
        dt = date(year, month, 1)

        year_num = get_year_number(month_idx)

        # Сезонность
        seasonality = SEASONALITY[month]

        # Объём продаж
        if month_idx > 0:
            growth = get_growth_rate(year_num)
            base_sales *= growth

        volume = int(base_sales * seasonality)

        # Выручка
        revenue = volume * AVG_PRICE

        # Себестоимость
        cogs = volume * AVG_COST

        # Валовая прибыль
        gross_profit = revenue - cogs
        gross_margin = gross_profit / revenue if revenue > 0 else 0

        # Комиссии МП (средневзвешенная)
        avg_commission = sum(
            CHANNELS[ch]["commission"] * get_channel_share(ch, year_num)
            for ch in CHANNELS
        )
        mp_commission = revenue * avg_commission

        # Логистика
        logistics = volume * LOGISTICS_PER_UNIT

        # Маркетинг
        marketing = revenue * get_marketing_pct(year_num)

        # Постоянные расходы
        fixed = get_fixed_costs(year_num)

        # Единоразовые (только в первый месяц)
        onetime = 0
        if month_idx == 0:
            onetime = ONETIME_BRAND + ONETIME_CERT + ONETIME_CONTENT

        # EBITDA
        ebitda = gross_profit - mp_commission - logistics - marketing - fixed - onetime
        ebitda_margin = ebitda / revenue if revenue > 0 else 0

        # Накопленный EBITDA
        cumulative_ebitda += ebitda

        data.append({
            "month_num": month_idx + 1,
            "date": dt,
            "year": year,
            "volume": volume,
            "avg_price": AVG_PRICE,
            "revenue": revenue,
            "cogs": cogs,
            "gross_profit": gross_profit,
            "gross_margin": gross_margin,
            "mp_commission": mp_commission,
            "logistics": logistics,
            "marketing": marketing,
            "fixed": fixed,
            "onetime": onetime,
            "ebitda": ebitda,
            "ebitda_margin": ebitda_margin,
            "cumulative_ebitda": cumulative_ebitda,
            "seasonality": seasonality,
        })

    return data

def create_excel(data, filename):
    """Создаёт Excel-файл с P&L."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PnL"

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    money_format = '#,##0'
    pct_format = '0.0%'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовок
    ws["A1"] = "P&L: Солнцезащитные кремы для спортсменов"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Инвестиции: {INVESTMENT:,}₽ | Цель: ROI 200% за 3 года"
    ws["A3"] = f"Средняя цена: {AVG_PRICE:.0f}₽ | Средняя себестоимость: {AVG_COST:.0f}₽"

    # Заголовки столбцов
    headers = [
        ("Месяц", 8),
        ("Дата", 12),
        ("Год", 8),
        ("Сезон.коэфф.", 12),
        ("Объём (шт)", 12),
        ("Ср.цена (₽)", 12),
        ("Выручка (₽)", 14),
        ("Себест. (₽)", 12),
        ("Вал.прибыль (₽)", 14),
        ("Вал.маржа (%)", 12),
        ("Комиссия МП (₽)", 14),
        ("Логистика (₽)", 12),
        ("Маркетинг (₽)", 14),
        ("Пост.расходы (₽)", 14),
        ("Единоразов. (₽)", 14),
        ("EBITDA (₽)", 14),
        ("EBITDA (%)", 12),
        ("EBITDA накопл. (₽)", 16),
    ]

    row = 5
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Данные
    for i, d in enumerate(data):
        row = 6 + i
        values = [
            d["month_num"],
            d["date"],
            d["year"],
            d["seasonality"],
            d["volume"],
            d["avg_price"],
            d["revenue"],
            d["cogs"],
            d["gross_profit"],
            d["gross_margin"],
            d["mp_commission"],
            d["logistics"],
            d["marketing"],
            d["fixed"],
            d["onetime"],
            d["ebitda"],
            d["ebitda_margin"],
            d["cumulative_ebitda"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border

            # Форматирование
            if col == 2:  # Дата
                cell.number_format = "YYYY-MM-DD"
            elif col in [4, 10, 17]:  # Проценты
                cell.number_format = pct_format
            elif col >= 5:  # Числа
                cell.number_format = money_format

            # Подсветка отрицательных
            if isinstance(val, (int, float)) and val < 0:
                cell.font = Font(color="FF0000")

            # Подсветка положительного накопленного EBITDA
            if col == 18 and val > 0:
                cell.font = Font(color="008000", bold=True)

    # Итоги
    row = 6 + len(data) + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = Font(bold=True)

    total_revenue = sum(d["revenue"] for d in data)
    total_cogs = sum(d["cogs"] for d in data)
    total_gross = sum(d["gross_profit"] for d in data)
    total_commission = sum(d["mp_commission"] for d in data)
    total_logistics = sum(d["logistics"] for d in data)
    total_marketing = sum(d["marketing"] for d in data)
    total_fixed = sum(d["fixed"] for d in data)
    total_onetime = sum(d["onetime"] for d in data)
    total_ebitda = sum(d["ebitda"] for d in data)

    totals = [
        (7, total_revenue),
        (8, total_cogs),
        (9, total_gross),
        (11, total_commission),
        (12, total_logistics),
        (13, total_marketing),
        (14, total_fixed),
        (15, total_onetime),
        (16, total_ebitda),
    ]

    for col, val in totals:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.number_format = money_format
        if val < 0:
            cell.font = Font(bold=True, color="FF0000")

    # Метрики
    row += 2
    final_cumulative = data[-1]["cumulative_ebitda"]
    roi = (final_cumulative + INVESTMENT) / INVESTMENT - 1

    ws.cell(row=row, column=1, value="КЛЮЧЕВЫЕ МЕТРИКИ").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value=f"Инвестиции:")
    ws.cell(row=row, column=2, value=INVESTMENT).number_format = money_format
    row += 1
    ws.cell(row=row, column=1, value=f"Чистая прибыль за 3 года:")
    cell = ws.cell(row=row, column=2, value=total_ebitda)
    cell.number_format = money_format
    if total_ebitda > 0:
        cell.font = Font(color="008000", bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Накопленный CF (с учётом инвестиций):")
    cell = ws.cell(row=row, column=2, value=final_cumulative)
    cell.number_format = money_format
    if final_cumulative > 0:
        cell.font = Font(color="008000", bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"ROI:")
    cell = ws.cell(row=row, column=2, value=roi)
    cell.number_format = pct_format
    if roi >= 2.0:
        cell.font = Font(color="008000", bold=True)

    # Точка безубыточности
    breakeven_month = None
    for d in data:
        if d["cumulative_ebitda"] > 0:
            breakeven_month = d["month_num"]
            breakeven_date = d["date"]
            break

    row += 2
    if breakeven_month:
        ws.cell(row=row, column=1, value=f"Точка безубыточности:")
        ws.cell(row=row, column=2, value=f"Месяц {breakeven_month} ({breakeven_date})")
    else:
        ws.cell(row=row, column=1, value="Точка безубыточности: не достигнута за 36 месяцев")
        ws.cell(row=row, column=1).font = Font(color="FF0000")

    wb.save(filename)
    return {
        "total_revenue": total_revenue,
        "total_ebitda": total_ebitda,
        "final_cf": final_cumulative,
        "roi": roi,
        "breakeven_month": breakeven_month,
    }

if __name__ == "__main__":
    print("Генерация P&L модели...")
    data = generate_pnl()

    filename = "DEMO_PRJ/finances/sunscreen_pnl_v1.xlsx"
    metrics = create_excel(data, filename)

    print(f"\nФайл создан: {filename}")
    print(f"\n--- КЛЮЧЕВЫЕ МЕТРИКИ ---")
    print(f"Выручка за 3 года: {metrics['total_revenue']:,.0f}₽")
    print(f"EBITDA за 3 года: {metrics['total_ebitda']:,.0f}₽")
    print(f"Накопленный CF: {metrics['final_cf']:,.0f}₽")
    print(f"ROI: {metrics['roi']:.1%}")
    if metrics['breakeven_month']:
        print(f"Точка безубыточности: месяц {metrics['breakeven_month']}")
    else:
        print("Точка безубыточности: НЕ ДОСТИГНУТА")
