"""
Скрипт 3: Проверка требований Wildberries к документам

API: https://content-api.wildberries.ru
Требует: WB_API_TOKEN в .env или переменных окружения

Подход:
1. Получить subjectId для наших категорий товаров
2. Для каждого subjectId запросить характеристики через /content/v2/object/charcs/{subjectId}
3. Найти обязательные характеристики, связанные с документами (сертификаты, декларации)

Вход: Товары_15.03.2026.csv + WB API
Выход: output/wb_requirements.xlsx
"""

import csv
import json
import os
import sys
import time
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
except ImportError:
    print("Нужен openpyxl. Установите: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Нужен requests. Установите: pip install requests")
    sys.exit(1)


CONTENT_API = "https://content-api.wildberries.ru"


def load_env():
    """Загрузить переменные из .env файла."""
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


def get_headers():
    """Получить заголовки для WB API."""
    token = os.environ.get("WB_API_TOKEN")
    if not token:
        print("Ошибка: не задан WB_API_TOKEN")
        print("Укажите в .env файле или переменных окружения:")
        print("  WB_API_TOKEN=your-token-here")
        sys.exit(1)

    return {
        "Authorization": token,
        "Content-Type": "application/json",
    }


def api_request(url: str, headers: dict, retries: int = 3):
    """GET-запрос к WB API с ретраями."""
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200:
                return resp.json()
            elif resp.status_code == 429:
                wait = min(2 ** attempt * 2, 30)
                print(f"  Rate limit, ждём {wait}с...")
                time.sleep(wait)
                continue
            else:
                print(f"  Ошибка {resp.status_code}: {resp.text[:200]}")
                if attempt < retries - 1:
                    time.sleep(2)
                    continue
                return None
        except requests.RequestException as e:
            print(f"  Ошибка сети: {e}")
            if attempt < retries - 1:
                time.sleep(2)
                continue
            return None
    return None


def search_subjects(headers: dict, name: str) -> list:
    """Поиск предметов (subjectId) по названию."""
    url = f"{CONTENT_API}/content/v2/object/all?name={requests.utils.quote(name)}&top=10"
    data = api_request(url, headers)
    if data and "data" in data:
        return data["data"]
    return []


def get_subject_charcs(headers: dict, subject_id: int) -> list:
    """Получить характеристики предмета (категории) по subjectId."""
    url = f"{CONTENT_API}/content/v2/object/charcs/{subject_id}"
    data = api_request(url, headers)
    if data and "data" in data:
        return data["data"]
    return []


def find_document_charcs(charcs: list) -> list:
    """Найти характеристики, связанные с документами/сертификатами."""
    doc_keywords = [
        "сертификат", "декларация", "документ", "разрешительн",
        "ГОСТ", "ТР ТС", "соответстви", "безопасност",
        "лицензи", "регистрац",
    ]
    result = []
    for charc in charcs:
        name = charc.get("name", "").lower()
        is_required = charc.get("required", False)
        for keyword in doc_keywords:
            if keyword.lower() in name:
                result.append({
                    "charc_id": charc.get("charcID", charc.get("id", "")),
                    "name": charc.get("name", ""),
                    "required": is_required,
                    "type": charc.get("charcType", charc.get("type", "")),
                    "max_count": charc.get("maxCount", ""),
                })
                break
    return result


def load_our_categories(csv_path: str) -> list:
    """Получить уникальные категории и типы из нашего CSV."""
    categories = set()
    types = set()
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            categories.add(row.get("Категория", "").strip())
            types.add(row.get("Тип", "").strip())
    return sorted(categories), sorted(types)


def write_results(
    subject_results: list,
    doc_charcs_results: list,
    all_charcs: dict,
    output_path: str,
):
    """Записать результаты в Excel."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # === Лист 1: Найденные предметы WB ===
    ws1 = wb.active
    ws1.title = "Предметы WB"
    headers1 = ["Наш запрос", "subjectId", "subjectName", "parentId", "parentName"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, item in enumerate(subject_results, 2):
        for col_idx, key in enumerate(headers1, 1):
            ws1.cell(row=row_idx, column=col_idx, value=str(item.get(key, ""))).border = thin_border

    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"
    for i, w in enumerate([30, 12, 35, 12, 35], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # === Лист 2: Характеристики-документы ===
    ws2 = wb.create_sheet("Требования документов")
    headers2 = ["subjectId", "subjectName", "Характеристика", "Обязательная", "Тип", "MaxCount"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, item in enumerate(doc_charcs_results, 2):
        for col_idx, key in enumerate(headers2, 1):
            ws2.cell(row=row_idx, column=col_idx, value=str(item.get(key, ""))).border = thin_border

    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    # === Лист 3: Все обязательные характеристики ===
    ws3 = wb.create_sheet("Обязательные характеристики")
    headers3 = ["subjectId", "subjectName", "charcID", "Характеристика", "Тип"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    for subject_id, charcs in all_charcs.items():
        subject_name = charcs.get("_name", "")
        for charc in charcs.get("required", []):
            ws3.cell(row=row_idx, column=1, value=subject_id).border = thin_border
            ws3.cell(row=row_idx, column=2, value=subject_name).border = thin_border
            ws3.cell(row=row_idx, column=3, value=charc.get("charcID", "")).border = thin_border
            ws3.cell(row=row_idx, column=4, value=charc.get("name", "")).border = thin_border
            ws3.cell(row=row_idx, column=5, value=charc.get("charcType", "")).border = thin_border
            row_idx += 1

    ws3.auto_filter.ref = ws3.dimensions
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n✓ Результат сохранён: {output_path}")


def main():
    base_dir = Path(__file__).parent.parent
    csv_path = base_dir / "Товары_15.03.2026.csv"
    output_path = base_dir / "output" / "wb_requirements.xlsx"

    load_env()
    headers = get_headers()

    # Загружаем наши категории
    categories, types = load_our_categories(str(csv_path))
    print(f"Наших категорий: {len(categories)}, типов: {len(types)}")

    # Ищем соответствующие предметы на WB
    print("\nПоиск предметов на WB...")
    subject_results = []
    found_subjects = {}  # subject_id -> name

    # Целевые поисковые запросы для автозапчастей
    # WB ищет по началу слова, поэтому используем короткие ключевые слова
    search_terms_auto = [
        # Прямые типы товаров
        "Фары", "Лампы автомобильные", "Прокладки автомобильные",
        "Сальники автомобильные", "Клипсы автомобильные",
        "Ступицы", "Датчики автомобильные", "Датчики парковки",
        "Свечи зажигания", "Свечи накаливания",
        "Фильтры топливные", "Фильтры масляные", "Фильтры воздушные",
        "Щётки стеклоочистителя", "Болты автомобильные",
        "Гайки автомобильные", "Форсунки омывателя",
        "Эмблемы автомобильные", "Реле автомобильные",
        "Термостаты автомобильные", "Подшипники автомобильные",
        "Стойки стабилизатора", "Пружины подвески",
        "Сайлентблоки", "Опоры амортизатора",
        # Общие автокатегории
        "Автозапчасти", "Автокрепеж", "Автоаксессуары",
        "Кузовные детали", "Электрика автомобильная",
        "Запчасти двигателя", "Запчасти подвески",
        "Запчасти тормозной", "Рулевое управление",
        # Наши категории из CSV
    ] + [t for t in types[:20]]  # + топ-20 типов на случай прямого совпадения

    # Дедупликация
    seen_terms = set()
    search_terms = []
    for t in search_terms_auto:
        if t.lower() not in seen_terms:
            seen_terms.add(t.lower())
            search_terms.append(t)

    for term in search_terms:
        subjects = search_subjects(headers, term)
        for s in subjects:
            sid = s.get("subjectID", s.get("id", ""))
            sname = s.get("subjectName", s.get("name", ""))
            pid = s.get("parentID", s.get("parentId", ""))
            pname = s.get("parentName", "")

            subject_results.append({
                "Наш запрос": term,
                "subjectId": sid,
                "subjectName": sname,
                "parentId": pid,
                "parentName": pname,
            })
            if sid:
                found_subjects[sid] = sname

        time.sleep(0.5)  # не нагружаем API

    print(f"  Найдено уникальных subjectId: {len(found_subjects)}")

    # Получаем характеристики для каждого найденного предмета
    print("\nЗагрузка характеристик...")
    doc_charcs_results = []
    all_charcs = {}

    for sid, sname in found_subjects.items():
        charcs = get_subject_charcs(headers, sid)
        if not charcs:
            continue

        # Сохраняем обязательные
        required_charcs = [c for c in charcs if c.get("required")]
        all_charcs[sid] = {"_name": sname, "required": required_charcs}

        # Ищем документные характеристики
        doc_charcs = find_document_charcs(charcs)
        for dc in doc_charcs:
            doc_charcs_results.append({
                "subjectId": sid,
                "subjectName": sname,
                "Характеристика": dc["name"],
                "Обязательная": "Да" if dc["required"] else "Нет",
                "Тип": dc["type"],
                "MaxCount": dc["max_count"],
            })

        time.sleep(0.3)

    print(f"  Категорий с требованиями документов: {len(set(r['subjectId'] for r in doc_charcs_results))}")

    # Сохраняем
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_results(subject_results, doc_charcs_results, all_charcs, str(output_path))

    # Raw JSON
    json_path = base_dir / "output" / "wb_raw_data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "subjects": subject_results,
            "doc_charcs": doc_charcs_results,
            "found_subjects": {str(k): v for k, v in found_subjects.items()},
        }, f, ensure_ascii=False, indent=2, default=str)
    print(f"✓ Raw JSON: {json_path}")


if __name__ == "__main__":
    main()
