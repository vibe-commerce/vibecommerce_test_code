"""
Скрипт 2: Проверка требований Ozon к сертификации

API: https://api-seller.ozon.ru
Требует: OZON_CLIENT_ID и OZON_API_KEY в .env или переменных окружения

Эндпоинты:
- POST /v2/product/certification/list — категории с требованиями
- POST /v1/brand/company-certification/list — бренды с требованиями
- GET /v1/product/certificate/types — справочник типов документов
- GET /v2/product/certificate/accordance-types/list — типы соответствия

Вход: Товары_15.03.2026.csv + Ozon API
Выход: output/ozon_requirements.xlsx
"""

import csv
import json
import os
import sys
import time
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
except ImportError:
    print("Нужен openpyxl. Установите: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Нужен requests. Установите: pip install requests")
    sys.exit(1)


BASE_URL = "https://api-seller.ozon.ru"


def load_env():
    """Загрузить переменные из .env файла."""
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


def get_headers():
    """Получить заголовки для Ozon API."""
    client_id = os.environ.get("OZON_CLIENT_ID")
    api_key = os.environ.get("OZON_API_KEY")

    if not client_id or not api_key:
        print("Ошибка: не заданы OZON_CLIENT_ID и/или OZON_API_KEY")
        print("Укажите их в .env файле или переменных окружения:")
        print("  OZON_CLIENT_ID=123456")
        print("  OZON_API_KEY=your-api-key-here")
        sys.exit(1)

    return {
        "Client-Id": client_id,
        "Api-Key": api_key,
        "Content-Type": "application/json",
    }


def api_request(method: str, endpoint: str, headers: dict, body: dict = None, retries: int = 3):
    """Выполнить запрос к Ozon API с ретраями."""
    url = f"{BASE_URL}{endpoint}"
    for attempt in range(retries):
        try:
            if method == "GET":
                resp = requests.get(url, headers=headers, timeout=30)
            else:
                resp = requests.post(url, headers=headers, json=body or {}, timeout=30)

            if resp.status_code == 200:
                return resp.json()
            elif resp.status_code == 429:
                wait = min(2 ** attempt * 2, 30)
                print(f"  Rate limit, ждём {wait}с...")
                time.sleep(wait)
                continue
            else:
                print(f"  Ошибка {resp.status_code}: {resp.text[:200]}")
                if attempt < retries - 1:
                    time.sleep(2)
                    continue
                return None
        except requests.RequestException as e:
            print(f"  Ошибка сети: {e}")
            if attempt < retries - 1:
                time.sleep(2)
                continue
            return None
    return None


def fetch_certification_categories(headers: dict) -> list:
    """Шаг 1: получить список категорий с требованиями к сертификации."""
    print("Шаг 1: Загрузка категорий с требованиями сертификации...")
    all_categories = []
    page = 1

    while True:
        data = api_request("POST", "/v2/product/certification/list", headers, {
            "page": page,
            "page_size": 1000,
        })
        if not data:
            break

        items = data.get("result", data.get("items", []))
        if isinstance(items, dict):
            items = items.get("items", [])

        if not items:
            break

        all_categories.extend(items)
        print(f"  Страница {page}: {len(items)} категорий")

        if len(items) < 1000:
            break
        page += 1

    print(f"  Итого: {len(all_categories)} категорий")
    return all_categories


def fetch_brand_certifications(headers: dict) -> list:
    """Шаг 2: получить бренды с требованиями к документам."""
    print("Шаг 2: Загрузка брендов с требованиями...")
    all_brands = []
    page = 1

    while True:
        data = api_request("POST", "/v1/brand/company-certification/list", headers, {
            "page": page,
            "page_size": 1000,
        })
        if not data:
            break

        items = data.get("result", data.get("items", []))
        if isinstance(items, dict):
            items = items.get("items", [])

        if not items:
            break

        all_brands.extend(items)
        print(f"  Страница {page}: {len(items)} брендов")

        if len(items) < 1000:
            break
        page += 1

    print(f"  Итого: {len(all_brands)} брендов")
    return all_brands


def fetch_certificate_types(headers: dict) -> list:
    """Шаг 3: справочник типов сертификатов."""
    print("Шаг 3: Загрузка справочника типов документов...")
    data = api_request("GET", "/v1/product/certificate/types", headers)
    if data:
        items = data.get("result", [])
        print(f"  Типы документов: {len(items)}")
        return items
    return []


def fetch_accordance_types(headers: dict) -> list:
    """Шаг 4: типы соответствия (ГОСТ, ТР ТС и т.д.)."""
    print("Шаг 4: Загрузка типов соответствия...")
    data = api_request("GET", "/v2/product/certificate/accordance-types/list", headers)
    if data:
        items = data.get("result", [])
        print(f"  Типы соответствия: {len(items)}")
        return items
    return []


def load_our_products(csv_path: str) -> list:
    """Загрузить наши товары из CSV."""
    products = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            products.append(row)
    return products


def match_categories(our_products: list, ozon_categories: list) -> dict:
    """Сопоставить наши категории с категориями Ozon.

    Простой matching по названию. В реальности нужен маппинг category_id,
    который мы получим, когда товары будут загружены на Ozon.
    """
    our_cats = set()
    for p in our_products:
        our_cats.add(p.get("Категория", "").strip())
        our_cats.add(p.get("Тип", "").strip())

    # Для каждой категории Ozon проверяем, есть ли пересечение по ключевым словам
    matches = {}
    for ozon_cat in ozon_categories:
        cat_name = ozon_cat.get("category_name", ozon_cat.get("name", ""))
        is_required = ozon_cat.get("is_required", False)
        type_name = ozon_cat.get("type_name", "")

        for our_cat in our_cats:
            # Простое сравнение по вхождению ключевых слов
            our_words = set(our_cat.lower().split())
            ozon_words = set(cat_name.lower().split())
            if our_words & ozon_words and len(our_words & ozon_words) >= 1:
                if our_cat not in matches:
                    matches[our_cat] = []
                matches[our_cat].append({
                    "ozon_category": cat_name,
                    "is_required": is_required,
                    "type_name": type_name,
                    "raw": ozon_cat,
                })

    return matches


def write_results(
    ozon_categories: list,
    ozon_brands: list,
    cert_types: list,
    accordance_types: list,
    category_matches: dict,
    output_path: str,
):
    """Записать результаты в Excel."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # === Лист 1: Категории Ozon ===
    ws1 = wb.active
    ws1.title = "Категории Ozon"
    headers1 = ["category_id", "category_name", "is_required", "type_id", "type_name"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, cat in enumerate(ozon_categories, 2):
        for col_idx, key in enumerate(headers1, 1):
            val = cat.get(key, "")
            ws1.cell(row=row_idx, column=col_idx, value=str(val)).border = thin_border

    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"
    for i, w in enumerate([15, 50, 12, 10, 30], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # === Лист 2: Бренды с требованиями ===
    ws2 = wb.create_sheet("Бренды Ozon")
    headers2 = ["brand_id", "brand_name", "has_certificate"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, brand in enumerate(ozon_brands, 2):
        for col_idx, key in enumerate(headers2, 1):
            val = brand.get(key, "")
            ws2.cell(row=row_idx, column=col_idx, value=str(val)).border = thin_border

    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    # === Лист 3: Типы документов ===
    ws3 = wb.create_sheet("Типы документов")
    ws3.cell(row=1, column=1, value="Тип").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Описание").font = Font(bold=True)
    for row_idx, ct in enumerate(cert_types, 2):
        if isinstance(ct, dict):
            ws3.cell(row=row_idx, column=1, value=str(ct.get("value", ct.get("type", ct))))
            ws3.cell(row=row_idx, column=2, value=str(ct.get("name", ct.get("description", ""))))
        else:
            ws3.cell(row=row_idx, column=1, value=str(ct))

    # === Лист 4: Маппинг наших категорий ===
    ws4 = wb.create_sheet("Маппинг категорий")
    headers4 = ["Наша категория/тип", "Категория Ozon", "Требуется документ", "Тип документа"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    for our_cat, matches in sorted(category_matches.items()):
        for match in matches:
            ws4.cell(row=row_idx, column=1, value=our_cat).border = thin_border
            ws4.cell(row=row_idx, column=2, value=match["ozon_category"]).border = thin_border
            ws4.cell(row=row_idx, column=3, value=str(match["is_required"])).border = thin_border
            ws4.cell(row=row_idx, column=4, value=match["type_name"]).border = thin_border
            row_idx += 1

    ws4.auto_filter.ref = ws4.dimensions
    ws4.freeze_panes = "A2"
    for i, w in enumerate([35, 50, 18, 30], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # === Лист 5: JSON-дамп (для отладки) ===
    ws5 = wb.create_sheet("Raw JSON")
    ws5.cell(row=1, column=1, value="Тип данных")
    ws5.cell(row=1, column=2, value="JSON")
    ws5.cell(row=2, column=1, value="categories_count")
    ws5.cell(row=2, column=2, value=len(ozon_categories))
    ws5.cell(row=3, column=1, value="brands_count")
    ws5.cell(row=3, column=2, value=len(ozon_brands))
    ws5.cell(row=4, column=1, value="cert_types")
    ws5.cell(row=4, column=2, value=json.dumps(cert_types, ensure_ascii=False, default=str)[:32000])
    ws5.cell(row=5, column=1, value="accordance_types")
    ws5.cell(row=5, column=2, value=json.dumps(accordance_types, ensure_ascii=False, default=str)[:32000])

    wb.save(output_path)
    print(f"\n✓ Результат сохранён: {output_path}")


def main():
    base_dir = Path(__file__).parent.parent
    csv_path = base_dir / "Товары_15.03.2026.csv"
    output_path = base_dir / "output" / "ozon_requirements.xlsx"

    load_env()
    headers = get_headers()

    # Загружаем данные из Ozon API
    ozon_categories = fetch_certification_categories(headers)
    ozon_brands = fetch_brand_certifications(headers)
    cert_types = fetch_certificate_types(headers)
    accordance_types = fetch_accordance_types(headers)

    # Загружаем наши товары
    our_products = load_our_products(str(csv_path))
    print(f"\nНаших товаров: {len(our_products)}")

    # Сопоставляем категории
    category_matches = match_categories(our_products, ozon_categories)
    print(f"Сопоставлено категорий/типов: {len(category_matches)}")

    # Сохраняем результат
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_results(
        ozon_categories, ozon_brands, cert_types, accordance_types,
        category_matches, str(output_path),
    )

    # Также сохраняем raw JSON для последующего анализа
    json_path = base_dir / "output" / "ozon_raw_data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "categories": ozon_categories,
            "brands": ozon_brands,
            "cert_types": cert_types,
            "accordance_types": accordance_types,
        }, f, ensure_ascii=False, indent=2)
    print(f"✓ Raw JSON: {json_path}")


if __name__ == "__main__":
    main()
