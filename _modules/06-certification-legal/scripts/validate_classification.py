"""
Скрипт 5: Валидация мастер-таблицы

Проверки из плана (раздел 6.5):
1. count == 2211
2. no_nulls(ИТОГО_документ)
3. all_same_within_type(Тип → ИТОГО_документ)
4. fares_are_certificate(Фары автомобильные)
5. labels_are_exempt(Этикетка)
6. no_duplicate_articles()
7. packaging_variants_consistent()
8. known_brands_resolved(«Нет бренда» ≤ 5% нерешённых)
9. distribution_summary()
10. discrepancies_law_vs_platform()

Вход: output/master_table.xlsx
Выход: PASS/FAIL + отчёт в консоль + output/validation_report.txt
"""

import sys
from pathlib import Path
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("Нужен openpyxl. Установите: pip install openpyxl")
    sys.exit(1)


EXPECTED_SKU_COUNT = 2211


def load_master_table(path: str) -> list:
    """Загрузить мастер-таблицу."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Мастер-таблица"]
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    wb.close()
    return data


class ValidationResult:
    def __init__(self):
        self.checks = []
        self.passed = 0
        self.failed = 0
        self.warnings = 0

    def check(self, name: str, passed: bool, details: str = ""):
        status = "PASS" if passed else "FAIL"
        self.checks.append((name, status, details))
        if passed:
            self.passed += 1
        else:
            self.failed += 1

    def warn(self, name: str, details: str = ""):
        self.checks.append((name, "WARN", details))
        self.warnings += 1

    def report(self) -> str:
        lines = ["=" * 70, "ОТЧЁТ ВАЛИДАЦИИ МАСТЕР-ТАБЛИЦЫ", "=" * 70, ""]

        for name, status, details in self.checks:
            icon = {"PASS": "✓", "FAIL": "✗", "WARN": "⚠"}[status]
            lines.append(f"  {icon} [{status}] {name}")
            if details:
                for line in details.split("\n"):
                    lines.append(f"       {line}")
            lines.append("")

        lines.append("-" * 70)
        lines.append(f"ИТОГО: {self.passed} PASS, {self.failed} FAIL, {self.warnings} WARN")

        overall = "PASS" if self.failed == 0 else "FAIL"
        lines.append(f"РЕЗУЛЬТАТ: {overall}")
        lines.append("=" * 70)

        return "\n".join(lines)


def validate(data: list) -> ValidationResult:
    """Выполнить все проверки."""
    v = ValidationResult()

    # 1. Полнота: count == 2211
    v.check(
        "1. Полнота (count == 2211)",
        len(data) == EXPECTED_SKU_COUNT,
        f"Фактически: {len(data)} SKU",
    )

    # 2. Нет пустых ИТОГО_документ
    nulls = [item["Артикул"] for item in data if not item.get("ИТОГО_документ")]
    v.check(
        "2. Нет пустых ИТОГО_документ",
        len(nulls) == 0,
        f"Пустых: {len(nulls)}" + (f" (первые 5: {nulls[:5]})" if nulls else ""),
    )

    # 3. Консистентность: один тип → один документ
    type_docs = defaultdict(set)
    for item in data:
        type_docs[item.get("Тип", "")].add(item.get("ИТОГО_документ", ""))
    inconsistent = {t: docs for t, docs in type_docs.items() if len(docs) > 1}
    v.check(
        "3. Консистентность (один тип → один документ)",
        len(inconsistent) == 0,
        "\n".join(f"{t}: {docs}" for t, docs in inconsistent.items()) if inconsistent else "Все типы согласованы",
    )

    # 4. Фары = СЕРТИФИКАТ
    fares = [item for item in data if item.get("Тип") == "Фары автомобильные"]
    fares_cert = all(item.get("ИТОГО_документ") == "СЕРТИФИКАТ" for item in fares)
    v.check(
        "4. Фары автомобильные = СЕРТИФИКАТ",
        fares_cert and len(fares) > 0,
        f"Найдено {len(fares)} фар, все СЕРТИФИКАТ: {fares_cert}",
    )

    # 5. Этикетки = НЕ_ТРЕБУЕТСЯ или ОТКАЗНОЕ
    labels = [item for item in data if item.get("Тип") == "Этикетка"]
    labels_ok = all(
        item.get("ИТОГО_документ") in ("НЕ_ТРЕБУЕТСЯ", "ОТКАЗНОЕ")
        for item in labels
    )
    v.check(
        "5. Этикетки = НЕ_ТРЕБУЕТСЯ или ОТКАЗНОЕ",
        labels_ok and len(labels) > 0,
        f"Найдено {len(labels)} этикеток, все корректны: {labels_ok}",
    )

    # 6. Нет дубликатов артикулов
    articles = [item.get("Артикул") for item in data]
    dupes = [a for a, c in Counter(articles).items() if c > 1]
    v.check(
        "6. Нет дубликатов артикулов",
        len(dupes) == 0,
        f"Дубликатов: {len(dupes)}" + (f" (первые 5: {dupes[:5]})" if dupes else ""),
    )

    # 7. Фасовки согласованы
    import re
    base_to_docs = defaultdict(set)
    for item in data:
        base = re.sub(r"_\d+pcs$", "", item.get("Артикул", ""), flags=re.IGNORECASE)
        base_to_docs[base].add(item.get("ИТОГО_документ", ""))
    inconsistent_pkg = {b: docs for b, docs in base_to_docs.items() if len(docs) > 1}
    v.check(
        "7. Фасовки согласованы",
        len(inconsistent_pkg) == 0,
        f"Несогласованных: {len(inconsistent_pkg)}" + (
            "\n" + "\n".join(f"{b}: {docs}" for b, docs in list(inconsistent_pkg.items())[:10])
            if inconsistent_pkg else ""
        ),
    )

    # 8. «Нет бренда» — ≤ 5% нерешённых (НЕ_ОПРЕДЕЛЕНО)
    no_brand = [item for item in data if item.get("Бренд") in ("Нет бренда", "", None)]
    no_brand_undetermined = [item for item in no_brand if item.get("ИТОГО_документ") == "НЕ_ОПРЕДЕЛЕНО"]
    pct = len(no_brand_undetermined) / len(no_brand) * 100 if no_brand else 0
    v.check(
        "8. «Нет бренда» — НЕ_ОПРЕДЕЛЕНО ≤ 5%",
        pct <= 5,
        f"«Нет бренда»: {len(no_brand)} SKU, из них НЕ_ОПРЕДЕЛЕНО: {len(no_brand_undetermined)} ({pct:.1f}%)",
    )

    # 9. Распределение — правдоподобность
    stats = Counter(item.get("ИТОГО_документ") for item in data)
    total = len(data)
    dist_lines = []
    for doc_type, count in stats.most_common():
        dist_lines.append(f"{doc_type}: {count} ({count/total*100:.1f}%)")

    cert_pct = stats.get("СЕРТИФИКАТ", 0) / total * 100
    otkaz_pct = stats.get("ОТКАЗНОЕ", 0) / total * 100

    # Ожидания из плана: А ~5–10%, В ~70%+
    plausible = 3 <= cert_pct <= 15 and otkaz_pct >= 40
    if plausible:
        v.check("9. Распределение правдоподобно", True, "\n".join(dist_lines))
    else:
        v.warn(
            "9. Распределение — проверить вручную",
            "\n".join(dist_lines) + f"\nСЕРТИФИКАТ {cert_pct:.1f}% (ожидание 5–10%), ОТКАЗНОЕ {otkaz_pct:.1f}% (ожидание ≥70%)",
        )

    # 10. Расхождения закон vs площадки
    discrepancies = [item for item in data if item.get("Расхождение") == "ДА"]
    if discrepancies:
        disc_lines = []
        for item in discrepancies[:10]:
            disc_lines.append(
                f"{item['Артикул']}: закон={item.get('Документ_по_закону','')}, "
                f"Ozon={item.get('Документ_Ozon','')}, WB={item.get('Документ_WB','')}"
            )
        v.warn(
            f"10. Расхождения закон vs площадки: {len(discrepancies)} SKU",
            "\n".join(disc_lines) + ("\n..." if len(discrepancies) > 10 else ""),
        )
    else:
        v.check(
            "10. Расхождения закон vs площадки",
            True,
            "Нет расхождений (или данные площадок отсутствуют)",
        )

    # Дополнительно: проверка маяковых товаров
    beacons = {
        "ЕС14000ОР": "СЕРТИФИКАТ",      # Фонарь — светотехника
        "CHHPS0698": "НЕ_ТРЕБУЕТСЯ",     # Этикетка
        "MD050317": "ДЕКЛАРАЦИЯ",         # Прокладка
        "8E0863727A": "ОТКАЗНОЕ",         # Клипса
        "MB814468": "ОТКАЗНОЕ",           # Эмблема
    }

    beacon_results = []
    for art_search, expected in beacons.items():
        found = [item for item in data if art_search in str(item.get("Артикул", ""))]
        if found:
            actual = found[0].get("ИТОГО_документ", "")
            status = "✓" if actual == expected else "✗"
            beacon_results.append(f"{status} {art_search}: ожидалось {expected}, получено {actual}")
        else:
            beacon_results.append(f"? {art_search}: не найден")

    all_beacons_pass = all(line.startswith("✓") for line in beacon_results)
    v.check(
        "11. Маяковые товары",
        all_beacons_pass,
        "\n".join(beacon_results),
    )

    return v


def main():
    base_dir = Path(__file__).parent.parent
    master_path = base_dir / "output" / "master_table.xlsx"

    if not master_path.exists():
        print(f"Ошибка: не найден файл {master_path}")
        print("Сначала запустите: python merge_classification.py")
        sys.exit(1)

    print(f"Загрузка: {master_path}")
    data = load_master_table(str(master_path))
    print(f"Загружено: {len(data)} SKU")

    result = validate(data)
    report = result.report()
    print(report)

    # Сохраняем отчёт
    report_path = base_dir / "output" / "validation_report.txt"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"\nОтчёт сохранён: {report_path}")

    # Exit code
    sys.exit(0 if result.failed == 0 else 1)


if __name__ == "__main__":
    main()
