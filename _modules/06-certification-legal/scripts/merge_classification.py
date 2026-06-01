"""
Скрипт 4: Объединение классификаций в мастер-таблицу

Объединяет результаты:
- Скрипт 1 (classify_by_legislation.py) → classification_legislation.xlsx
- Скрипт 2 (check_ozon_requirements.py) → ozon_requirements.xlsx (если есть)
- Скрипт 3 (check_wb_requirements.py) → wb_requirements.xlsx (если есть)

Логика: ИТОГО_документ = MAX(закон, Ozon, WB)
Приоритет: СЕРТИФИКАТ > ДЕКЛАРАЦИЯ > ОТКАЗНОЕ > НЕ_ТРЕБУЕТСЯ

Вход: output/classification_legislation.xlsx + ozon_requirements.xlsx + wb_requirements.xlsx
Выход: output/master_table.xlsx
"""

import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Нужен openpyxl. Установите: pip install openpyxl")
    sys.exit(1)


# Приоритет документов (больше = строже)
DOC_PRIORITY = {
    "СЕРТИФИКАТ": 4,
    "ДЕКЛАРАЦИЯ": 3,
    "ОТКАЗНОЕ": 2,
    "НЕ_ТРЕБУЕТСЯ": 1,
    "НЕ_ОПРЕДЕЛЕНО": 0,
    "": 0,
}

DOC_FILLS = {
    "СЕРТИФИКАТ": PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid"),
    "ДЕКЛАРАЦИЯ": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "ОТКАЗНОЕ": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
    "НЕ_ТРЕБУЕТСЯ": PatternFill(start_color="D0E0E3", end_color="D0E0E3", fill_type="solid"),
    "НЕ_ОПРЕДЕЛЕНО": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}


def max_doc(doc1: str, doc2: str) -> str:
    """Вернуть документ с наибольшим приоритетом."""
    p1 = DOC_PRIORITY.get(doc1, 0)
    p2 = DOC_PRIORITY.get(doc2, 0)
    return doc1 if p1 >= p2 else doc2


def load_legislation(path: str) -> dict:
    """Загрузить классификацию по законодательству."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Классификация"]

    # Получаем заголовки
    headers = [cell.value for cell in ws[1]]
    result = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        article = data.get("Артикул", "")
        if article:
            result[article] = {
                "base_article": data.get("Базовый_артикул", ""),
                "name": data.get("Название", ""),
                "brand": data.get("Бренд", ""),
                "category": data.get("Категория", ""),
                "type": data.get("Тип", ""),
                "doc_law": data.get("Документ_по_закону", ""),
                "basis_law": data.get("Основание", ""),
                "confidence": data.get("Уверенность_%", 0),
            }

    wb.close()
    return result


def load_ozon_mapping(path: str) -> dict:
    """Загрузить маппинг категорий Ozon (если файл существует)."""
    if not Path(path).exists():
        print(f"  Файл Ozon не найден: {path} — пропускаем")
        return {}

    wb = openpyxl.load_workbook(path, read_only=True)

    # Маппинг: наша категория/тип → требование Ozon
    result = {}
    if "Маппинг категорий" in wb.sheetnames:
        ws = wb["Маппинг категорий"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            our_cat = data.get("Наша категория/тип", "")
            is_required = str(data.get("Требуется документ", "")).lower() in ("true", "да", "yes", "1")
            doc_type = data.get("Тип документа", "")
            if our_cat and is_required:
                # Маппим тип документа Ozon к нашей классификации
                ozon_doc = map_ozon_doc_type(doc_type)
                if our_cat not in result or DOC_PRIORITY.get(ozon_doc, 0) > DOC_PRIORITY.get(result[our_cat], 0):
                    result[our_cat] = ozon_doc

    wb.close()
    return result


def map_ozon_doc_type(ozon_type: str) -> str:
    """Маппинг типов документов Ozon на нашу классификацию."""
    ozon_type = str(ozon_type).lower()
    if "certificate_of_conformity" in ozon_type or "сертификат" in ozon_type:
        return "СЕРТИФИКАТ"
    elif "declaration" in ozon_type or "декларация" in ozon_type:
        return "ДЕКЛАРАЦИЯ"
    elif "refused" in ozon_type or "отказ" in ozon_type:
        return "ОТКАЗНОЕ"
    elif "registration" in ozon_type or "регистрац" in ozon_type:
        return "СЕРТИФИКАТ"  # Свидетельство о регистрации ≈ сертификат
    else:
        return "ОТКАЗНОЕ"  # По умолчанию — отказное


def load_wb_mapping(path: str) -> dict:
    """Загрузить требования WB (если файл существует)."""
    if not Path(path).exists():
        print(f"  Файл WB не найден: {path} — пропускаем")
        return {}

    wb = openpyxl.load_workbook(path, read_only=True)
    result = {}

    if "Требования документов" in wb.sheetnames:
        ws = wb["Требования документов"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            subject_name = data.get("subjectName", "")
            is_required = str(data.get("Обязательная", "")).lower() in ("да", "true", "yes", "1")
            if subject_name and is_required:
                result[subject_name] = "ОТКАЗНОЕ"  # WB требует минимум отказное

    wb.close()
    return result


def merge_all(
    legislation: dict,
    ozon_mapping: dict,
    wb_mapping: dict,
) -> list:
    """Объединить все источники в мастер-таблицу."""
    master = []

    for article, data in legislation.items():
        doc_law = data["doc_law"] or ""
        doc_ozon = ""
        doc_wb = ""

        # Сопоставляем с Ozon по категории и типу
        if data["type"] in ozon_mapping:
            doc_ozon = ozon_mapping[data["type"]]
        elif data["category"] in ozon_mapping:
            doc_ozon = ozon_mapping[data["category"]]

        # Сопоставляем с WB по типу
        for wb_subject, wb_doc in wb_mapping.items():
            if data["type"].lower() in wb_subject.lower() or wb_subject.lower() in data["type"].lower():
                doc_wb = max_doc(doc_wb, wb_doc)
                break

        # Итоговый документ = MAX(закон, Ozon, WB)
        doc_total = doc_law
        doc_total = max_doc(doc_total, doc_ozon)
        doc_total = max_doc(doc_total, doc_wb)

        # Расхождение?
        sources = [doc_law, doc_ozon, doc_wb]
        non_empty = [s for s in sources if s]
        has_discrepancy = len(set(non_empty)) > 1 if len(non_empty) > 1 else False

        master.append({
            "Артикул": article,
            "Базовый_артикул": data["base_article"],
            "Название": data["name"],
            "Бренд": data["brand"],
            "Категория": data["category"],
            "Тип": data["type"],
            "Документ_по_закону": doc_law,
            "Документ_Ozon": doc_ozon,
            "Документ_WB": doc_wb,
            "ИТОГО_документ": doc_total,
            "Расхождение": "ДА" if has_discrepancy else "",
            "Основание": data["basis_law"],
            "Уверенность_%": data["confidence"],
        })

    return master


def write_master_table(master: list, output_path: str):
    """Записать мастер-таблицу в Excel."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    discrepancy_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    # === Лист 1: Мастер-таблица ===
    ws = wb.active
    ws.title = "Мастер-таблица"

    headers = [
        "Артикул", "Базовый_артикул", "Название", "Бренд", "Категория", "Тип",
        "Документ_по_закону", "Документ_Ozon", "Документ_WB",
        "ИТОГО_документ", "Расхождение", "Основание", "Уверенность_%",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, item in enumerate(master, 2):
        for col_idx, key in enumerate(headers, 1):
            val = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

            # Цвет для документных колонок
            if key in ("Документ_по_закону", "Документ_Ozon", "Документ_WB", "ИТОГО_документ"):
                fill = DOC_FILLS.get(val)
                if fill:
                    cell.fill = fill

            # Подсветка расхождений
            if key == "Расхождение" and val == "ДА":
                cell.fill = discrepancy_fill

    # Ширина колонок
    col_widths = [18, 18, 50, 15, 30, 35, 18, 15, 15, 18, 12, 55, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # === Лист 2: Статистика ===
    ws2 = wb.create_sheet("Статистика")
    from collections import Counter

    stats = Counter(item["ИТОГО_документ"] for item in master)
    ws2.cell(row=1, column=1, value="ИТОГО_документ").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="SKU").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Доля %").font = Font(bold=True)

    total = len(master)
    for row_idx, (doc, count) in enumerate(stats.most_common(), 2):
        ws2.cell(row=row_idx, column=1, value=doc)
        ws2.cell(row=row_idx, column=2, value=count)
        ws2.cell(row=row_idx, column=3, value=round(count / total * 100, 1))
        fill = DOC_FILLS.get(doc)
        if fill:
            ws2.cell(row=row_idx, column=1).fill = fill

    # Расхождения
    discrepancies = sum(1 for item in master if item["Расхождение"] == "ДА")
    ws2.cell(row=row_idx + 2, column=1, value="Расхождения закон/площадки").font = Font(bold=True)
    ws2.cell(row=row_idx + 2, column=2, value=discrepancies)

    # НЕ_ОПРЕДЕЛЕНО
    undetermined = sum(1 for item in master if item["ИТОГО_документ"] == "НЕ_ОПРЕДЕЛЕНО")
    ws2.cell(row=row_idx + 3, column=1, value="НЕ_ОПРЕДЕЛЕНО (требуют ручной классификации)")
    ws2.cell(row=row_idx + 3, column=2, value=undetermined)

    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    # === Лист 3: Расхождения (детали) ===
    ws3 = wb.create_sheet("Расхождения")
    headers3 = ["Артикул", "Название", "Тип", "Закон", "Ozon", "WB", "ИТОГО"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        cell.border = thin_border

    row_idx = 2
    for item in master:
        if item["Расхождение"] == "ДА":
            ws3.cell(row=row_idx, column=1, value=item["Артикул"]).border = thin_border
            ws3.cell(row=row_idx, column=2, value=item["Название"]).border = thin_border
            ws3.cell(row=row_idx, column=3, value=item["Тип"]).border = thin_border
            ws3.cell(row=row_idx, column=4, value=item["Документ_по_закону"]).border = thin_border
            ws3.cell(row=row_idx, column=5, value=item["Документ_Ozon"]).border = thin_border
            ws3.cell(row=row_idx, column=6, value=item["Документ_WB"]).border = thin_border
            ws3.cell(row=row_idx, column=7, value=item["ИТОГО_документ"]).border = thin_border
            row_idx += 1

    ws3.auto_filter.ref = ws3.dimensions
    ws3.freeze_panes = "A2"
    for i, w in enumerate([18, 50, 35, 15, 15, 15, 15], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # === Лист 4: Группировка по бренду (для Этапа 2 — запрос поставщикам) ===
    ws4 = wb.create_sheet("По брендам")
    headers4 = ["Бренд", "Всего SKU", "СЕРТИФИКАТ", "ДЕКЛАРАЦИЯ", "ОТКАЗНОЕ", "НЕ_ТРЕБУЕТСЯ", "НЕ_ОПРЕДЕЛЕНО"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    brand_stats = defaultdict(lambda: Counter())
    for item in master:
        brand_stats[item["Бренд"]][item["ИТОГО_документ"]] += 1

    for row_idx, (brand, counts) in enumerate(
        sorted(brand_stats.items(), key=lambda x: sum(x[1].values()), reverse=True), 2
    ):
        ws4.cell(row=row_idx, column=1, value=brand).border = thin_border
        ws4.cell(row=row_idx, column=2, value=sum(counts.values())).border = thin_border
        for col_idx, doc_type in enumerate(["СЕРТИФИКАТ", "ДЕКЛАРАЦИЯ", "ОТКАЗНОЕ", "НЕ_ТРЕБУЕТСЯ", "НЕ_ОПРЕДЕЛЕНО"], 3):
            ws4.cell(row=row_idx, column=col_idx, value=counts.get(doc_type, 0)).border = thin_border

    ws4.auto_filter.ref = ws4.dimensions
    ws4.freeze_panes = "A2"
    ws4.column_dimensions["A"].width = 25

    wb.save(output_path)
    print(f"\n✓ Мастер-таблица сохранена: {output_path}")


def main():
    base_dir = Path(__file__).parent.parent
    output_dir = base_dir / "output"

    legislation_path = output_dir / "classification_legislation.xlsx"
    ozon_path = output_dir / "ozon_requirements.xlsx"
    wb_path = output_dir / "wb_requirements.xlsx"
    master_path = output_dir / "master_table.xlsx"

    if not legislation_path.exists():
        print(f"Ошибка: не найден файл классификации {legislation_path}")
        print("Сначала запустите: python classify_by_legislation.py")
        sys.exit(1)

    print("Загрузка данных...")
    print(f"  Законодательство: {legislation_path}")
    legislation = load_legislation(str(legislation_path))
    print(f"  Загружено {len(legislation)} SKU из классификации по закону")

    ozon_mapping = load_ozon_mapping(str(ozon_path))
    print(f"  Ozon: {len(ozon_mapping)} маппингов категорий")

    wb_mapping = load_wb_mapping(str(wb_path))
    print(f"  WB: {len(wb_mapping)} маппингов категорий")

    print("\nОбъединение...")
    master = merge_all(legislation, ozon_mapping, wb_mapping)

    print(f"\nМастер-таблица: {len(master)} SKU")

    # Статистика
    from collections import Counter
    stats = Counter(item["ИТОГО_документ"] for item in master)
    for doc_type, count in stats.most_common():
        pct = count / len(master) * 100
        print(f"  {doc_type:20s}: {count:5d} ({pct:.1f}%)")

    discrepancies = sum(1 for item in master if item["Расхождение"] == "ДА")
    print(f"  Расхождения: {discrepancies}")

    write_master_table(master, str(master_path))


if __name__ == "__main__":
    main()
