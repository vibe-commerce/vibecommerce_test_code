"""
Cashflow model for marketplace business (WB + Ozon).
Input:  _modules/05-ads-optimization/data-demo/{sales_data_v1.0,ads_data_v2.0}.xlsx
Output: cashflow_model.xlsx

Allowed assumptions (explicit, tunable):
  START_CASH            — стартовый денежный остаток, ₽
  WB_PAYOUT_LAG_DAYS    — лаг выплаты WB, дней (T+7 по умолчанию)
  OZON_PAYOUT_LAG_DAYS  — лаг выплаты Ozon, дней (T+14 по умолчанию)
  SUPPLIER_PREPAY_DAYS  — предоплата поставщику за N дней до продажи
  SUPPLIER_PREPAY_SHARE — доля предоплаты (1.0 = 100%, 0.5 = 50/50)
  ADS_SHARE_OF_REVENUE  — доля рекламного бюджета от выручки (предоплата МП)
  OPEX_MONTHLY          — операционные расходы, ₽/мес (ФОТ, связь, налоги кроме УСН)
  USN_RATE              — ставка УСН "доходы" = 6%
  STORAGE_SHARE         — хранение на складе МП, % от выручки
  FORECAST_WEEKS        — недель прогноза вперёд
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

MODULES = Path(__file__).resolve().parents[2]
# Демо-данные общие с модулем 05 (генерируются его скриптами в data-demo/)
SALES_FILE = MODULES / '05-ads-optimization' / 'data-demo' / 'sales_data_v1.0.xlsx'
ADS_FILE   = MODULES / '05-ads-optimization' / 'data-demo' / 'ads_data_v2.0.xlsx'
OUT        = Path(__file__).resolve().parent / 'cashflow_model.xlsx'

# ==== ДОПУЩЕНИЯ (меняй здесь) ========================================
START_CASH            = 1_500_000
WB_PAYOUT_LAG_DAYS    = 7
OZON_PAYOUT_LAG_DAYS  = 14
SUPPLIER_PREPAY_DAYS  = 30
SUPPLIER_PREPAY_SHARE = 1.00      # 100% предоплата поставщику
ADS_PAY_LAG_DAYS      = 0         # предоплата рекламы (пополнение баланса)
STORAGE_SHARE         = 0.015     # 1.5% от выручки на хранение МП
OPEX_MONTHLY          = 250_000   # ФОТ + связь + бухгалтерия + прочее
USN_RATE              = 0.06      # УСН 6%
FORECAST_WEEKS        = 8         # прогноз вперёд
GROWTH_WOW            = 0.00      # ожидаемый рост неделя-к-неделе в прогнозе
# =====================================================================


def load_data():
    sales = pd.read_excel(SALES_FILE, sheet_name='Продажи')
    cat   = pd.read_excel(ADS_FILE,   sheet_name='Каталог')
    ads   = pd.read_excel(ADS_FILE,   sheet_name='Сводка')

    sales = sales.merge(
        cat[['SKU','Себестоимость, ₽','Комиссия WB, %','Комиссия Ozon, %',
             'Логистика WB, ₽/шт','Логистика Ozon, ₽/шт']],
        on='SKU', how='left', suffixes=('','_cat')
    )
    sales['Выручка']   = sales['Продажи, шт'] * sales['Цена, ₽']
    sales['Себест']    = sales['Продажи, шт'] * sales['Себестоимость, ₽']
    sales['Комис_%']   = np.where(sales['Площадка']=='WB',
                                  sales['Комиссия WB, %'],
                                  sales['Комиссия Ozon, %'])
    sales['Лог_ед']    = np.where(sales['Площадка']=='WB',
                                  sales['Логистика WB, ₽/шт'],
                                  sales['Логистика Ozon, ₽/шт'])
    sales['Комиссия']  = sales['Выручка'] * sales['Комис_%']
    sales['Логистика'] = sales['Продажи, шт'] * sales['Лог_ед']
    sales['Хранение']  = sales['Выручка'] * STORAGE_SHARE

    w_sales = sales.groupby(['Неделя','Площадка']).agg(
        выручка=('Выручка','sum'),
        себест=('Себест','sum'),
        комиссия=('Комиссия','sum'),
        логистика=('Логистика','sum'),
        хранение=('Хранение','sum'),
    ).reset_index()

    ads_w = ads.groupby('Неделя').agg(реклама=('Расход рекл., ₽','sum')).reset_index()
    return w_sales, ads_w


def build_forecast(w_sales: pd.DataFrame, ads_w: pd.DataFrame, weeks: int):
    """Лёгкий прогноз: берём средние 4 последних недель + рост WoW."""
    last4 = w_sales[w_sales['Неделя'] >= w_sales['Неделя'].max() - pd.Timedelta(weeks=3)]
    base_wb   = last4[last4['Площадка']=='WB'][['выручка','себест','комиссия','логистика','хранение']].mean()
    base_oz   = last4[last4['Площадка']=='Ozon'][['выручка','себест','комиссия','логистика','хранение']].mean()
    base_ads  = ads_w.tail(4)['реклама'].mean()

    last_week = w_sales['Неделя'].max()
    rows_sales, rows_ads = [], []
    for i in range(1, weeks+1):
        wk = last_week + pd.Timedelta(weeks=i)
        mult = (1 + GROWTH_WOW) ** i
        for plat, base in [('WB', base_wb), ('Ozon', base_oz)]:
            rows_sales.append({
                'Неделя': wk, 'Площадка': plat,
                'выручка':   base['выручка']   * mult,
                'себест':    base['себест']    * mult,
                'комиссия':  base['комиссия']  * mult,
                'логистика': base['логистика'] * mult,
                'хранение':  base['хранение']  * mult,
            })
        rows_ads.append({'Неделя': wk, 'реклама': base_ads * mult})

    w_sales_f = pd.concat([w_sales, pd.DataFrame(rows_sales)], ignore_index=True)
    ads_w_f   = pd.concat([ads_w,   pd.DataFrame(rows_ads)],   ignore_index=True)
    return w_sales_f, ads_w_f


def build_cashflow(w_sales: pd.DataFrame, ads_w: pd.DataFrame) -> pd.DataFrame:
    """
    Шкала — недельная. Каждому факту приписываем дату денежного события
    (shift по лагу), затем агрегируем по неделе кассы.
    """
    # --- притоки от МП ---
    w_sales['нетто_поступление'] = (w_sales['выручка']
                                     - w_sales['комиссия']
                                     - w_sales['логистика']
                                     - w_sales['хранение'])
    def payout_date(r):
        lag = WB_PAYOUT_LAG_DAYS if r['Площадка']=='WB' else OZON_PAYOUT_LAG_DAYS
        return r['Неделя'] + pd.Timedelta(days=lag)
    w_sales['дата_денег'] = w_sales.apply(payout_date, axis=1)
    w_sales['неделя_кассы'] = w_sales['дата_денег'].dt.to_period('W-SUN').dt.start_time

    inflows = w_sales.groupby(['неделя_кассы','Площадка'])['нетто_поступление'].sum().unstack(fill_value=0)
    inflows.columns = [f'Поступление {c}' for c in inflows.columns]

    # --- оттоки ---
    # закупка: предоплата поставщику за N дней до недели продажи
    supply = w_sales.groupby('Неделя')['себест'].sum().reset_index()
    supply['дата_оплаты'] = supply['Неделя'] - pd.Timedelta(days=SUPPLIER_PREPAY_DAYS)
    supply['неделя_кассы'] = supply['дата_оплаты'].dt.to_period('W-SUN').dt.start_time
    supply['закупка'] = supply['себест'] * SUPPLIER_PREPAY_SHARE
    supply_out = supply.groupby('неделя_кассы')['закупка'].sum()

    # постоплата поставщику (остаток)
    supply_post = supply.copy()
    supply_post['дата_оплаты'] = supply_post['Неделя'] + pd.Timedelta(days=14)
    supply_post['неделя_кассы'] = supply_post['дата_оплаты'].dt.to_period('W-SUN').dt.start_time
    supply_post['постоплата'] = supply_post['себест'] * (1 - SUPPLIER_PREPAY_SHARE)
    supply_post_out = supply_post.groupby('неделя_кассы')['постоплата'].sum()

    # реклама: предоплата в неделю показов
    ads_w['неделя_кассы'] = (ads_w['Неделя'] - pd.Timedelta(days=ADS_PAY_LAG_DAYS)).dt.to_period('W-SUN').dt.start_time
    ads_out = ads_w.groupby('неделя_кассы')['реклама'].sum()

    # налог УСН — раз в квартал по факту выручки
    # АП: 25 апр (Q1), 25 июл (Q2), 25 окт (Q3), налог 31 мар след.года (Q4)
    rev_by_q = w_sales.groupby(w_sales['Неделя'].dt.to_period('Q'))['выручка'].sum()
    usn_rows = []
    for q, rev in rev_by_q.items():
        due = {
            1: pd.Timestamp(q.year, 4, 25),
            2: pd.Timestamp(q.year, 7, 25),
            3: pd.Timestamp(q.year, 10, 25),
            4: pd.Timestamp(q.year+1, 3, 31),
        }[q.quarter]
        usn_rows.append({'неделя_кассы': due.to_period('W-SUN').start_time,
                         'усн': rev * USN_RATE})
    usn_df = pd.DataFrame(usn_rows)
    usn_out = usn_df.groupby('неделя_кассы')['усн'].sum() if not usn_df.empty else pd.Series(dtype=float)

    # окно: от самого раннего события до самого позднего (включая предоплату поставщику)
    all_starts = [w_sales['неделя_кассы'].min()]
    all_ends   = [w_sales['неделя_кассы'].max()]
    for s in (supply_out, supply_post_out, ads_out, usn_out):
        if len(s):
            all_starts.append(s.index.min())
            all_ends.append(s.index.max())
    all_weeks = pd.date_range(min(all_starts), max(all_ends), freq='W-MON')
    opex_weekly = OPEX_MONTHLY / 4.33  # средненедельный opex
    opex = pd.Series(opex_weekly, index=all_weeks, name='opex')

    # --- собираем таблицу cashflow по неделям ---
    cf = pd.DataFrame({'Неделя': all_weeks})
    cf = cf.merge(inflows.reset_index().rename(columns={'неделя_кассы':'Неделя'}), on='Неделя', how='left')
    cf['Закупка (предоплата)']  = cf['Неделя'].map(supply_out).fillna(0)
    cf['Поставщик (постоплата)'] = cf['Неделя'].map(supply_post_out).fillna(0)
    cf['Реклама']               = cf['Неделя'].map(ads_out).fillna(0)
    cf['OPEX']                  = cf['Неделя'].map(opex).fillna(0)
    cf['УСН']                   = cf['Неделя'].map(usn_out).fillna(0)
    cf = cf.fillna(0)

    cf['Итого притоки'] = cf[[c for c in cf.columns if c.startswith('Поступление')]].sum(axis=1)
    cf['Итого оттоки']  = cf[['Закупка (предоплата)','Поставщик (постоплата)','Реклама','OPEX','УСН']].sum(axis=1)
    cf['Чистый денежный поток'] = cf['Итого притоки'] - cf['Итого оттоки']
    cf['Остаток на конец'] = START_CASH + cf['Чистый денежный поток'].cumsum()
    cf['Остаток на начало'] = cf['Остаток на конец'].shift(1).fillna(START_CASH)
    cf['Разрыв?'] = np.where(cf['Остаток на конец'] < 0, 'ДА', '')
    return cf.round(0)


def write_excel(cf: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cashflow'

    # порядок колонок
    cols = ['Неделя','Остаток на начало',
            'Поступление WB','Поступление Ozon','Итого притоки',
            'Закупка (предоплата)','Поставщик (постоплата)','Реклама','OPEX','УСН','Итого оттоки',
            'Чистый денежный поток','Остаток на конец','Разрыв?']
    cols = [c for c in cols if c in cf.columns]

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    neg_fill    = PatternFill('solid', fgColor='F8CBAD')
    warn_fill   = PatternFill('solid', fgColor='FFF2CC')
    border      = Border(*[Side(style='thin', color='D9D9D9')]*4)

    # заголовок
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # данные
    for i, row in enumerate(cf[cols].itertuples(index=False), 2):
        for j, v in enumerate(row, 1):
            cell = ws.cell(i, j, v)
            cell.border = border
            colname = cols[j-1]
            if colname == 'Неделя':
                cell.number_format = 'yyyy-mm-dd'
            elif colname == 'Разрыв?':
                if v == 'ДА':
                    cell.fill = neg_fill
                    cell.font = Font(bold=True, color='C00000')
            else:
                cell.number_format = '#,##0 ₽;[Red]-#,##0 ₽'
                if colname == 'Остаток на конец' and v < 500_000:
                    cell.fill = warn_fill if v >= 0 else neg_fill

    # ширина колонок
    widths = [12, 15, 15, 15, 14, 15, 17, 12, 12, 12, 14, 16, 16, 10]
    for j, w in enumerate(widths[:len(cols)], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'B2'

    # график остатка
    chart = LineChart()
    chart.title = 'Денежный остаток на конец недели, ₽'
    chart.y_axis.title = 'Остаток, ₽'
    chart.x_axis.title = 'Неделя'
    chart.height = 10; chart.width = 22
    data_col = cols.index('Остаток на конец') + 1
    data_ref = Reference(ws, min_col=data_col, min_row=1, max_row=len(cf)+1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(cf)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, 'P2')

    # лист допущений
    ws2 = wb.create_sheet('Допущения')
    assumptions = [
        ('Стартовый остаток, ₽',           START_CASH),
        ('Лаг выплат WB, дней',             WB_PAYOUT_LAG_DAYS),
        ('Лаг выплат Ozon, дней',           OZON_PAYOUT_LAG_DAYS),
        ('Предоплата поставщику, дней ДО продажи', SUPPLIER_PREPAY_DAYS),
        ('Доля предоплаты поставщику',      SUPPLIER_PREPAY_SHARE),
        ('Хранение МП, % от выручки',       STORAGE_SHARE),
        ('OPEX, ₽/мес (ФОТ+прочее)',        OPEX_MONTHLY),
        ('Ставка УСН',                      USN_RATE),
        ('Недель прогноза вперёд',          FORECAST_WEEKS),
        ('Рост выручки неделя-к-неделе',    GROWTH_WOW),
    ]
    ws2.cell(1,1,'Параметр').font = Font(bold=True)
    ws2.cell(1,2,'Значение').font = Font(bold=True)
    for i,(k,v) in enumerate(assumptions, 2):
        ws2.cell(i,1,k); ws2.cell(i,2,v)
    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 18

    wb.save(out_path)


def print_summary(cf: pd.DataFrame):
    print('='*72)
    print('ИТОГИ CASHFLOW-МОДЕЛИ')
    print('='*72)
    print(f'Период модели: {cf["Неделя"].min():%Y-%m-%d} → {cf["Неделя"].max():%Y-%m-%d}')
    print(f'Стартовый остаток: {START_CASH:>15,.0f} ₽'.replace(',', ' '))
    print(f'Итого притоки:     {cf["Итого притоки"].sum():>15,.0f} ₽'.replace(',', ' '))
    print(f'Итого оттоки:      {cf["Итого оттоки"].sum():>15,.0f} ₽'.replace(',', ' '))
    print(f'Чистый поток:      {cf["Чистый денежный поток"].sum():>15,.0f} ₽'.replace(',', ' '))
    print(f'Конечный остаток:  {cf["Остаток на конец"].iloc[-1]:>15,.0f} ₽'.replace(',', ' '))
    print(f'Минимальный остаток: {cf["Остаток на конец"].min():>13,.0f} ₽ на неделе {cf.loc[cf["Остаток на конец"].idxmin(),"Неделя"]:%Y-%m-%d}'.replace(',', ' '))
    print()

    gaps = cf[cf['Остаток на конец'] < 0]
    if len(gaps):
        print(f'⚠️  КАССОВЫЕ РАЗРЫВЫ: {len(gaps)} недель(и)')
        print(gaps[['Неделя','Остаток на конец']].to_string(index=False))
    else:
        print('✅ Кассовых разрывов НЕТ при текущих допущениях')

    warn = cf[(cf['Остаток на конец'] >= 0) & (cf['Остаток на конец'] < 500_000)]
    if len(warn):
        print(f'\n🟡 Опасная зона (< 500k): {len(warn)} недель(и)')
        print(warn[['Неделя','Остаток на конец']].to_string(index=False))


def main():
    w_sales, ads_w = load_data()
    w_sales_f, ads_w_f = build_forecast(w_sales, ads_w, FORECAST_WEEKS)
    cf = build_cashflow(w_sales_f, ads_w_f)
    write_excel(cf, OUT)
    print_summary(cf)
    print(f'\n📊 Excel: {OUT}')


if __name__ == '__main__':
    main()
