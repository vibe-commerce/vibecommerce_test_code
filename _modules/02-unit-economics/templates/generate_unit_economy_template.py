"""Генератор пустого шаблона unit-economy-template.xlsx с формулами.

Запуск:
    uv run --with openpyxl _modules/02-unit-economics/templates/generate_unit_economy_template.py

Создаёт unit-economy-template.xlsx в той же папке.
Без реальных цифр — только структура и формулы для студента.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Unit Economy"

    bold = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="00BCD4", end_color="00BCD4", fill_type="solid")
    formula_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Header
    headers = [
        "SKU",
        "Название",
        "Цена продажи",
        "Себестоимость",
        "Комиссия МП %",
        "Логистика МП",
        "Эквайринг %",
        "Реклама ДРР %",
        "% возвратов",
        "Стоимость возврата",
        "Дней на складе",
        "Хранение/день",
        "Налог режим",
        "— Комиссия МП ₽",
        "— Логистика ₽",
        "— Эквайринг ₽",
        "— Реклама ₽",
        "— Возвраты ₽",
        "— Хранение ₽",
        "Прибыль до налога",
        "Налог ₽",
        "Чистая прибыль ₽",
        "Чистая маржа %",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Demo row (synthetic)
    # SKU | Name | Price | COGS | Comm% | Log | Acq% | Ads% | Ret% | RetCost | Days | Stor/d | Tax
    demo = [
        "DEMO-001",
        "Demo product",
        2000,
        800,
        0.17,
        45,
        0.015,
        0.10,
        0.10,
        80,
        15,
        0.5,
        "usn_6",
    ]
    for col, v in enumerate(demo, start=1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.border = border

    # Formulas (yellow background)
    formulas = {
        14: "=C2*E2",            # Commission ₽
        15: "=F2",               # Logistics ₽
        16: "=C2*G2",            # Acquiring ₽
        17: "=C2*H2",            # Ads ₽
        18: "=J2*I2",            # Returns ₽
        19: "=L2*K2",            # Storage ₽
        20: "=C2-D2-N2-O2-P2-Q2-R2-S2",   # Gross profit (Price - COGS - all MP costs)
        21: '=IF(M2="usn_6",C2*0.06,IF(M2="usn_15",MAX(T2*0.15,C2*0.01),IF(M2="osno",T2*0.13,0)))',
        22: "=T2-U2",            # Net profit
        23: "=V2/C2",            # Net margin %
    }
    for col, formula in formulas.items():
        cell = ws.cell(row=2, column=col, value=formula)
        cell.fill = formula_fill
        cell.border = border
        if col == 23:
            cell.number_format = "0.00%"

    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes_text = [
        "# Unit Economy Template",
        "",
        "Заполни строки 3+ своими SKU.",
        "Жёлтые ячейки (колонки N-W) — формулы, не редактируй.",
        "Белые ячейки (колонки A-M) — твои данные.",
        "",
        "## Колонки M — налог режим",
        "usn_6   — УСН 6% от выручки",
        "usn_15  — УСН 15% от прибыли (мин. 1% от выручки)",
        "osno    — ОСНО, упрощённо НДФЛ 13% от прибыли",
        "",
        "## Тарифы МП (E - комиссия в долях, F - руб/единица)",
        "WB:    E~0.14-0.19   F~30-50",
        "Ozon:  E~0.07-0.25   F~35-70",
        "YM:    E~0.00-0.19   F~40-80",
        "",
        "Подробности → _knowledge/marketplaces/",
        "",
        "## Откуда смотреть формулы",
        "_modules/02-unit-economics/scripts/calc_marketplace_tax.py",
        "_modules/02-unit-economics/scripts/calc_unit_economics_tax.py",
        "",
        "📈 Полная версия в VIP-репо:",
        "- Multi-SKU портфельная оптимизация",
        "- Сценарии чувствительности к ДРР",
        "- Cohort-юнит-экономика (LTV/ARPU/CAC)",
        "- Импорт + НДС-агент",
    ]
    for i, line in enumerate(notes_text, start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 80

    output = __file__.replace("generate_unit_economy_template.py", "unit-economy-template.xlsx")
    wb.save(output)
    print(f"✅ Generated: {output}")


if __name__ == "__main__":
    main()
