"""Генератор шаблона P&L (Profit & Loss) для селлера.

Запуск:
    uv run --with openpyxl _modules/07-financial-reporting/templates/generate_pnl_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    bold = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="00BCD4", end_color="00BCD4", fill_type="solid")
    sum_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    margin_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Header
    headers = ["Статья", "Кв. 1", "Кв. 2", "Кв. 3", "Кв. 4", "Итого год"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # P&L structure
    rows = [
        # (label, is_subtotal, is_negative)
        ("Выручка WB", False, False),
        ("Выручка Ozon", False, False),
        ("Выручка Я.Маркет", False, False),
        ("Выручка DTC (свой сайт)", False, False),
        ("ВАЛОВАЯ ВЫРУЧКА", True, False),
        ("", False, False),
        ("Комиссия WB", False, True),
        ("Комиссия Ozon", False, True),
        ("Комиссия Я.Маркет", False, True),
        ("Логистика МП", False, True),
        ("Эквайринг", False, True),
        ("ЧИСТАЯ ВЫРУЧКА", True, False),
        ("", False, False),
        ("Себестоимость (COGS)", False, True),
        ("ВАЛОВАЯ ПРИБЫЛЬ", True, False),
        ("", False, False),
        ("Реклама WB Promo", False, True),
        ("Реклама Ozon Performance", False, True),
        ("Я.Директ", False, True),
        ("Прочая реклама", False, True),
        ("ФОТ + взносы", False, True),
        ("Прочие операционные расходы", False, True),
        ("ОПЕРАЦИОННАЯ ПРИБЫЛЬ (EBITDA)", True, False),
        ("", False, False),
        ("Налог УСН/ОСНО", False, True),
        ("ЧИСТАЯ ПРИБЫЛЬ", True, False),
        ("", False, False),
        ("Чистая маржа %", True, False),
    ]

    row = 2
    revenue_row = 6  # ВАЛОВАЯ ВЫРУЧКА
    net_revenue_row = 13  # ЧИСТАЯ ВЫРУЧКА
    gross_profit_row = 16  # ВАЛОВАЯ ПРИБЫЛЬ
    ebitda_row = 24  # EBITDA
    net_profit_row = 27  # ЧИСТАЯ ПРИБЫЛЬ
    net_margin_row = 29  # Чистая маржа %

    for label, is_sub, is_neg in rows:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold if is_sub else Font(size=10)
        if is_sub:
            cell.fill = sum_fill
        cell.border = border

        for col in range(2, 7):
            c = ws.cell(row=row, column=col)
            c.border = border

            # Year column = SUM of 4 quarters
            if col == 6 and label:
                if is_sub or not is_neg:
                    c.value = f"=SUM(B{row}:E{row})"
                else:
                    c.value = f"=SUM(B{row}:E{row})"

            if is_sub:
                c.fill = sum_fill
                c.font = bold

    # Subtotal formulas (sum of relevant rows in each quarter)
    for col_letter in ["B", "C", "D", "E"]:
        # ВАЛОВАЯ ВЫРУЧКА = sum WB + Ozon + ЯМ + DTC (rows 2-5)
        ws[f"{col_letter}{revenue_row}"] = f"=SUM({col_letter}2:{col_letter}5)"
        # ЧИСТАЯ ВЫРУЧКА = ВАЛ.ВЫРУЧКА - commission/logistics/acquiring (rows 8-12 negative)
        ws[f"{col_letter}{net_revenue_row}"] = f"={col_letter}{revenue_row}-SUM({col_letter}8:{col_letter}12)"
        # ВАЛОВАЯ ПРИБЫЛЬ = ЧИСТАЯ ВЫРУЧКА - COGS
        ws[f"{col_letter}{gross_profit_row}"] = f"={col_letter}{net_revenue_row}-{col_letter}{gross_profit_row - 1}"
        # EBITDA = ВАЛ.ПРИБЫЛЬ - реклама - ФОТ - прочее
        ws[f"{col_letter}{ebitda_row}"] = f"={col_letter}{gross_profit_row}-SUM({col_letter}18:{col_letter}23)"
        # ЧИСТАЯ ПРИБЫЛЬ = EBITDA - налог
        ws[f"{col_letter}{net_profit_row}"] = f"={col_letter}{ebitda_row}-{col_letter}{net_profit_row - 1}"
        # Маржа % = ЧИСТАЯ ПРИБЫЛЬ / ВАЛОВАЯ ВЫРУЧКА
        ws[f"{col_letter}{net_margin_row}"] = f"=IFERROR({col_letter}{net_profit_row}/{col_letter}{revenue_row}, 0)"
        ws[f"{col_letter}{net_margin_row}"].number_format = "0.00%"
        ws[f"{col_letter}{net_margin_row}"].fill = margin_fill

    # Year totals
    for r in [revenue_row, net_revenue_row, gross_profit_row, ebitda_row, net_profit_row]:
        ws[f"F{r}"] = f"=SUM(B{r}:E{r})"
        ws[f"F{r}"].fill = sum_fill
        ws[f"F{r}"].font = bold
    ws[f"F{net_margin_row}"] = f"=IFERROR(F{net_profit_row}/F{revenue_row}, 0)"
    ws[f"F{net_margin_row}"].number_format = "0.00%"
    ws[f"F{net_margin_row}"].fill = margin_fill
    ws[f"F{net_margin_row}"].font = bold

    # Column widths
    ws.column_dimensions["A"].width = 35
    for c in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 14

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes_text = [
        "# P&L Template (Profit & Loss)",
        "",
        "Заполни числа в колонках B-E (Кв. 1-4) по каждой статье.",
        "Жёлтые ячейки (итоги, чистая прибыль) пересчитываются автоматически.",
        "Зелёная ячейка (Чистая маржа %) — главный KPI селлера.",
        "",
        "## Логика P&L",
        "Валовая выручка (revenue по каналам)",
        "- Комиссии МП, логистика, эквайринг",
        "= Чистая выручка",
        "- Себестоимость (COGS)",
        "= Валовая прибыль",
        "- Реклама + ФОТ + прочее",
        "= Операционная прибыль (EBITDA)",
        "- Налог УСН/ОСНО",
        "= Чистая прибыль",
        "",
        "## Связь с другими шаблонами",
        "- ДДС (cashflow-template.xlsx) — реальные деньги (учитывает дебиторку/кредиторку)",
        "- Баланс (balance-template.xlsx) — снимок активов и обязательств",
        "",
        "📈 Полная версия в VIP-репо:",
        "- Дашборд с автоматическим расчётом",
        "- Сценарии (оптимистический / реалистический / пессимистический)",
        "- Сравнение с прошлым годом",
    ]
    for i, line in enumerate(notes_text, start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 80

    output = __file__.replace("generate_pnl_template.py", "profit-loss-template.xlsx")
    wb.save(output)
    print(f"✅ Generated: {output}")


if __name__ == "__main__":
    main()
