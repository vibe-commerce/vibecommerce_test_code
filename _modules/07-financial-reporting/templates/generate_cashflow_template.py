"""Генератор шаблона ДДС (Движение Денежных Средств) для селлера.

Запуск:
    uv run --with openpyxl _modules/07-financial-reporting/templates/generate_cashflow_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cashflow"

    bold = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="00BCD4", end_color="00BCD4", fill_type="solid")
    sum_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Статья", "Янв", "Фев", "Мар", "Апр", "Май", "Июн",
               "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек", "Итого"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    rows = [
        # (label, is_subtotal)
        ("ПОСТУПЛЕНИЯ", True),
        ("Продажи WB (реальные деньги)", False),
        ("Продажи Ozon (реальные деньги)", False),
        ("Продажи ЯМ (реальные деньги)", False),
        ("Продажи DTC", False),
        ("Кредиты / займы", False),
        ("Прочие поступления", False),
        ("ИТОГО ПОСТУПЛЕНИЯ", True),
        ("", False),
        ("ВЫБЫТИЯ", True),
        ("Закупка товара (COGS)", False),
        ("Реклама", False),
        ("ФОТ + взносы", False),
        ("Налоги", False),
        ("Аренда (офис/склад)", False),
        ("Логистика (доп.)", False),
        ("Возврат кредитов", False),
        ("Прочие выбытия", False),
        ("ИТОГО ВЫБЫТИЯ", True),
        ("", False),
        ("ЧИСТЫЙ CASHFLOW (приток-отток)", True),
        ("", False),
        ("Остаток на начало месяца", False),
        ("Остаток на конец месяца", True),
    ]

    income_start_row = 3
    income_total_row = 9
    expense_start_row = 12
    expense_total_row = 20
    cashflow_row = 22
    open_balance_row = 24
    close_balance_row = 25

    row = 2
    for label, is_sub in rows:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold if is_sub else Font(size=10)
        if is_sub:
            cell.fill = sum_fill
        cell.border = border
        for col in range(2, 15):
            c = ws.cell(row=row, column=col)
            c.border = border
            if is_sub:
                c.fill = sum_fill
                c.font = bold
        row += 1

    # Formulas
    for col in range(2, 14):  # cols B-M (12 months)
        col_letter = ws.cell(row=1, column=col).column_letter
        # ИТОГО ПОСТУПЛЕНИЯ = sum income rows
        ws[f"{col_letter}{income_total_row}"] = f"=SUM({col_letter}{income_start_row}:{col_letter}{income_total_row - 1})"
        # ИТОГО ВЫБЫТИЯ = sum expense rows
        ws[f"{col_letter}{expense_total_row}"] = f"=SUM({col_letter}{expense_start_row}:{col_letter}{expense_total_row - 1})"
        # ЧИСТЫЙ CASHFLOW = поступления - выбытия
        ws[f"{col_letter}{cashflow_row}"] = f"={col_letter}{income_total_row}-{col_letter}{expense_total_row}"
        # Остаток на конец = остаток на начало + чистый cashflow
        ws[f"{col_letter}{close_balance_row}"] = f"={col_letter}{open_balance_row}+{col_letter}{cashflow_row}"
        # Остаток на начало следующего месяца = остаток на конец предыдущего
        if col < 13:
            next_col_letter = ws.cell(row=1, column=col + 1).column_letter
            ws[f"{next_col_letter}{open_balance_row}"] = f"={col_letter}{close_balance_row}"

    # Year totals (column N)
    for r in [income_total_row, expense_total_row, cashflow_row]:
        ws.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})").font = bold
        ws.cell(row=r, column=14).fill = sum_fill

    ws.column_dimensions["A"].width = 35
    for c_idx in range(2, 15):
        ws.cell(row=1, column=c_idx).column_letter
        col = ws.cell(row=1, column=c_idx).column_letter
        ws.column_dimensions[col].width = 12

    # Notes
    notes = wb.create_sheet("Notes")
    notes_text = [
        "# ДДС Template (Cashflow)",
        "",
        "Заполни числа по месяцам (колонки B-M).",
        "Жёлтые ячейки (итоги + остатки) пересчитываются автоматически.",
        "Главный KPI: \"Остаток на конец месяца\" не должен уходить в минус.",
        "",
        "## Отличие от P&L",
        "P&L — про прибыль/убыток (даже если деньги ещё не получены)",
        "ДДС — про реальное движение денег",
        "",
        "Пример: продажа на WB сделана в Январе, деньги пришли в Феврале.",
        "В P&L: выручка в Январе",
        "В ДДС: поступление в Феврале",
        "",
        "## Структура ДДС",
        "Поступления:",
        "  Операционная деятельность (продажи)",
        "  Финансовая деятельность (кредиты)",
        "  Инвестиционная деятельность (продажа активов)",
        "",
        "Выбытия:",
        "  Операционная (закупка, реклама, ФОТ, налоги)",
        "  Финансовая (возврат кредитов, дивиденды)",
        "  Инвестиционная (покупка оборудования)",
        "",
        "📈 Полная версия в VIP-репо:",
        "- Прогноз cashflow на 6-12 месяцев вперёд",
        "- Сценарии нехватки cash + меры (кредитная линия, факторинг)",
    ]
    for i, line in enumerate(notes_text, start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 80

    output = __file__.replace("generate_cashflow_template.py", "cashflow-template.xlsx")
    wb.save(output)
    print(f"✅ Generated: {output}")


if __name__ == "__main__":
    main()
