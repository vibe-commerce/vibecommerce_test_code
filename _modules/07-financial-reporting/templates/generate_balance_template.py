"""Генератор шаблона Баланса (упрощённый) для селлера на МП.

Снимок активов и обязательств на конец каждого квартала. Без реальных цифр —
только структура и формулы (контрольная строка: Активы = Пассивы).

Запуск:
    uv run --with openpyxl _modules/07-financial-reporting/templates/generate_balance_template.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"

    bold = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="00BCD4", end_color="00BCD4", fill_type="solid")
    sum_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    check_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Статья", "Кон. Кв1", "Кон. Кв2", "Кон. Кв3", "Кон. Кв4"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    rows = [
        # (label, is_subtotal)
        ("АКТИВЫ", True),
        ("Денежные средства (р/с, касса)", False),
        ("Дебиторская задолженность (МП к выплате)", False),
        ("Товарные запасы на складах МП", False),
        ("Товарные запасы на своём складе", False),
        ("Авансы поставщикам", False),
        ("Прочие активы", False),
        ("ИТОГО АКТИВЫ", True),
        ("", False),
        ("ОБЯЗАТЕЛЬСТВА", True),
        ("Кредиторская задолженность (поставщикам)", False),
        ("Кредиты и займы", False),
        ("Налоги к уплате", False),
        ("Прочие обязательства", False),
        ("ИТОГО ОБЯЗАТЕЛЬСТВА", True),
        ("", False),
        ("Собственный капитал (вклады + нераспр. прибыль)", False),
        ("ИТОГО ПАССИВЫ (обязательства + капитал)", True),
        ("", False),
        ("Контроль: Активы − Пассивы (должно быть 0)", True),
    ]

    row = 2
    for label, is_sub in rows:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold if is_sub else Font(size=10)
        if is_sub:
            cell.fill = sum_fill
        cell.border = border
        for col in range(2, 6):
            c = ws.cell(row=row, column=col)
            c.border = border
            if is_sub:
                c.fill = sum_fill
                c.font = bold
        row += 1

    # Абсолютные индексы строк (start row=2)
    total_assets = 9    # ИТОГО АКТИВЫ
    total_liab = 16     # ИТОГО ОБЯЗАТЕЛЬСТВА
    equity = 18         # Собственный капитал
    total_passive = 19  # ИТОГО ПАССИВЫ
    check = 21          # Контроль

    for col in ["B", "C", "D", "E"]:
        ws[f"{col}{total_assets}"] = f"=SUM({col}3:{col}8)"
        ws[f"{col}{total_liab}"] = f"=SUM({col}12:{col}15)"
        ws[f"{col}{total_passive}"] = f"={col}{total_liab}+{col}{equity}"
        ws[f"{col}{check}"] = f"={col}{total_assets}-{col}{total_passive}"
        ws[f"{col}{check}"].fill = check_fill
        ws[f"{col}{check}"].font = bold

    ws.column_dimensions["A"].width = 45
    for c in ["B", "C", "D", "E"]:
        ws.column_dimensions[c].width = 14

    notes = wb.create_sheet("Notes")
    notes_text = [
        "# Balance Template (упрощённый баланс селлера)",
        "",
        "Снимок активов и обязательств на конец каждого квартала.",
        "Заполни числа в колонках B-E. Жёлтые итоги и зелёная контрольная строка — формулы.",
        "",
        "## Логика",
        "АКТИВЫ (что есть): деньги + дебиторка МП + товарные запасы + авансы",
        "ОБЯЗАТЕЛЬСТВА (что должен): кредиторка + кредиты + налоги",
        "КАПИТАЛ: вклады собственника + накопленная прибыль",
        "Баланс сходится: ИТОГО АКТИВЫ = ОБЯЗАТЕЛЬСТВА + КАПИТАЛ (контроль = 0)",
        "",
        "## Связь с другими шаблонами",
        "- P&L (profit-loss-template.xlsx) — прибыль за период",
        "- ДДС (cashflow-template.xlsx) — движение денег",
        "",
        "⚠️ Без реальных цифр — только структура и формулы.",
        "📈 Полная версия (дашборд, авто-сведение балансов) — в VIP-репо.",
    ]
    for i, line in enumerate(notes_text, start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 80

    output = __file__.replace("generate_balance_template.py", "balance-template.xlsx")
    wb.save(output)
    print(f"✅ Generated: {output}")


if __name__ == "__main__":
    main()
