"""
Анализ ценовой эластичности спроса по данным продаж.

Использование:
  uv run --with openpyxl --with matplotlib PRJ_PRICING/price_elasticity.py <путь_к_xlsx>

  Если путь не указан — используется PRJ_MARKETPLACE/sales_data_v1.0.xlsx

Формат входного файла (.xlsx):
  Лист "Каталог": SKU | Название | Категория | Себестоимость, ₽ | ...
  Лист "Продажи": Неделя | SKU | Название | Категория | Площадка | Продажи, шт | Цена, ₽ | ...

Что считает:
  - Эластичность спроса по цене (arc elasticity, метод средней точки)
  - Классификация: эластичный / единичная / неэластичный
  - Целевая цена для удвоения объёма (log-linear модель)

Выходные файлы (в reports/):
  01_elasticity_all_skus.png      — эластичность по всем SKU
  02_elasticity_by_category.png   — эластичность по категориям
  03_price_vs_qty_top12.png       — scatter цена vs продажи (ТОП-12)
  04_demand_curves_multi_price.png — кривые спроса (3+ ценовых уровня)
  05_price_vs_demand_change.png   — матрица ΔЦена% vs ΔСпрос%

Эластичность спроса по цене: E = (ΔQ/Q̄) / (ΔP/P̄)
  E < -1  → эластичный спрос (цена вниз → выручка вверх)
  E > -1  → неэластичный спрос (цена вниз → выручка вниз)
  E = -1  → единичная эластичность
"""

import sys
import os
import math
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from collections import defaultdict

# ── Конфигурация ──
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['font.size'] = 10
plt.rcParams['axes.titlesize'] = 12
plt.rcParams['figure.dpi'] = 150

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_DATA = os.path.join(REPO_ROOT, 'PRJ_MARKETPLACE', 'sales_data_v1.0.xlsx')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'reports')


def load_data(data_file):
    """Загрузить данные из Excel."""
    wb = openpyxl.load_workbook(data_file, data_only=True)
    ws = wb['Продажи']

    sku_price_qty = defaultdict(lambda: defaultdict(list))
    sku_platform_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for row in ws.iter_rows(min_row=2, values_only=True):
        week, sku, name, cat, platform, qty, price = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if qty is not None and price is not None and qty > 0:
            sku_price_qty[(sku, name, cat)][price].append(qty)
            sku_platform_data[(sku, name, cat)][platform][price].append(qty)

    # Себестоимость из каталога
    costs = {}
    if 'Каталог' in wb.sheetnames:
        ws_cat = wb['Каталог']
        for row in ws_cat.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[3] is not None:
                costs[row[0]] = row[3]

    wb.close()
    return sku_price_qty, sku_platform_data, costs


def calc_elasticity(sku_price_qty, costs):
    """Рассчитать эластичность для каждого SKU."""
    results = []

    for (sku, name, cat), price_data in sorted(sku_price_qty.items()):
        prices = sorted(price_data.keys())
        if len(prices) < 2:
            continue

        avg_qty_by_price = {p: sum(qs) / len(qs) for p, qs in price_data.items()}
        p_low, p_high = prices[0], prices[-1]
        q_low, q_high = avg_qty_by_price[p_low], avg_qty_by_price[p_high]

        p_avg = (p_low + p_high) / 2
        q_avg = (q_low + q_high) / 2

        if p_avg == 0 or q_avg == 0:
            continue

        elasticity = ((q_high - q_low) / q_avg) / ((p_high - p_low) / p_avg)

        total_qty = sum(sum(qs) for qs in price_data.values())
        total_rev = sum(p * sum(qs) for p, qs in price_data.items())
        avg_price = total_rev / total_qty if total_qty > 0 else 0

        # Log-linear эластичность и целевая цена для 2× объёма
        e_log = None
        p_target_2x = None
        if q_low > 0 and q_high > 0 and p_low != p_high:
            e_log = math.log(q_high / q_low) / math.log(p_high / p_low)
            a = math.log(q_low) - e_log * math.log(p_low)
            total_weeks = sum(len(qs) for qs in price_data.values())
            avg_qty_overall = total_qty / total_weeks
            q_target = 2 * avg_qty_overall
            try:
                ln_p_target = (math.log(q_target) - a) / e_log
                p_target_2x = math.exp(ln_p_target)
            except (ValueError, ZeroDivisionError):
                p_target_2x = None

        results.append({
            'sku': sku,
            'name': name,
            'category': cat,
            'p_low': p_low,
            'p_high': p_high,
            'q_at_low': q_low,
            'q_at_high': q_high,
            'elasticity': elasticity,
            'e_log': e_log,
            'total_qty': total_qty,
            'total_rev': total_rev,
            'avg_price': avg_price,
            'price_delta_pct': (p_high - p_low) / p_low * 100,
            'cost': costs.get(sku),
            'p_target_2x': p_target_2x,
        })

    results.sort(key=lambda x: x['elasticity'])
    return results


def print_table(results):
    """Вывести таблицу эластичности в консоль."""
    print("=" * 110)
    print(f"{'SKU':>4} {'Название':<38} {'Категория':<18} {'Цена ↓':>7} {'Цена ↑':>7} {'ΔP%':>5} {'Q↓':>5} {'Q↑':>5} {'E':>7} {'Тип':<12}")
    print("=" * 110)

    for r in results:
        e = r['elasticity']
        if e < -1:
            etype = "ЭЛАСТИЧН."
        elif e > -0.5:
            etype = "НЕЭЛАСТИЧН."
        else:
            etype = "~ЕДИНИЧН."

        print(f"{r['sku']:>4} {r['name']:<38} {r['category']:<18} {r['p_low']:>7.0f} {r['p_high']:>7.0f} "
              f"{r['price_delta_pct']:>4.1f}% {r['q_at_low']:>5.1f} {r['q_at_high']:>5.1f} {e:>7.2f} {etype:<12}")

    elastic = [r for r in results if r['elasticity'] < -1]
    inelastic = [r for r in results if r['elasticity'] > -0.5]
    unit_elastic = [r for r in results if -1 <= r['elasticity'] <= -0.5]

    print(f"\n{'=' * 60}")
    print(f"ИТОГИ")
    print(f"{'=' * 60}")
    print(f"Всего SKU: {len(results)}")
    print(f"  Эластичный (E < -1):     {len(elastic):>3} ({len(elastic)/len(results)*100:.0f}%)")
    print(f"  Единичная эластичность:   {len(unit_elastic):>3} ({len(unit_elastic)/len(results)*100:.0f}%)")
    print(f"  Неэластичный (E > -0.5): {len(inelastic):>3} ({len(inelastic)/len(results)*100:.0f}%)")


def print_target_prices(results, category=None):
    """Вывести целевые цены для удвоения объёма."""
    filtered = results
    if category:
        filtered = [r for r in results if r['category'] == category]

    print(f"\n{'=' * 100}")
    title = f"ЦЕЛЕВЫЕ ЦЕНЫ ДЛЯ 2× ОБЪЁМА"
    if category:
        title += f" — {category}"
    print(title)
    print(f"{'=' * 100}\n")

    for r in filtered:
        if r['p_target_2x'] is None:
            continue

        avg_qty = r['total_qty'] / (sum(1 for _ in []) or 1)
        total_weeks = 24  # приблизительно
        avg_qty_w = r['total_qty'] / total_weeks
        p_target = r['p_target_2x']
        drop_pct = (1 - p_target / r['avg_price']) * 100
        cost = r['cost'] or 0

        rev_current = avg_qty_w * r['avg_price']
        rev_target = (2 * avg_qty_w) * p_target

        profit_current = (r['avg_price'] - cost) * avg_qty_w
        profit_target = (p_target - cost) * (2 * avg_qty_w)

        margin_current = (r['avg_price'] - cost) / r['avg_price'] * 100 if r['avg_price'] > 0 else 0
        margin_target = (p_target - cost) / p_target * 100 if p_target > 0 else 0

        print(f"SKU{r['sku']}: {r['name']}")
        print(f"  Сейчас: {r['avg_price']:.0f} ₽  →  Цель: {p_target:.0f} ₽  (снижение {drop_pct:.1f}%)")
        if cost > 0:
            profit_change = (profit_target / profit_current - 1) * 100 if profit_current > 0 else 0
            status = "ПРИБЫЛЬ РАСТЁТ" if profit_target > profit_current else "ПРИБЫЛЬ ПАДАЕТ"
            print(f"  Маржа: {margin_current:.0f}% → {margin_target:.0f}%  |  Прибыль: {profit_change:+.0f}%  |  {status}")
        print()


def plot_all(results, sku_price_qty, sku_platform_data, output_dir):
    """Построить все графики."""
    os.makedirs(output_dir, exist_ok=True)

    # ── ГРАФИК 1: Барчарт эластичности ──
    fig, ax = plt.subplots(figsize=(16, max(8, len(results) * 0.3)))
    names_short = [f"SKU{r['sku']} {r['name'][:25]}" for r in results]
    elasticities = [r['elasticity'] for r in results]
    colors = ['#d32f2f' if e < -1 else '#388e3c' if e > -0.5 else '#f57c00' for e in elasticities]

    ax.barh(range(len(results)), elasticities, color=colors, height=0.7, edgecolor='white', linewidth=0.5)
    ax.set_yticks(range(len(results)))
    ax.set_yticklabels(names_short, fontsize=8)
    ax.axvline(x=-1, color='red', linestyle='--', alpha=0.5, label='E = -1 (граница эластичности)')
    ax.axvline(x=0, color='gray', linestyle='-', alpha=0.3)
    ax.set_xlabel('Эластичность спроса по цене (E)')
    ax.set_title('Ценовая эластичность спроса по всем товарам', fontsize=14, fontweight='bold')
    ax.legend(loc='lower right')
    ax.invert_yaxis()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '01_elasticity_all_skus.png'), bbox_inches='tight')
    plt.close()
    print(f"  01_elasticity_all_skus.png")

    # ── ГРАФИК 2: По категориям ──
    cat_elast = defaultdict(list)
    for r in results:
        cat_elast[r['category']].append(r['elasticity'])

    categories = sorted(cat_elast.keys(), key=lambda c: sum(cat_elast[c]) / len(cat_elast[c]))
    cat_means = [sum(cat_elast[c]) / len(cat_elast[c]) for c in categories]
    cat_mins = [min(cat_elast[c]) for c in categories]
    cat_maxs = [max(cat_elast[c]) for c in categories]

    fig, ax = plt.subplots(figsize=(12, max(4, len(categories) * 0.8)))
    colors_cat = ['#d32f2f' if m < -1 else '#388e3c' if m > -0.5 else '#f57c00' for m in cat_means]
    bars = ax.barh(categories, cat_means, color=colors_cat, height=0.6, edgecolor='white')
    for i, (cat, mn, mx) in enumerate(zip(categories, cat_mins, cat_maxs)):
        ax.plot([mn, mx], [i, i], color='gray', linewidth=2, alpha=0.5)
        ax.plot(mn, i, 'o', color='gray', markersize=4, alpha=0.5)
        ax.plot(mx, i, 'o', color='gray', markersize=4, alpha=0.5)
    ax.axvline(x=-1, color='red', linestyle='--', alpha=0.5, label='E = -1')
    ax.set_xlabel('Средняя эластичность (E)')
    ax.set_title('Эластичность спроса по категориям', fontsize=14, fontweight='bold')
    ax.legend()
    for bar, val in zip(bars, cat_means):
        ax.text(val - 0.1, bar.get_y() + bar.get_height() / 2, f'{val:.2f}',
                va='center', ha='right', fontsize=10, fontweight='bold', color='white')
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '02_elasticity_by_category.png'), bbox_inches='tight')
    plt.close()
    print(f"  02_elasticity_by_category.png")

    # ── ГРАФИК 3: Scatter ТОП-12 ──
    top12 = sorted(results, key=lambda x: x['total_qty'], reverse=True)[:12]
    fig, axes = plt.subplots(3, 4, figsize=(18, 12))
    axes = axes.flatten()
    for idx, r in enumerate(top12):
        ax = axes[idx]
        sku_key = (r['sku'], r['name'], r['category'])
        for platform, marker, color in [('WB', 'o', '#7b1fa2'), ('Ozon', 's', '#1565c0')]:
            if platform in sku_platform_data[sku_key]:
                pdata = sku_platform_data[sku_key][platform]
                xs, ys = [], []
                for price, qtys in pdata.items():
                    for q in qtys:
                        xs.append(price)
                        ys.append(q)
                ax.scatter(xs, ys, marker=marker, color=color, alpha=0.6, s=40, label=platform)
        e = r['elasticity']
        ecolor = '#d32f2f' if e < -1 else '#388e3c' if e > -0.5 else '#f57c00'
        ax.set_title(f"SKU{r['sku']}: {r['name'][:22]}\nE = {e:.2f}", fontsize=9, color=ecolor, fontweight='bold')
        ax.set_xlabel('Цена, ₽', fontsize=8)
        ax.set_ylabel('Продажи, шт', fontsize=8)
        ax.tick_params(labelsize=7)
        ax.legend(fontsize=7, loc='best')
        ax.grid(True, alpha=0.2)
    fig.suptitle('Цена vs Объём продаж — ТОП-12 товаров', fontsize=14, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '03_price_vs_qty_top12.png'), bbox_inches='tight')
    plt.close()
    print(f"  03_price_vs_qty_top12.png")

    # ── ГРАФИК 4: Кривые спроса ──
    multi_price_skus = [r for r in results if len(sku_price_qty[(r['sku'], r['name'], r['category'])]) >= 3]
    if multi_price_skus:
        n = min(8, len(multi_price_skus))
        rows = (n + 3) // 4
        fig, axes = plt.subplots(rows, 4, figsize=(18, rows * 4))
        if rows == 1:
            axes = [axes] if n == 1 else axes
        axes_flat = axes.flatten() if hasattr(axes, 'flatten') else [axes]
        for idx, r in enumerate(multi_price_skus[:n]):
            ax = axes_flat[idx]
            sku_key = (r['sku'], r['name'], r['category'])
            price_data = sku_price_qty[sku_key]
            prices_sorted = sorted(price_data.keys())
            avg_qtys = [sum(price_data[p]) / len(price_data[p]) for p in prices_sorted]
            revenues = [p * q for p, q in zip(prices_sorted, avg_qtys)]
            ax.plot(prices_sorted, avg_qtys, 'o-', color='#1565c0', linewidth=2, markersize=8)
            ax2 = ax.twinx()
            ax2.bar(prices_sorted, revenues, width=(prices_sorted[-1] - prices_sorted[0]) * 0.15,
                    alpha=0.3, color='#388e3c')
            ax2.set_ylabel('Выручка, ₽', fontsize=8, color='#388e3c')
            e = r['elasticity']
            ecolor = '#d32f2f' if e < -1 else '#388e3c' if e > -0.5 else '#f57c00'
            ax.set_title(f"SKU{r['sku']}: {r['name'][:22]}\nE = {e:.2f}", fontsize=9, color=ecolor, fontweight='bold')
            ax.set_xlabel('Цена, ₽', fontsize=8)
            ax.set_ylabel('Ср. продажи, шт/нед', fontsize=8, color='#1565c0')
            ax.grid(True, alpha=0.2)
        for idx in range(n, len(axes_flat)):
            axes_flat[idx].set_visible(False)
        fig.suptitle('Кривые спроса — товары с 3+ ценовыми уровнями', fontsize=14, fontweight='bold', y=1.02)
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, '04_demand_curves_multi_price.png'), bbox_inches='tight')
        plt.close()
        print(f"  04_demand_curves_multi_price.png")

    # ── ГРАФИК 5: Матрица ──
    fig, ax = plt.subplots(figsize=(12, 8))
    for r in results:
        dp = r['price_delta_pct']
        dq = ((r['q_at_high'] - r['q_at_low']) / r['q_at_low']) * 100 if r['q_at_low'] > 0 else 0
        size = max(20, min(200, r['total_qty'] / 5))
        e = r['elasticity']
        color = '#d32f2f' if e < -1 else '#388e3c' if e > -0.5 else '#f57c00'
        ax.scatter(dp, dq, s=size, c=color, alpha=0.6, edgecolors='white', linewidth=0.5)
        if abs(e) > 3 or abs(dq) > 50:
            ax.annotate(f"SKU{r['sku']}", (dp, dq), fontsize=7, alpha=0.7,
                       xytext=(5, 5), textcoords='offset points')
    ax.axhline(y=0, color='gray', linestyle='-', alpha=0.3)
    ax.set_xlabel('Δ Цена, %', fontsize=11)
    ax.set_ylabel('Δ Объём продаж, %', fontsize=11)
    ax.set_title('Чувствительность спроса: изменение цены vs изменение объёма', fontsize=14, fontweight='bold')
    legend_elements = [
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#d32f2f', markersize=10, label='Эластичный (E < -1)'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#f57c00', markersize=10, label='Ед. эластичность'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#388e3c', markersize=10, label='Неэластичный (E > -0.5)'),
    ]
    ax.legend(handles=legend_elements, loc='upper right')
    ax.grid(True, alpha=0.2)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '05_price_vs_demand_change.png'), bbox_inches='tight')
    plt.close()
    print(f"  05_price_vs_demand_change.png")


def main():
    data_file = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_DATA

    if not os.path.exists(data_file):
        print(f"Файл не найден: {data_file}")
        sys.exit(1)

    print(f"Источник данных: {data_file}")
    print(f"Результаты: {OUTPUT_DIR}/\n")

    sku_price_qty, sku_platform_data, costs = load_data(data_file)
    results = calc_elasticity(sku_price_qty, costs)

    if not results:
        print("Нет данных с вариацией цен. Для анализа нужны продажи одного SKU по разным ценам.")
        sys.exit(1)

    print_table(results)
    print_target_prices(results)

    print(f"\nСтрою графики...")
    plot_all(results, sku_price_qty, sku_platform_data, OUTPUT_DIR)
    print(f"\nГотово! Графики в {OUTPUT_DIR}/")


if __name__ == '__main__':
    main()
