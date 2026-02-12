"""
Генератор тестового датасета продаж товаров для сна.
50 SKU × 2 площадки (WB + Ozon) × 12 недель (ноя 2025 – фев 2026).
Реалистичные паттерны: сезонность, Black Friday, НГ-пик, пост-НГ спад,
out-of-stock, ценовые акции, тренды.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import math

random.seed(42)

# ============================================================
# 1. PRODUCT CATALOG
# ============================================================
# (id, name, category, cost, price_wb, price_ozon, weight_kg, base_vol_wb, base_vol_ozon)
products = [
    # --- Подушки (10 SKU) ---
    (1,  "Подушка бамбук 50×70",            "Подушки",  350,  990,  1050, 0.8,  45, 25),
    (2,  "Подушка бамбук 70×70",            "Подушки",  400,  1190, 1250, 1.0,  35, 20),
    (3,  "Подушка лебяжий пух 50×70",       "Подушки",  300,  890,  940,  0.7,  55, 30),
    (4,  "Подушка лебяжий пух 70×70",       "Подушки",  350,  1090, 1140, 0.9,  40, 22),
    (5,  "Подушка холлофайбер 50×70",       "Подушки",  200,  590,  620,  0.6,  70, 40),
    (6,  "Подушка холлофайбер 70×70",       "Подушки",  250,  690,  730,  0.8,  50, 28),
    (7,  "Подушка ортопед. Memory Foam",    "Подушки",  800,  2490, 2590, 1.2,  20, 15),
    (8,  "Подушка ортопед. с гелем",        "Подушки",  900,  2990, 3090, 1.3,  12, 10),
    (9,  "Подушка для беременных U-обр.",   "Подушки",  600,  1990, 2090, 2.5,  15, 12),
    (10, "Подушка-валик для шеи",           "Подушки",  250,  790,  840,  0.4,  30, 18),
    # --- Одеяла (8 SKU) ---
    (11, "Одеяло бамбук 1.5-сп",           "Одеяла",   500,  1490, 1550, 1.8,  35, 20),
    (12, "Одеяло бамбук 2-сп",             "Одеяла",   650,  1890, 1990, 2.2,  25, 15),
    (13, "Одеяло бамбук Евро",             "Одеяла",   750,  2190, 2290, 2.5,  20, 12),
    (14, "Одеяло лебяжий пух 1.5-сп",      "Одеяла",   400,  1290, 1350, 1.5,  40, 22),
    (15, "Одеяло лебяжий пух 2-сп",        "Одеяла",   550,  1590, 1690, 1.9,  28, 16),
    (16, "Одеяло лебяжий пух Евро",        "Одеяла",   650,  1890, 1990, 2.3,  18, 10),
    (17, "Одеяло шерстяное 1.5-сп",        "Одеяла",   700,  2490, 2590, 2.0,  12,  8),
    (18, "Одеяло облегч. летнее 1.5-сп",   "Одеяла",   350,  990,  1050, 1.0,   5,  3),
    # --- Постельное бельё (10 SKU) ---
    (19, "КПБ сатин 1.5-сп",              "Постельное бельё", 600, 1790, 1890, 1.5, 30, 18),
    (20, "КПБ сатин 2-сп",                "Постельное бельё", 700, 2190, 2290, 1.8, 22, 14),
    (21, "КПБ сатин Евро",                "Постельное бельё", 800, 2490, 2590, 2.0, 18, 11),
    (22, "КПБ поплин 1.5-сп",             "Постельное бельё", 400, 1190, 1250, 1.3, 40, 25),
    (23, "КПБ поплин 2-сп",               "Постельное бельё", 500, 1490, 1550, 1.6, 30, 18),
    (24, "КПБ поплин Евро",               "Постельное бельё", 600, 1690, 1790, 1.8, 22, 14),
    (25, "КПБ бязь 1.5-сп",               "Постельное бельё", 300,  890,  940, 1.2, 55, 35),
    (26, "КПБ бязь 2-сп",                 "Постельное бельё", 380, 1090, 1150, 1.4, 40, 25),
    (27, "Простыня на рез. сатин 160×200", "Постельное бельё", 300,  890,  940, 0.5, 35, 20),
    (28, "Простыня на рез. поплин 160×200","Постельное бельё", 200,  590,  640, 0.4, 50, 30),
    # --- Наматрасники (6 SKU) ---
    (29, "Наматрасник стёганый 160×200",      "Наматрасники", 400, 1290, 1350, 1.5, 20, 12),
    (30, "Наматрасник стёганый 180×200",      "Наматрасники", 450, 1490, 1550, 1.7, 15, 10),
    (31, "Наматрасник водонепрон. 160×200",   "Наматрасники", 500, 1590, 1690, 1.6, 18, 14),
    (32, "Наматрасник водонепрон. 180×200",   "Наматрасники", 550, 1790, 1890, 1.8, 12,  9),
    (33, "Наматрасник с бортами 160×200",     "Наматрасники", 350, 1090, 1150, 1.4, 15,  8),
    (34, "Наматрасник с бортами 180×200",     "Наматрасники", 400, 1290, 1350, 1.6, 10,  6),
    # --- Пледы (5 SKU) ---
    (35, "Плед микрофибра 150×200",   "Пледы", 300,  990, 1050, 1.0, 40, 22),
    (36, "Плед микрофибра 200×220",   "Пледы", 400, 1390, 1450, 1.4, 25, 15),
    (37, "Плед шерпа 150×200",        "Пледы", 450, 1590, 1690, 1.3, 30, 18),
    (38, "Плед велсофт 150×200",      "Пледы", 250,  790,  840, 0.9, 35, 20),
    (39, "Плед с рукавами",           "Пледы", 350, 1190, 1250, 1.2, 20, 15),
    # --- Аксессуары для сна (6 SKU) ---
    (40, "Маска для сна шёлковая",     "Аксессуары", 150,  590,  640, 0.05, 60, 35),
    (41, "Маска для сна 3D",           "Аксессуары", 100,  390,  420, 0.04, 80, 50),
    (42, "Маска для сна с охлаждением","Аксессуары", 200,  690,  740, 0.08, 30, 18),
    (43, "Беруши силиконовые (набор)", "Аксессуары",  80,  290,  320, 0.02,100, 60),
    (44, "Спрей для подушки лаванда",  "Аксессуары", 120,  490,  540, 0.15, 45, 28),
    (45, "Ночник-проектор звёзд",      "Аксессуары", 350, 1290, 1350, 0.30, 25, 20),
    # --- Детский сон (5 SKU) ---
    (46, "Подушка детская 40×60",  "Детский сон", 200,  690,  740, 0.4, 25, 15),
    (47, "Одеяло детское 110×140", "Детский сон", 300,  990, 1050, 1.0, 20, 12),
    (48, "КПБ детский поплин",     "Детский сон", 350, 1090, 1150, 0.8, 18, 12),
    (49, "Бортики в кроватку",     "Детский сон", 400, 1390, 1450, 1.5, 12,  8),
    (50, "Ночник детский LED",     "Детский сон", 250,  890,  940, 0.2, 20, 15),
]

# ============================================================
# 2. WEEKS (12 full weeks ≈ 3 months)
# ============================================================
weeks = [
    datetime(2025, 11, 17),  # W1
    datetime(2025, 11, 24),  # W2 — Black Friday
    datetime(2025, 12,  1),  # W3
    datetime(2025, 12,  8),  # W4
    datetime(2025, 12, 15),  # W5
    datetime(2025, 12, 22),  # W6 — НГ-пик
    datetime(2025, 12, 29),  # W7 — НГ-каникулы
    datetime(2026,  1,  5),  # W8 — пост-НГ спад
    datetime(2026,  1, 12),  # W9
    datetime(2026,  1, 19),  # W10
    datetime(2026,  1, 26),  # W11
    datetime(2026,  2,  2),  # W12
]

# ============================================================
# 3. SEASONAL MULTIPLIERS
# ============================================================
# Overall demand multiplier by week
season_overall = [0.90, 1.25, 1.00, 1.05, 1.15, 1.45, 1.15, 0.55, 0.65, 0.78, 0.85, 0.90]

# Category-specific adjustments (multiply on top of overall)
cat_season = {
    "Подушки":          [1.00, 1.10, 1.00, 1.00, 1.05, 1.15, 1.00, 0.85, 0.90, 1.00, 1.00, 1.00],
    "Одеяла":           [1.10, 1.10, 1.10, 1.15, 1.20, 1.30, 1.10, 0.70, 0.80, 0.90, 0.95, 1.00],
    "Постельное бельё": [1.00, 1.05, 1.00, 1.00, 1.00, 1.15, 1.00, 0.90, 0.95, 1.00, 1.00, 1.00],
    "Наматрасники":     [1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 1.00, 0.90, 0.95, 1.00, 1.00, 1.00],
    "Пледы":            [1.10, 1.15, 1.10, 1.15, 1.20, 1.40, 1.10, 0.60, 0.70, 0.80, 0.85, 0.90],
    "Аксессуары":       [0.90, 1.00, 1.00, 1.00, 1.10, 1.50, 1.15, 0.70, 0.80, 0.90, 0.95, 1.00],
    "Детский сон":      [1.00, 1.00, 1.00, 1.00, 1.05, 1.30, 1.05, 0.80, 0.85, 0.90, 0.95, 1.00],
}

# ============================================================
# 4. COMMISSION & LOGISTICS
# ============================================================
commission_rates = {
    "WB":  {"Подушки": 0.15, "Одеяла": 0.15, "Постельное бельё": 0.15,
            "Наматрасники": 0.15, "Пледы": 0.15, "Аксессуары": 0.18, "Детский сон": 0.15},
    "Ozon":{"Подушки": 0.12, "Одеяла": 0.12, "Постельное бельё": 0.12,
            "Наматрасники": 0.12, "Пледы": 0.12, "Аксессуары": 0.15, "Детский сон": 0.12},
}

# DRR (% от выручки на рекламу)
drr_base = {
    "Подушки": 0.08,  "Одеяла": 0.07,  "Постельное бельё": 0.09,
    "Наматрасники": 0.06, "Пледы": 0.07, "Аксессуары": 0.10, "Детский сон": 0.07,
}

# DRR multiplier by week (more ad spend around BF/NY)
drr_week_mult = [1.0, 1.3, 1.0, 1.0, 1.1, 1.2, 0.9, 0.5, 0.7, 0.8, 0.9, 1.0]


def logistics_per_unit(weight_kg, platform):
    """Стоимость логистики за единицу (руб)."""
    if platform == "WB":
        return round(55 + weight_kg * 12)
    else:
        return round(45 + weight_kg * 14)


def storage_weekly(weight_kg, qty, platform):
    """Хранение за неделю (руб), зависит от объёма и количества."""
    per_unit = 2.5 + weight_kg * 1.8 if platform == "WB" else 2.0 + weight_kg * 1.5
    return round(per_unit * qty)


# ============================================================
# 5. SPECIAL PATTERNS
# ============================================================
# Out-of-stock: SKU 5 (бестселлер) — WB, недели 9-10
oos_rules = {(5, "WB", 8): 0, (5, "WB", 9): 0}

# Price discount: одеяла WB — скидка 10% в январе (недели 8-12)
discount_rules = {}
for p in products:
    if p[3] and p[0] in range(11, 19):  # одеяла, SKU 11-18
        for wi in range(7, 12):
            discount_rules[(p[0], "WB", wi)] = 0.90  # 10% off

# Growing trend: SKU 7 (Memory Foam) — растёт +5% в неделю
trend_rules = {7: 1.05}

# Promo boost: SKU 37 (Плед шерпа) — усиленная реклама в недели 5-6
promo_boost = {(37, 4): 1.8, (37, 5): 2.0}  # sales and DRR boost

# Cannibalization: SKU 1 and 3 — корреляция (когда один растёт, другой падает)
# Implemented via inverse random noise


# ============================================================
# 6. GENERATE DATA
# ============================================================
def gen_qty(base, overall_m, cat_m, noise=0.20):
    val = base * overall_m * cat_m
    val *= random.uniform(1 - noise, 1 + noise)
    return max(0, round(val))


sales_rows = []  # list of dicts

for wi, week_date in enumerate(weeks):
    for p in products:
        pid, name, cat, cost, price_wb, price_ozon, weight, base_wb, base_ozon = p

        for platform in ["WB", "Ozon"]:
            base_vol = base_wb if platform == "WB" else base_ozon
            price = price_wb if platform == "WB" else price_ozon

            # --- seasonal multiplier ---
            overall_m = season_overall[wi]
            cat_m = cat_season[cat][wi]

            # --- trend ---
            trend_m = 1.0
            if pid in trend_rules:
                trend_m = trend_rules[pid] ** wi  # compound growth

            # --- out-of-stock ---
            if (pid, platform, wi) in oos_rules:
                qty = 0
            else:
                qty = gen_qty(base_vol, overall_m, cat_m, noise=0.20)
                qty = round(qty * trend_m)

            # --- cannibalization (SKU 1 ↔ 3) ---
            # If SKU 1 WB, save noise; SKU 3 WB gets inverse
            # (simplified: alternate good/bad weeks)
            if pid == 1 and wi % 2 == 0:
                qty = round(qty * 1.12)
            elif pid == 1 and wi % 2 == 1:
                qty = round(qty * 0.88)
            elif pid == 3 and wi % 2 == 0:
                qty = round(qty * 0.88)
            elif pid == 3 and wi % 2 == 1:
                qty = round(qty * 1.12)

            # --- price discount ---
            if (pid, platform, wi) in discount_rules:
                price = round(price * discount_rules[(pid, platform, wi)])

            # --- promo boost ---
            promo_m = 1.0
            if (pid, wi) in promo_boost:
                promo_m = promo_boost[(pid, wi)]
                qty = round(qty * promo_m)

            # --- summer duvet sells very poorly in winter ---
            if pid == 18:  # летнее одеяло
                qty = max(0, round(qty * 0.3))

            # --- commission ---
            comm_rate = commission_rates[platform][cat]

            # --- logistics ---
            logi = logistics_per_unit(weight, platform) * qty

            # --- storage ---
            stor = storage_weekly(weight, qty, platform)

            # --- advertising ---
            revenue = price * qty
            drr = drr_base[cat] * drr_week_mult[wi]
            if (pid, wi) in promo_boost:
                drr *= 1.5  # extra spend during promo
            ad_spend = round(revenue * drr)

            # --- add random variance to logistics & storage ---
            if qty > 0:
                logi = round(logi * random.uniform(0.92, 1.08))
                stor = round(stor * random.uniform(0.90, 1.10))

            sales_rows.append({
                "week": week_date,
                "sku_id": pid,
                "name": name,
                "category": cat,
                "platform": platform,
                "qty": qty,
                "price": price,
                "cost": cost,
                "comm_rate": comm_rate,
                "logistics": logi,
                "storage": stor,
                "ad_spend": ad_spend,
            })


# ============================================================
# 7. WRITE EXCEL
# ============================================================
wb = openpyxl.Workbook()

# --- Styles ---
header_font = Font(bold=True, size=11)
header_fill = PatternFill("solid", fgColor="4472C4")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
num_fmt_rub = '#,##0" ₽"'
num_fmt_pct = '0.0%'
date_fmt = 'DD.MM.YYYY'

# ==================== SHEET 1: Каталог ====================
ws_cat = wb.active
ws_cat.title = "Каталог"

cat_headers = ["SKU", "Название", "Категория", "Себестоимость, ₽",
               "Цена WB, ₽", "Цена Ozon, ₽", "Вес, кг"]
cat_widths = [6, 35, 18, 16, 12, 13, 9]

for ci, (h, w) in enumerate(zip(cat_headers, cat_widths), 1):
    cell = ws_cat.cell(row=1, column=ci, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws_cat.column_dimensions[get_column_letter(ci)].width = w

for ri, p in enumerate(products, 2):
    pid, name, cat, cost, price_wb, price_ozon, weight = p[0], p[1], p[2], p[3], p[4], p[5], p[6]
    vals = [pid, name, cat, cost, price_wb, price_ozon, weight]
    for ci, v in enumerate(vals, 1):
        cell = ws_cat.cell(row=ri, column=ci, value=v)
        cell.border = thin_border
        if ci in (4, 5, 6):
            cell.number_format = num_fmt_rub

ws_cat.auto_filter.ref = f"A1:G{len(products)+1}"
ws_cat.freeze_panes = "A2"

# ==================== SHEET 2: Продажи ====================
ws = wb.create_sheet("Продажи")

sale_headers = [
    "Неделя",           # A
    "SKU",              # B
    "Название",         # C
    "Категория",        # D
    "Площадка",         # E
    "Продажи, шт",     # F
    "Цена, ₽",         # G
    "Выручка, ₽",      # H  (formula)
    "Себест. ед., ₽",  # I
    "Себест. итого, ₽",# J  (formula)
    "Комиссия МП, %",  # K
    "Комиссия МП, ₽",  # L  (formula)
    "Логистика, ₽",    # M
    "Хранение, ₽",     # N
    "Реклама, ₽",      # O
    "Прибыль, ₽",      # P  (formula)
    "Маржа, %",        # Q  (formula)
]
sale_widths = [12, 6, 35, 18, 10, 13, 10, 13, 13, 15, 13, 14, 12, 12, 12, 14, 10]

for ci, (h, w) in enumerate(zip(sale_headers, sale_widths), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(ci)].width = w

for ri, row in enumerate(sales_rows, 2):
    r = ri  # row number in Excel

    # A: Неделя
    ws.cell(row=r, column=1, value=row["week"]).number_format = date_fmt
    # B: SKU
    ws.cell(row=r, column=2, value=row["sku_id"])
    # C: Название
    ws.cell(row=r, column=3, value=row["name"])
    # D: Категория
    ws.cell(row=r, column=4, value=row["category"])
    # E: Площадка
    ws.cell(row=r, column=5, value=row["platform"])
    # F: Продажи, шт
    ws.cell(row=r, column=6, value=row["qty"])
    # G: Цена, ₽
    ws.cell(row=r, column=7, value=row["price"]).number_format = num_fmt_rub
    # H: Выручка (формула)
    ws.cell(row=r, column=8).value = f"=F{r}*G{r}"
    ws.cell(row=r, column=8).number_format = num_fmt_rub
    # I: Себестоимость ед.
    ws.cell(row=r, column=9, value=row["cost"]).number_format = num_fmt_rub
    # J: Себестоимость итого (формула)
    ws.cell(row=r, column=10).value = f"=F{r}*I{r}"
    ws.cell(row=r, column=10).number_format = num_fmt_rub
    # K: Комиссия %
    ws.cell(row=r, column=11, value=row["comm_rate"]).number_format = num_fmt_pct
    # L: Комиссия ₽ (формула)
    ws.cell(row=r, column=12).value = f"=H{r}*K{r}"
    ws.cell(row=r, column=12).number_format = num_fmt_rub
    # M: Логистика
    ws.cell(row=r, column=13, value=row["logistics"]).number_format = num_fmt_rub
    # N: Хранение
    ws.cell(row=r, column=14, value=row["storage"]).number_format = num_fmt_rub
    # O: Реклама
    ws.cell(row=r, column=15, value=row["ad_spend"]).number_format = num_fmt_rub
    # P: Прибыль (формула)
    ws.cell(row=r, column=16).value = f"=H{r}-J{r}-L{r}-M{r}-N{r}-O{r}"
    ws.cell(row=r, column=16).number_format = num_fmt_rub
    # Q: Маржа % (формула с защитой от деления на 0)
    ws.cell(row=r, column=17).value = f'=IF(H{r}=0,0,P{r}/H{r})'
    ws.cell(row=r, column=17).number_format = num_fmt_pct

    # borders
    for ci in range(1, 18):
        ws.cell(row=r, column=ci).border = thin_border

ws.auto_filter.ref = f"A1:Q{len(sales_rows)+1}"
ws.freeze_panes = "A2"

# ============================================================
# 8. SAVE
# ============================================================
output_path = "/Users/vadimbakanov/Documents/_CODE/vibecommerce_test_code/MARKETPLACE_PRJ/sales_data_v1.0.xlsx"
wb.save(output_path)
print(f"✓ Файл сохранён: {output_path}")
print(f"  Каталог: {len(products)} SKU")
print(f"  Продажи: {len(sales_rows)} строк ({len(weeks)} недель × {len(products)} SKU × 2 площадки)")
